# -*- coding: utf-8 -*-
"""Парсер «Программы максимум» (ПМ) — выгрузки ПЦ в разрезе решений.

Файл приходит из 1С («ПМ ПЦ_Решения к ДД.ММ.ГГГГ.xlsx»), лист `TDSheet`:
шапка не в первой строке, над ней сидит заголовок отчёта и объединённый блок «Замечания».

Что важно знать о цифрах файла:
  * «Прогноз + факт по ПЦ (итого)» = факт + прогноз; «Прогноз по ПЦ (итого)» = сумма квартальных
    колонок. Значит факт строки = итого − прогноз, отдельной колонки факта в этом файле нет.
  * Квартальные колонки названы по финансовому году («Прогноз 2FQ2026»), а финансовый год
    начинается 1 апреля. Поэтому 2FQ2026 — это КАЛЕНДАРНЫЙ июль–сентябрь 2026 (2026-Q3),
    3FQ2026 — октябрь–декабрь (2026-Q4). Проверено сверкой контрольных сумм файлов
    «Решения» и «ЦРБ» (в последнем те же колонки подписаны «3 квартал»/«4 квартал 2026 г.»).
  * Одна строка файла = связка клиент+договор+раздел ФП+решение. Доля (< 1) означает, что сумма
    договора уже поделена между решениями — складывать строки можно без опаски.

Парсер разворачивает строку файла в несколько записей: по одной на каждый квартал плюс
запись «Факт», если он есть. Так суммы ПМ ложатся на тот же ключ (quarter_label), по которому
живут строки финплана из СЦ, и обе таблицы становятся сравнимыми.
"""
import datetime
import re

import openpyxl

# Имена колонок → внутренние поля. Первое совпавшее имя выигрывает.
COLUMN_ALIASES = {
    "client_name": ("Клиент",),
    "contract_num": ("Договор",),
    "findoc": ("Финдок",),
    "presale": ("Пресейл",),
    "pc": ("ПЦ",),
    "section": ("Раздел ФП",),
    "solution": ("Решение",),
    "strategic_solution": ("Страт. решение", "Страт.решение", "Стратегическое решение"),
    "direction": ("Направление",),
    "share": ("Доля",),
    "total_amount": ("Прогноз + факт по ПЦ (итого)",),
    "forecast_amount": ("Прогноз по ПЦ (итого)",),
    "manager": ("Менеджер УП финдока", "Менеджер УП"),
    "kt": ("КТ",),
    "directorate": ("Дирекция",),
}

REQUIRED_FIELDS = ("client_name", "section", "total_amount")

QUARTER_HEADER_RE = re.compile(r"^Прогноз\s*(\d)FQ\s*(\d{4})$", re.IGNORECASE)

FACT_KIND = "Факт"
FORECAST_KIND = "Прогноз"


class PMParseError(Exception):
    """Файл не похож на выгрузку ПМ — шапку найти не удалось."""


def fiscal_to_calendar_quarter(fq_num, fy_year):
    """«2FQ2026» → «2026-Q3». Финансовый год стартует 1 апреля, поэтому ФКВ1 = календарный Q2."""
    calendar_q = fq_num + 1
    year = fy_year
    if calendar_q > 4:            # ФКВ4 (январь–март) уезжает в следующий календарный год
        calendar_q -= 4
        year += 1
    return f"{year}-Q{calendar_q}"


def _clean(value):
    if value is None:
        return ""
    return " ".join(str(value).split())


def _num(value):
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def find_header_row(ws, max_scan=15):
    """Возвращает номер строки шапки. Шапку опознаём по паре «Клиент» + «Раздел ФП»."""
    for r in range(1, min(max_scan, ws.max_row) + 1):
        names = {_clean(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)}
        if "Клиент" in names and "Раздел ФП" in names:
            return r
    raise PMParseError(
        "Не нашёл строку заголовков: в первых строках нет колонок «Клиент» и «Раздел ФП». "
        "Ожидается лист TDSheet из выгрузки «ПМ ПЦ_Решения»."
    )


def build_column_map(ws, header_row):
    """{внутреннее поле → номер колонки} + список квартальных колонок."""
    headers = {}
    for c in range(1, ws.max_column + 1):
        name = _clean(ws.cell(row=header_row, column=c).value)
        if name and name not in headers:
            headers[name] = c

    col_map = {}
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in headers:
                col_map[field] = headers[alias]
                break

    quarters = []
    for name, col in headers.items():
        m = QUARTER_HEADER_RE.match(name)
        if m:
            fq_num, fy_year = int(m.group(1)), int(m.group(2))
            quarters.append({
                "column": col,
                "header": name,
                "quarter_label": fiscal_to_calendar_quarter(fq_num, fy_year),
                "fq_label": f"{fq_num}FQ{fy_year}",
            })
    quarters.sort(key=lambda q: q["quarter_label"])
    return col_map, quarters


def detect_mismatch(ws):
    """Список проблем файла; пустой — файл можно грузить."""
    try:
        header_row = find_header_row(ws)
    except PMParseError as exc:
        return [str(exc)]

    col_map, quarters = build_column_map(ws, header_row)
    problems = []
    for field in REQUIRED_FIELDS:
        if field not in col_map:
            problems.append("нет обязательной колонки «%s»" % COLUMN_ALIASES[field][0])
    if not quarters:
        problems.append("не нашёл ни одной квартальной колонки вида «Прогноз 2FQ2026»")
    return problems


def parse_workbook(filepath, strategic_filter=None):
    """Разбирает файл ПМ.

    strategic_filter — необязательная подстрока стратегического решения (например «Реестры»).
    Строка попадает в выборку, если подстрока встречается в «Страт. решение» ИЛИ в «Решение»:
    у части строк продуктовое решение указано точнее, чем стратегическое (например
    продукт указан точнее решения), и наоборот —
    строки с пустым стратегическим решением находятся по названию продукта.

    Возвращает dict: rows (развёрнутые по кварталам записи), quarters, meta.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        for name in wb.sheetnames:        # TDSheet — основной лист, «Свод» это сводная таблица
            if name.lower() == "tdsheet":
                ws = wb[name]
                break

        header_row = find_header_row(ws)
        col_map, quarters = build_column_map(ws, header_row)
        if not quarters:
            raise PMParseError(
                "В файле нет квартальных колонок «Прогноз NFQГГГГ» — нечего сравнивать."
            )

        needle = (strategic_filter or "").strip().casefold()
        rows = []
        source_rows = 0
        skipped_empty = 0
        sheet_title = ws.title

        for raw in ws.iter_rows(min_row=header_row + 1, values_only=True):
            def cell(field):
                col = col_map.get(field)
                return raw[col - 1] if col and col <= len(raw) else None

            client = _clean(cell("client_name"))
            if not client:
                skipped_empty += 1
                continue

            solution = _clean(cell("solution"))
            strategic = _clean(cell("strategic_solution"))
            if needle and needle not in solution.casefold() and needle not in strategic.casefold():
                continue

            source_rows += 1
            total = _num(cell("total_amount"))
            forecast = _num(cell("forecast_amount"))
            fact = round(total - forecast, 2)

            base = {
                "client_name": client,
                "contract_num": _clean(cell("contract_num")),
                "findoc": _clean(cell("findoc")),
                "presale": _clean(cell("presale")),
                "pc": _clean(cell("pc")),
                "section": _clean(cell("section")),
                "solution": solution,
                "strategic_solution": strategic,
                "direction": _clean(cell("direction")),
                "share": _num(cell("share")),
                "manager": _clean(cell("manager")),
                "kt": _clean(cell("kt")),
                "directorate": _clean(cell("directorate")),
                "total_amount": total,
                "forecast_amount": forecast,
            }

            if fact:
                rows.append(dict(base, quarter_label="", kind=FACT_KIND, amount=fact))
            for q in quarters:
                col = q["column"]
                amount = _num(raw[col - 1] if col <= len(raw) else None)
                if amount:
                    rows.append(dict(base, quarter_label=q["quarter_label"],
                                     kind=FORECAST_KIND, amount=round(amount, 2)))
    finally:
        # read_only держит файл открытым до close(): на Windows иначе не удалить/переписать
        wb.close()

    return {
        "rows": rows,
        "quarters": quarters,
        "meta": {
            "header_row": header_row,
            "sheet": sheet_title,
            "source_rows": source_rows,
            "skipped_empty": skipped_empty,
            "parsed_at": datetime.datetime.now().isoformat(timespec="seconds"),
            "strategic_filter": strategic_filter or "",
        },
    }
