"""
Парсер файлов «Отчёт распределение ресурсов» (плоская структура).

Форматы:
  .xlsb  — читается напрямую через pyxlsb (pip3 install pyxlsb)
  .xlsx  — читается через openpyxl

Автоматически ищет строку заголовков по ключевым словам («РП», «Трудозатраты»,
«Сотрудник») — не зависит от фиксированного номера строки.
"""

import os
import subprocess
import tempfile
import openpyxl

# ─── Синонимы для поиска заголовков (нижний регистр, подстрока) ──────────────
_HEADER_KEYS = {
    'rp':           ('рп',),
    'hours':        ('трудозатрат',),
    'dept':         ('департамент',),
    'division':     ('управление',),
    'employee':     ('сотрудник', 'фио'),
    'project':      ('проект',),
    'task':         ('задача',),
    'client':       ('клиент',),
    'project_type': ('тип проекта',),
    'work_type':    ('тип работ', 'вид работ'),
}
_REQUIRED = {'rp', 'hours'}


def _strip(v):
    return str(v).strip() if v is not None else None


def _detect_columns(header_row):
    """Определяет индексы колонок по строке заголовков."""
    cols = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        name = str(cell).strip().lower()
        for field, synonyms in _HEADER_KEYS.items():
            if field in cols:
                continue
            for syn in synonyms:
                if syn in name:
                    cols[field] = idx
                    break
    if not _REQUIRED.issubset(cols.keys()):
        return None
    return cols


def _make_record(row, cols):
    """Строит словарь из строки данных по найденным индексам колонок."""
    def get(field):
        idx = cols.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    rp = _strip(get('rp'))
    if not rp:
        return None

    hours_raw = get('hours')
    if hours_raw is None or hours_raw == '':
        return None
    try:
        hours = float(hours_raw)
    except (ValueError, TypeError):
        return None
    if hours == 0:
        return None

    return {
        'rp':           rp,
        'dept':         _strip(get('dept')),
        'division':     _strip(get('division')),
        'employee':     _strip(get('employee')),
        'project':      _strip(get('project')),
        'task':         _strip(get('task')),
        'client':       _strip(get('client')),
        'project_type': _strip(get('project_type')),
        'work_type':    _strip(get('work_type')),
        'hours':        hours,
    }


# ─── Чтение .xlsb через pyxlsb ───────────────────────────────────────────────

def _parse_xlsb(path, rp_filter=None, pc_filter=None):
    """Читает .xlsb напрямую через pyxlsb."""
    try:
        import pyxlsb
    except ImportError:
        raise RuntimeError(
            "Для чтения .xlsb установите pyxlsb: откройте Терминал и выполните\n"
            "pip3 install pyxlsb"
        )

    records = []
    cols = None
    header_found = False

    with pyxlsb.open_workbook(path) as wb:
        # Ищем лист с именем 'data', иначе первый лист
        sheet_name = None
        for name in wb.sheets:
            if name.strip().lower() == 'data':
                sheet_name = name
                break
        if sheet_name is None:
            sheet_name = wb.sheets[0]

        with wb.get_sheet(sheet_name) as ws:
            for i, row in enumerate(ws.rows()):
                # pyxlsb возвращает объекты Cell; берём .v (value)
                row_vals = [c.v for c in row]

                if not header_found:
                    if i < 30:
                        detected = _detect_columns(row_vals)
                        if detected:
                            cols = detected
                            header_found = True
                    continue

                if cols is None:
                    break

                rec = _make_record(row_vals, cols)
                if rec is None:
                    continue
                if rp_filter and rec['rp'] != rp_filter:
                    continue
                if pc_filter and not _pc_match(rec.get('dept'), pc_filter):
                    continue
                records.append(rec)

    if not header_found:
        raise RuntimeError(
            "Строка заголовков не найдена в первых 30 строках. "
            "Убедитесь, что загружаете правильный файл «Отчёт распределение ресурсов»."
        )
    return records


# ─── Чтение .xlsx через openpyxl ─────────────────────────────────────────────

def _pc_match(dept, pc_filter):
    """Проверяет, содержит ли dept подстроку pc_filter (без учёта регистра)."""
    if not dept or not pc_filter:
        return False
    return pc_filter.lower() in dept.lower()


def _parse_xlsx(path, rp_filter=None, pc_filter=None):
    """Читает .xlsx через openpyxl."""
    wb = openpyxl.load_workbook(path, data_only=True)

    if not wb.sheetnames:
        raise RuntimeError("Файл не содержит листов. Возможно, он повреждён.")

    sheet_name = next(
        (n for n in wb.sheetnames if n.strip().lower() == 'data'),
        wb.sheetnames[0]
    )
    ws = wb[sheet_name]

    cols = None
    header_found = False
    records = []

    for i, row_vals in enumerate(ws.iter_rows(values_only=True)):
        if row_vals is None:
            continue
        row = list(row_vals)

        if not header_found:
            if i < 30:
                detected = _detect_columns(row)
                if detected:
                    cols = detected
                    header_found = True
            continue

        if cols is None:
            break

        rec = _make_record(row, cols)
        if rec is None:
            continue
        if rp_filter and rec['rp'] != rp_filter:
            continue
        if pc_filter and not _pc_match(rec.get('dept'), pc_filter):
            continue
        records.append(rec)

    wb.close()

    if not header_found:
        raise RuntimeError(
            "Строка заголовков не найдена в первых 30 строках. "
            "Убедитесь, что загружаете правильный файл «Отчёт распределение ресурсов»."
        )
    return records


# ─── Формат «Детализация до задачи» ────────────────────────────────────────

# Заголовки нового формата (нижний регистр, подстрока)
_DETAIL_COL_MAP = {
    'employee':    ('фио сотрудника',),
    'dept':        ('подразделение',),
    'rp_team':     ('рп команды',),
    'client':      ('клиент',),
    'project_num': ('номер\nпроекта', 'номер проекта'),
    'project':     ('наименование проекта',),
    'work_type':   ('продукт',),
    'rp_product':  ('рп продукта',),
    'task_num':    ('номер\nзадачи', 'номер задачи'),
    'task':        ('наименование\nзадачи', 'наименование задачи'),
    'hours':       ('ч/д', 'ч.д.', 'ч.д', 'чел.дней', 'чел/дней'),
}


def _is_detail_format(first_row):
    """True если первая строка содержит заголовки формата «Детализация»."""
    for cell in first_row[:4]:
        if cell and 'фио' in str(cell).strip().lower():
            return True
    return False


def _parse_detail_xlsx(path, rp_filter=None, pc_filter=None):
    """
    Читает «Списания по РП — детализация до задачи» (.xlsx).
    Единица исходного файла — Ч/Д (человеко-дни), конвертируется в часы (×8).
    Включает строки, где РП команды = rp_filter ИЛИ РП продукта = rp_filter.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    cols = {}
    records = []

    for i, row_vals in enumerate(ws.iter_rows(values_only=True)):
        row = list(row_vals) if row_vals else []

        if i == 0:
            # Первая строка — заголовок
            for idx, cell in enumerate(row):
                if cell is None:
                    continue
                name = str(cell).strip().lower()
                for field, synonyms in _DETAIL_COL_MAP.items():
                    if field in cols:
                        continue
                    for syn in synonyms:
                        if syn in name:
                            cols[field] = idx
                            break
            continue

        if not cols:
            break

        def get(field, _row=row):
            idx = cols.get(field)
            if idx is None or idx >= len(_row):
                return None
            v = _row[idx]
            return str(v).strip() if v is not None else None

        rp_team    = get('rp_team')
        rp_product = get('rp_product')

        # Фильтр РП: строка проходит если совпадает хоть одно из двух
        if rp_filter:
            if rp_team != rp_filter and rp_product != rp_filter:
                continue

        dept = get('dept')
        if pc_filter and not _pc_match(dept, pc_filter):
            continue

        hours_raw = get('hours')
        if hours_raw is None or hours_raw == '':
            continue
        try:
            hours = float(hours_raw) * 8   # Ч/Д → часы
        except (ValueError, TypeError):
            continue
        if hours == 0:
            continue

        proj_num  = get('project_num') or ''
        proj_name = get('project') or ''
        # Склеиваем номер + название, чтобы _RE_PROJ_NUM в ts_compare работал
        project_combined = f"{proj_num} {proj_name}".strip() if proj_num else proj_name

        task_num  = get('task_num') or ''
        task_name = get('task') or ''
        task_combined = f"{task_num} {task_name}".strip() if task_num else task_name

        records.append({
            'rp':           rp_team,
            'rp_product':   rp_product,
            'dept':         dept,
            'division':     dept,   # нет уровня дивизии — используем подразделение
            'employee':     get('employee'),
            'project':      project_combined,
            'task':         task_combined,
            'client':       get('client'),
            'project_type': None,
            'work_type':    get('work_type'),
            'hours':        hours,
        })

    wb.close()

    if not cols:
        raise RuntimeError(
            "Строка заголовков не распознана. "
            "Убедитесь, что файл содержит колонку «ФИО СОТРУДНИКА»."
        )
    return records


# ─── Публичный API ────────────────────────────────────────────────────────────

def parse_ts_file(filepath, rp_filter=None, pc_filter=None):
    """
    Разбирает файл трудозатрат (.xlsx или .xlsb).

    :param filepath:  путь к файлу
    :param rp_filter: если задан — строки где rp ИЛИ rp_product равны значению
    :param pc_filter: если задан — только строки, где dept содержит эту подстроку
    :return: (list[dict], file_type)
             dict-ключи: rp, rp_product, dept, division, employee, project,
                         task, client, project_type, work_type, hours
             file_type: 'office' | 'detail'
    """
    path = str(filepath)

    if path.lower().endswith('.xlsb'):
        records = _parse_xlsb(path, rp_filter=rp_filter, pc_filter=pc_filter)
        # У старого парсера нет rp_product — добавляем None
        for r in records:
            r.setdefault('rp_product', None)
        return records, 'office'

    # Для .xlsx определяем формат по первой строке
    wb_peek = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws_peek = wb_peek[wb_peek.sheetnames[0]]
    first_row = next(ws_peek.iter_rows(values_only=True), None) or []
    wb_peek.close()

    if _is_detail_format(first_row):
        return _parse_detail_xlsx(path, rp_filter=rp_filter, pc_filter=pc_filter), 'detail'
    else:
        records = _parse_xlsx(path, rp_filter=rp_filter, pc_filter=pc_filter)
        for r in records:
            r.setdefault('rp_product', None)
        return records, 'office'


def parse_employees_file(filepath):
    """
    Читает список сотрудников из Excel (.xlsx).
    Строка 1 — заголовок, строки 2+ — ФИО в первом столбце.
    """
    path = str(filepath)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    names = []
    for i, row_vals in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row_vals:
            continue
        v = row_vals[0] if len(row_vals) > 0 else None
        name = _strip(v)
        if name:
            names.append(name)
    wb.close()
    return names
