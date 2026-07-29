import os
import re
import glob
import shutil
import datetime
import json
import tempfile
import urllib.request
import urllib.error
import urllib.parse
from pathlib import Path
from functools import wraps
from collections import Counter

from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

from app.db import (
    get_conn, init_db, TEAM_LEADS, OWNER_SECTIONS, TEAM_LEAD_SECTIONS, OWNER_LABEL, ALL_RESPONSIBLE,
    TRACKING_ONLY_SECTIONS, FOCUS_SECTIONS, AGGREGATE_CATEGORY_MAP, AGGREGATE_CATEGORY_ORDER,
    AGGREGATE_CATEGORY_SECTIONS, RISK_LEVELS, get_setting, set_setting,
    next_quarter_label, quarter_options, compute_quarter_label,
    get_available_quarters, get_available_fiscal_years, calendar_to_fiscal, fiscal_year_quarters,
)
from app.importer import import_excel
from app.stt import transcribe, SttError
from app.ts_parser import parse_ts_file, parse_employees_file

BASE_DIR = Path(__file__).parent
UPLOAD_DIR = BASE_DIR / "data" / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

app = Flask(__name__)
app.secret_key = os.environ.get("FP_PORTAL_SECRET", "dev-secret-change-me-in-production")
app.config["MAX_CONTENT_LENGTH"] = 200 * 1024 * 1024  # 200MB (для больших xlsb)

# Версия статики (по времени изменения style.css) -- добавляется к ссылке на CSS в base.html
# как ?v=..., чтобы браузер мог кэшировать файл надолго, но сразу подхватывал новую версию
# после следующего перезапуска приложения (т.е. после деплоя с изменённым style.css).
try:
    ASSET_VERSION = str(int((BASE_DIR / "static" / "css" / "style.css").stat().st_mtime))
except OSError:
    ASSET_VERSION = "1"

# Чат «Спросить портал» отвечает через локальную модель в Ollama -- запрос идёт на тот же
# компьютер (по умолчанию http://localhost:11434), данные никуда в интернет не уходят.
# Модель и адрес можно переопределить переменными окружения, если у вас другой порт/модель.
OLLAMA_HOST = os.environ.get("OLLAMA_HOST", "http://localhost:11434").rstrip("/")
OLLAMA_MODEL = os.environ.get("OLLAMA_MODEL", "qwen3:8b")
OLLAMA_TIMEOUT = float(os.environ.get("OLLAMA_TIMEOUT", "60"))

MONTH_ORDER = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
               "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
MONTH_INDEX = {m: i for i, m in enumerate(MONTH_ORDER)}

QUARTER_MONTHS = {
    1: ["Январь", "Февраль", "Март"],
    2: ["Апрель", "Май", "Июнь"],
    3: ["Июль", "Август", "Сентябрь"],
    4: ["Октябрь", "Ноябрь", "Декабрь"],
}

# Блок «К получению в ближайшие N дней» на дашборде квартала -- суммы по активным статьям,
# у которых дата ДПА попадает в ближайшие DPA_DUE_WINDOW_DAYS дней. Факт не учитывается --
# деньги по таким статьям уже получены, забирать повторно нечего. Это календарный, а не
# плановый разрез (в отличие от остального дашборда, который смотрит на месяцы текущего
# квартала) -- считается по всем активным строкам независимо от месяца/периода.
# Разбивка -- по явному списку категорий из запроса пользователя: Проекты, Заказные
# доработки и Лицензии отдельно, Сопровождение и TaaS -- вместе. Докупки сюда не входят.
DPA_DUE_WINDOW_DAYS = 4
DPA_DUE_CATEGORIES = [
    ("Проекты", ["Проекты"]),
    ("Заказные доработки", ["Заказные доработки"]),
    ("Лицензии", ["Лицензии"]),
    ("Сопровождение + TaaS", ["Сопровождение", "TaaS"]),
]

# Постоянная очередь изменений (row_events), не привязанная к одной загрузке: три вкладки
TAB_EVENT_TYPES = {
    "new": ("new", "reactivated"),
    "closed": ("zeroed", "deactivated"),
    "other": ("amount_changed", "field_changed"),
}
TAB_LABELS = {
    "new": "Новые статьи",
    "closed": "Закрытые / обнулённые",
    "other": "Прочие изменения",
}

# ---------- Чат «Спросить портал»: данные для ИИ + запасной поиск без ИИ ----------
# Разбор вопроса по ключевым словам и названиям клиентов/разделов используется двояко:
# 1) чтобы собрать релевантный кусок данных из БД и передать его локальной модели в Ollama
#    (см. answer_portal_query/_build_chat_context ниже) -- модель отвечает обычным языком;
# 2) как запасной вариант (_answer_portal_query_local), если Ollama недоступна -- тогда
#    пользователь получает готовый моноширенный текст-ответ напрямую, без ИИ.

OBLIGATION_STATUS_LABELS = {
    "not_started": "не начато", "in_progress": "в работе", "done": "выполнено", "blocked": "блокировано",
}

SECTION_KEYWORDS = [
    ("сопровожд", "Сопровождение"),
    ("доработ", "Заказные доработки"),
    ("лицензи", "Лицензии"),
    ("докупк", "Докупки"),
    ("taas", "TaaS"), ("таас", "TaaS"),
    ("проект", "Проекты"),
]
PORTFOLIO_KEYWORDS = [
    ("возможност", "Возможности"),
    ("факт", "Факт"),
    ("план", "0-100"), ("0-100", "0-100"), ("0 100", "0-100"),
]
RISK_WORDS = ("риск",)
OBLIGATION_WORDS = ("обязательств",)
OVERDUE_WORDS = ("просроч",)
DUE_SOON_WORDS = ("горящ", "горит", "ближайш")
HELP_WORDS = ("помощь", "что ты умеешь", "команды", "/help")

HELP_TEXT = (
    "Я ищу по данным портала локально, без ИИ. Понимаю такие запросы:\n\n"
    "  Сбербанк                  - сводка и строки по клиенту\n"
    "  2606/23-2В-О               - найти статью по номеру договора/проекта/ЕЗ\n"
    "  обязательства Сбербанк    - открытые обязательства по клиенту\n"
    "  риски                     - статьи без ответственного + отмеченные риском вручную\n"
    "  обязательства             - все открытые обязательства\n"
    "  просрочки                 - просроченные обязательства\n"
    "  горящие сроки             - обязательства со сроком до 3 дней\n"
    "  возможности по TaaS        - сводка по клиентам для раздела/портфеля\n"
    "  факт по сопровождению     - то же для Факта\n\n"
    "Клиента и раздел/портфель можно сочетать в одном вопросе.\n\n"
    "Также через чат можно выполнять действия (если ваш аккаунт привязан):\n"
    "  поставь высокий риск по Сберу\n"
    "  добавь обязательство по проекту 1234/25-1А: подписать акт, ответственный Иванов, срок 2026-07-01\n"
    "  обязательство по Сберу выполнено\n"
    "  поменяй комментарий по Сберу на ...\n\n"
    "Перед любым действием чат сначала покажет, что собирается сделать, и ничего не "
    "изменит в БД, пока вы не ответите «да» (отменить — «нет»).\n\n"
    "Вопрос можно не печатать, а наговорить — кнопка микрофона рядом с полем вопроса.\n"
)


def fmt_money(amount):
    return "{:,.0f}".format(amount or 0).replace(",", " ") + " ₽"


def _truncate(text, max_len):
    text = text or "—"
    if len(text) <= max_len:
        return text
    return text[: max_len - 1] + "…"


def _tokenize(text):
    return re.findall(r"[a-zа-яё0-9]+", (text or "").lower())


def _find_matching_clients(question, all_clients):
    """Сильное совпадение -- полное (многословное) имя клиента встречается в вопросе как подстрока.
    Слабое -- хотя бы одно характерное слово клиента (>=4 букв) совпадает по префиксу со словом
    вопроса (покрывает падежные окончания типа 'Сбербанка', 'Сбербанку')."""
    q_lower = question.lower()
    q_tokens = _tokenize(question)
    strong, weak = [], []
    for c in all_clients:
        c_lower = c.lower()
        if c_lower in q_lower:
            strong.append(c)
            continue
        for ct in (t for t in _tokenize(c) if len(t) >= 4):
            if any((qt.startswith(ct) or (len(qt) >= 4 and ct.startswith(qt))) for qt in q_tokens):
                weak.append(c)
                break
    return strong if strong else weak


def _find_keywords(question, mapping):
    q_lower = question.lower()
    found = []
    for kw, label in mapping:
        if kw in q_lower and label not in found:
            found.append(label)
    return found


def _any_word(question, words):
    q_lower = question.lower()
    return any(w in q_lower for w in words)


def _find_responsible(question):
    q_lower = question.lower()
    for name in ALL_RESPONSIBLE:
        if name.lower() in q_lower:
            return name
        surname = name.split()[0].lower()
        if len(surname) >= 5 and surname in q_lower:
            return name
    return None


def _format_obligations_table(obls, show_client=True):
    today_iso = datetime.date.today().isoformat()
    prefix_hdr = f"{'Клиент':<18} {'Проект':<22} {'Проект №':<10} " if show_client else ""
    header = prefix_hdr + f"{'Обязательство':<30} {'Ответственный':<18} {'Срок':>10} {'Статус':<12}"
    lines = [header, "-" * len(header)]
    for o in obls:
        overdue = bool(o["due_date"]) and o["due_date"] < today_iso and o["status"] != "done"
        status_label = OBLIGATION_STATUS_LABELS.get(o["status"], o["status"] or "—")
        if overdue:
            status_label += " ⚠"
        prefix = (
            f"{_truncate(o['client_name'], 18):<18} {_truncate(o['project_name'], 22):<22} "
            f"{_truncate(o['project_num'] or '-', 10):<10} "
        ) if show_client else ""
        due = o["due_date"] or "без срока"
        lines.append(
            f"{prefix}{_truncate(o['title'], 30):<30} {_truncate(o['responsible_name'], 18):<18} "
            f"{due:>10} {status_label:<12}"
        )
    return "\n".join(lines)


def _answer_for_clients(conn, clients, sections, portfolios, show_obligations):
    ph = ",".join("?" * len(clients))
    rows = conn.execute(
        f"SELECT * FROM fp_rows WHERE is_active = 1 AND client_name IN ({ph}) "
        f"ORDER BY dpa_date IS NULL, dpa_date ASC", clients,
    ).fetchall()
    if sections:
        rows = [r for r in rows if r["section"] in sections]
    if portfolios:
        rows = [r for r in rows if (r["portfolio"] or "Прочее") in portfolios]
    if not rows:
        return f"По клиенту(ам) «{', '.join(clients)}» с такими фильтрами активных строк не найдено."

    row_ids = [r["id"] for r in rows]
    obl_by_row = {}
    ph2 = ",".join("?" * len(row_ids))
    for o in conn.execute(f"SELECT * FROM obligations WHERE fp_row_id IN ({ph2})", row_ids).fetchall():
        obl_by_row.setdefault(o["fp_row_id"], []).append(o)

    totals = {"Факт": 0.0, "0-100": 0.0, "Возможности": 0.0}
    for r in rows:
        p = r["portfolio"] or "Прочее"
        totals[p] = totals.get(p, 0.0) + (r["amount_0_100"] or 0)

    lines = [", ".join(sorted(set(r["client_name"] for r in rows))), ""]
    lines.append(
        f"Факт: {fmt_money(totals.get('Факт', 0))}   "
        f"План 0-100: {fmt_money(totals.get('0-100', 0))}   "
        f"Возможности: {fmt_money(totals.get('Возможности', 0))}"
    )
    lines.append(f"Строк: {len(rows)}")
    lines.append("")

    header = f"{'Проект':<26} {'Проект №':<10} {'Раздел':<16} {'Портфель':<11} {'Сумма':>13} {'ДПА':>10} {'Об.':>3}"
    lines.append(header)
    lines.append("-" * len(header))
    for r in rows:
        obls = obl_by_row.get(r["id"], [])
        obl_count = str(len(obls)) if obls else "-"
        lines.append(
            f"{_truncate(r['project_name'], 26):<26} {_truncate(r['project_num'] or '-', 10):<10} "
            f"{_truncate(r['section'], 16):<16} "
            f"{_truncate(r['portfolio'] or '-', 11):<11} "
            f"{'{:,.0f}'.format(r['amount_0_100'] or 0).replace(',', ' '):>13} "
            f"{(r['dpa_date'] or '-'):>10} {obl_count:>3}"
        )

    if show_obligations:
        open_obls = [o for r in rows for o in obl_by_row.get(r["id"], []) if o["status"] != "done"]
        lines.append("")
        if open_obls:
            ids_by_row = {r["id"]: r for r in rows}
            decorated = []
            for o in open_obls:
                r = ids_by_row.get(o["fp_row_id"])
                decorated.append({**dict(o), "client_name": r["client_name"] if r else "-",
                                   "project_name": r["project_name"] if r else "-",
                                   "project_num": r["project_num"] if r else "-"})
            lines.append("Открытые обязательства:")
            lines.append(_format_obligations_table(decorated, show_client=False))
        else:
            lines.append("Открытых обязательств по этим строкам нет.")

    return "\n".join(lines)


def _answer_obligations(conn, responsible, only_overdue, only_due_soon):
    query = (
        "SELECT o.*, f.client_name, f.project_name, f.project_num FROM obligations o "
        "JOIN fp_rows f ON f.id = o.fp_row_id WHERE o.status != 'done'"
    )
    params = []
    if responsible:
        query += " AND o.responsible_name = ?"
        params.append(responsible)
    query += " ORDER BY o.due_date IS NULL, o.due_date ASC"
    obls = conn.execute(query, params).fetchall()

    today = datetime.date.today()

    def is_overdue(o):
        return bool(o["due_date"]) and datetime.date.fromisoformat(o["due_date"]) < today

    def is_due_soon(o):
        if not o["due_date"]:
            return False
        d = datetime.date.fromisoformat(o["due_date"])
        return today <= d and (d - today).days <= 3

    if only_overdue:
        obls = [o for o in obls if is_overdue(o)]
        title = "Просроченные обязательства"
    elif only_due_soon:
        obls = [o for o in obls if is_due_soon(o)]
        title = "Горящие сроки (≤3 дня)"
    else:
        title = "Открытые обязательства"
    if responsible:
        title += f" — {responsible}"

    if not obls:
        return f"{title}: ничего не найдено."

    lines = [f"{title} ({len(obls)}):", "", _format_obligations_table(obls, show_client=True)]
    return "\n".join(lines)


def _answer_risky_rows(conn):
    months, qnum, outlier_months = get_period_info()
    placeholders = ",".join("?" * len(months)) if months else "''"
    tl_ph = ",".join("?" * len(TEAM_LEAD_SECTIONS))
    own_ph = ",".join("?" * len(OWNER_SECTIONS))
    rows = conn.execute(f"""
        SELECT f.* FROM fp_rows f
        WHERE f.is_active = 1
        AND f.month IN ({placeholders})
        AND (
            (f.section IN ({tl_ph}) AND NOT EXISTS (
                SELECT 1 FROM obligations o WHERE o.fp_row_id = f.id AND o.responsible_type = 'team_lead'
            ))
            OR
            (f.section IN ({own_ph}) AND NOT EXISTS (
                SELECT 1 FROM obligations o WHERE o.fp_row_id = f.id AND o.responsible_type = 'owner'
            ))
        )
    """, months + list(TEAM_LEAD_SECTIONS) + list(OWNER_SECTIONS)).fetchall()
    rows = [r for r in rows if (r["amount_0_100"] or 0) >= 300000]
    rows = sorted(rows, key=lambda r: -(r["amount_0_100"] or 0))

    if not rows:
        return "Статей без оформленного обязательства/ответственного (от 300 000 ₽) сейчас нет."

    total = sum((r["amount_0_100"] or 0) for r in rows)
    lines = [f"Статьи без оформленного обязательства/ответственного, от 300 000 ₽ "
             f"({len(rows)}, итого {fmt_money(total)}):", ""]
    header = f"{'Клиент':<20} {'Проект':<26} {'Проект №':<10} {'Раздел':<20} {'Сумма':>13} {'Нужен':<14}"
    lines.append(header)
    lines.append("-" * len(header))
    for r in rows[:30]:
        who = "рук. команды" if r["section"] in TEAM_LEAD_SECTIONS else "Owner"
        lines.append(
            f"{_truncate(r['client_name'], 20):<20} {_truncate(r['project_name'], 26):<26} "
            f"{_truncate(r['project_num'] or '-', 10):<10} "
            f"{_truncate(r['section'], 20):<20} "
            f"{'{:,.0f}'.format(r['amount_0_100'] or 0).replace(',', ' '):>13} {who:<14}"
        )
    if len(rows) > 30:
        lines.append(f"... и ещё {len(rows) - 30} строк")
    return "\n".join(lines)


def _answer_manual_risk_rows(conn):
    """Статьи, которые пользователь вручную отметил риском (поле risk_level на детальной
    строке/в агрегированной таблице: Низкий/Средний/Высокий). Это ДРУГОЕ понятие, чем в
    _answer_risky_rows выше -- там речь про статьи без оформленного обязательства/
    ответственного. Раньше чат вообще не видел эту ручную отметку, из-за чего не находил
    статьи, которые пользователь сам пометил как риск."""
    rows = conn.execute(
        "SELECT * FROM fp_rows WHERE is_active = 1 AND risk_level > 0 "
        "ORDER BY risk_level DESC, amount_0_100 DESC"
    ).fetchall()
    if not rows:
        return "Статей, отмеченных риском вручную, сейчас нет."

    lines = [f"Статьи, отмеченные риском вручную ({len(rows)}):", ""]
    header = f"{'Уровень':<14} {'Клиент':<20} {'Проект':<24} {'Проект №':<10} {'Договор':<16} {'Сумма':>13}"
    lines.append(header)
    lines.append("-" * len(header))
    for r in rows[:30]:
        level_label = RISK_LEVELS.get(r["risk_level"], {}).get("label", "—")
        lines.append(
            f"{level_label:<14} {_truncate(r['client_name'], 20):<20} "
            f"{_truncate(r['project_name'], 24):<24} {_truncate(r['project_num'] or '-', 10):<10} "
            f"{_truncate(r['contract_num'] or '-', 16):<16} "
            f"{'{:,.0f}'.format(r['amount_0_100'] or 0).replace(',', ' '):>13}"
        )
    if len(rows) > 30:
        lines.append(f"... и ещё {len(rows) - 30} строк")
    return "\n".join(lines)


def _normalize_digits(s):
    """Убирает пробелы (включая неразрывные) из числа -- в БД project_num/contract_num часто
    хранятся с разделителями тысяч вида "3\xa0856\xa0399", а в вопросе пользователь обычно
    пишет номер слитно ("3856399")."""
    return re.sub(r"\s+", "", s or "")


def _find_code_matches(conn, question):
    """Ищет в вопросе подстроки, похожие на номер договора/проекта/ЕЗ (буквенно-цифровые
    куски, соединённые "/" или "-", например "2606/23-2В-О"), и возвращает активные строки
    ФП, у которых contract_num/project_num/ez_num содержит такую подстроку. Раньше чат умел
    находить статьи только по названию клиента -- вопрос с одним лишь номером договора (без
    клиента) не давал вообще никакого совпадения.

    Отдельно также ищет голые числа из 5+ цифр (например, просто "3856399", без "/" или "-") --
    это типичная форма номера проекта, который пользователь вводит слитно, хотя в БД он хранится
    с пробелами-разделителями тысяч ("3 856 399") и поэтому не находится обычным LIKE."""
    candidates = [c for c in re.findall(
        r"[A-Za-zА-Яа-яЁё0-9]+(?:[/\-][A-Za-zА-Яа-яЁё0-9]+)+", question,
    ) if len(c) >= 4]
    seen_ids, matches = set(), []
    for code in candidates:
        like = f"%{code}%"
        rows = conn.execute(
            "SELECT * FROM fp_rows WHERE is_active = 1 AND "
            "(contract_num LIKE ? OR project_num LIKE ? OR ez_num LIKE ?)",
            (like, like, like),
        ).fetchall()
        for r in rows:
            if r["id"] not in seen_ids:
                seen_ids.add(r["id"])
                matches.append(r)

    bare_numbers = re.findall(r"\d{5,}", question)
    if bare_numbers:
        for r in conn.execute("SELECT * FROM fp_rows WHERE is_active = 1").fetchall():
            if r["id"] in seen_ids:
                continue
            norm_project = _normalize_digits(r["project_num"])
            norm_contract = _normalize_digits(r["contract_num"])
            if any(n == norm_project or n == norm_contract for n in bare_numbers):
                seen_ids.add(r["id"])
                matches.append(r)
    return matches


def _answer_code_matches(conn, matches):
    """Подробный блок по строкам, найденным через _find_code_matches -- включает все
    идентификаторы (договор/проект/ЕЗ), сумму, ручную отметку риска, заметку/комментарий и
    открытые обязательства, чтобы по вопросу с конкретным номером можно было ответить
    полностью, без дополнительных уточнений."""
    row_ids = [r["id"] for r in matches]
    ph = ",".join("?" * len(row_ids))
    obl_by_row = {}
    for o in conn.execute(f"SELECT * FROM obligations WHERE fp_row_id IN ({ph})", row_ids).fetchall():
        obl_by_row.setdefault(o["fp_row_id"], []).append(o)

    lines = [f"Найдено по коду из вопроса ({len(matches)}):", ""]
    for r in matches:
        risk_label = RISK_LEVELS.get(r["risk_level"], {}).get("label", "не отмечен")
        lines.append(
            f"Клиент: {r['client_name']}  |  Проект: {r['project_name']}  |  "
            f"Раздел: {r['section']}  |  Портфель: {r['portfolio'] or '-'}"
        )
        lines.append(
            f"  Договор: {r['contract_num'] or '-'}   Проект №: {r['project_num'] or '-'}   "
            f"ЕЗ: {r['ez_num'] or '-'}"
        )
        lines.append(
            f"  Сумма 0-100: {fmt_money(r['amount_0_100'])}   Риск (вручную): {risk_label}   "
            f"ДПА: {r['dpa_date'] or '-'}"
        )
        if r["note"]:
            lines.append(f"  Заметка: {r['note']}")
        if r["mp_comment"]:
            lines.append(f"  Комментарий: {r['mp_comment']}")
        obls = obl_by_row.get(r["id"], [])
        if obls:
            lines.append("  Обязательства:")
            for o in obls:
                status_label = OBLIGATION_STATUS_LABELS.get(o["status"], o["status"] or "—")
                lines.append(
                    f"    - {o['title']} | {o['responsible_name']} | "
                    f"{o['due_date'] or 'без срока'} | {status_label}"
                )
        lines.append("")
    return "\n".join(lines).strip()


def _answer_section_breakdown(conn, sections, portfolios):
    months, qnum, outlier_months = get_period_info()
    placeholders = ",".join("?" * len(months)) if months else "''"
    rows = conn.execute(f"SELECT * FROM fp_rows WHERE is_active = 1 AND month IN ({placeholders})", months).fetchall()
    if sections:
        rows = [r for r in rows if r["section"] in sections]

    by_client = {}
    for r in rows:
        p = r["portfolio"] or "Прочее"
        if portfolios and p not in portfolios:
            continue
        d = by_client.setdefault(r["client_name"], {"Факт": 0.0, "0-100": 0.0, "Возможности": 0.0})
        d[p] = d.get(p, 0.0) + (r["amount_0_100"] or 0)
    by_client = {c: v for c, v in by_client.items() if any(v.values())}

    if not by_client:
        scope = ", ".join(sections) if sections else "все разделы"
        return f"По разделу(ам) «{scope}» с такими фильтрами данных не нашлось."

    totals = {k: sum(v.get(k, 0.0) for v in by_client.values()) for k in ("Факт", "0-100", "Возможности")}
    scope_label = ", ".join(sections) if sections else "все разделы"
    portfolio_label = ", ".join(portfolios) if portfolios else "Факт/0-100/Возможности"
    lines = [f"{scope_label} — {portfolio_label}:", ""]
    lines.append(
        f"Итого: Факт {fmt_money(totals['Факт'])}   "
        f"0-100 {fmt_money(totals['0-100'])}   "
        f"Возможности {fmt_money(totals['Возможности'])}"
    )
    lines.append("")

    sort_key = portfolios[0] if len(portfolios) == 1 else "0-100"
    ordered = sorted(
        by_client.items(),
        key=lambda kv: -(kv[1].get(sort_key, 0.0) + kv[1].get("Возможности", 0.0)),
    )

    header = f"{'Клиент':<26} {'Факт':>13} {'0-100':>13} {'Возможности':>13}"
    lines.append(header)
    lines.append("-" * len(header))
    for client, vals in ordered[:25]:
        lines.append(
            f"{_truncate(client, 26):<26} {fmt_money(vals['Факт']):>13} "
            f"{fmt_money(vals['0-100']):>13} {fmt_money(vals['Возможности']):>13}"
        )
    if len(ordered) > 25:
        lines.append(f"... и ещё {len(ordered) - 25} клиентов")
    return "\n".join(lines)


def _answer_portal_query_local(question):
    """Запасной вариант ответа: локальный разбор вопроса по ключевым словам, без обращения
    к внешним сервисам и без ИИ. Используется, когда Ollama недоступна (не запущена, модель
    не скачана, таймаут и т.п.) -- см. answer_portal_query ниже. Возвращает готовый
    моноширенный текст-ответ."""
    q = (question or "").strip()
    if not q or _any_word(q, HELP_WORDS):
        return HELP_TEXT

    conn = get_conn()
    try:
        code_matches = _find_code_matches(conn, q)
        all_clients = [r["client_name"] for r in conn.execute(
            "SELECT DISTINCT client_name FROM fp_rows WHERE is_active = 1"
        ).fetchall()]

        matched_clients = _find_matching_clients(q, all_clients)
        sections = _find_keywords(q, SECTION_KEYWORDS)
        portfolios = _find_keywords(q, PORTFOLIO_KEYWORDS)
        want_risk = _any_word(q, RISK_WORDS)
        want_obligations = _any_word(q, OBLIGATION_WORDS)
        want_overdue = _any_word(q, OVERDUE_WORDS)
        want_due_soon = _any_word(q, DUE_SOON_WORDS)
        responsible = _find_responsible(q)

        if code_matches:
            answer = _answer_code_matches(conn, code_matches)
        elif matched_clients:
            answer = _answer_for_clients(
                conn, matched_clients[:5], sections, portfolios,
                show_obligations=want_obligations or want_overdue or want_due_soon,
            )
        elif want_overdue or want_due_soon or (want_obligations and not sections and not portfolios):
            answer = _answer_obligations(
                conn, responsible,
                only_overdue=want_overdue and not want_due_soon,
                only_due_soon=want_due_soon and not want_overdue,
            )
        elif want_risk and not sections and not portfolios:
            answer = _answer_risky_rows(conn) + "\n\n---\n\n" + _answer_manual_risk_rows(conn)
        elif sections or portfolios:
            answer = _answer_section_breakdown(conn, sections, portfolios)
        else:
            answer = (
                "Не нашёл в вопросе известного клиента, раздела или ключевого слова "
                "(риски/обязательства/просрочки).\n\n" + HELP_TEXT
            )
    finally:
        conn.close()
    return answer


def _strip_think_tags(text):
    """Некоторые модели оборачивают рассуждения в <think>...</think> перед финальным ответом --
    эти теги пользователю в чате не нужны, вырезаем их на всякий случай."""
    return re.sub(r"<think>.*?</think>", "", text, flags=re.DOTALL).strip()


def _call_ollama(messages, timeout=None):
    """Отправляет запрос локальной модели в Ollama по его REST API (голый urllib, без
    дополнительных пакетов). Бросает исключение при любой проблеме (Ollama не запущена,
    модель не скачана, таймаут, плохой ответ) -- вызывающий код сам решает, что делать
    дальше (см. answer_portal_query)."""
    payload = json.dumps({
        "model": OLLAMA_MODEL,
        "messages": messages,
        "stream": False,
        # Современные модели (например, Qwen3.x) по умолчанию включают режим "размышлений"
        # перед ответом -- это в 5-10 раз медленнее и нам тут не нужно: вопрос простой,
        # ответ должен быть быстрым и по делу. think=false отключает это (модели без
        # поддержки режима размышлений просто игнорируют параметр).
        "think": False,
        "options": {"temperature": 0.2, "num_ctx": 8192},
    }).encode("utf-8")
    req = urllib.request.Request(
        f"{OLLAMA_HOST}/api/chat", data=payload,
        headers={"Content-Type": "application/json"}, method="POST",
    )
    with urllib.request.urlopen(req, timeout=timeout or OLLAMA_TIMEOUT) as resp:
        data = json.loads(resp.read().decode("utf-8"))
    content = (data.get("message") or {}).get("content", "").strip()
    if not content:
        raise RuntimeError("Ollama вернула пустой ответ")
    return _strip_think_tags(content)


def _build_chat_context(conn, question):
    """Собирает текстовый контекст с данными портала по теме вопроса -- те же блоки, что
    раньше показывались пользователю как готовый ответ при локальном поиске (см.
    _answer_portal_query_local), плюс блок рисков по умолчанию, чтобы модель ориентировалась
    даже в общих вопросах без точного совпадения по клиенту/разделу."""
    months, qnum, _ = get_period_info()
    period_label = "-".join(months) if months else "период"
    blocks = [f"Текущий рассматриваемый период: {period_label}."]

    # Поиск по номеру договора/проекта/ЕЗ -- проверяем первым: если в вопросе есть конкретный
    # код, это самое точное совпадение, даже если клиент/раздел не упомянуты или не распознаны.
    code_matches = _find_code_matches(conn, question)
    if code_matches:
        blocks.append(_answer_code_matches(conn, code_matches))

    all_clients = [r["client_name"] for r in conn.execute(
        "SELECT DISTINCT client_name FROM fp_rows WHERE is_active = 1"
    ).fetchall()]
    matched_clients = _find_matching_clients(question, all_clients)
    sections = _find_keywords(question, SECTION_KEYWORDS)
    portfolios = _find_keywords(question, PORTFOLIO_KEYWORDS)
    want_risk = _any_word(question, RISK_WORDS)
    want_obligations = _any_word(question, OBLIGATION_WORDS)
    want_overdue = _any_word(question, OVERDUE_WORDS)
    want_due_soon = _any_word(question, DUE_SOON_WORDS)
    responsible = _find_responsible(question)

    if matched_clients:
        blocks.append(_answer_for_clients(
            conn, matched_clients[:5], sections, portfolios, show_obligations=True,
        ))
    if sections or portfolios:
        blocks.append(_answer_section_breakdown(conn, sections, portfolios))
    if want_obligations or want_overdue or want_due_soon or responsible:
        blocks.append(_answer_obligations(
            conn, responsible,
            only_overdue=want_overdue and not want_due_soon,
            only_due_soon=want_due_soon and not want_overdue,
        ))
    # Риски показываем по умолчанию (если вопрос не попал ни в один блок выше), а также
    # всегда, если про риски спросили явно -- та же логика, что раньше была в локальном поиске.
    # Два разных блока: _answer_risky_rows -- статьи без оформленного обязательства (порог
    # 300k), _answer_manual_risk_rows -- то, что пользователь вручную отметил полем "Риск".
    # Раньше второй блок не собирался вообще, поэтому чат не видел ручные отметки риска.
    if want_risk or not (matched_clients or sections or portfolios or want_obligations
                          or want_overdue or want_due_soon or code_matches):
        blocks.append(_answer_risky_rows(conn))
        blocks.append(_answer_manual_risk_rows(conn))

    context = "\n\n---\n\n".join(blocks)
    max_chars = 9000
    if len(context) > max_chars:
        context = context[:max_chars] + "\n... (контекст обрезан, дальше есть ещё строки)"
    return context


# ---------- Действия через чат (риск, ответственные, статус обязательств, комментарий) ----------
# Кроме вопросов, чат умеет выполнять те же действия, что доступны на страницах портала:
# проставить/снять отметку риска, добавить обязательство (= назначить ответственного), изменить
# статус обязательства, изменить внутренний комментарий по статье. Сначала отдельным строгим
# JSON-запросом к Ollama определяем, действие это или обычный вопрос (_detect_chat_action), и
# если действие -- выполняем его прямо в БД с той же проверкой прав, что у веб-роутов (_exec_*).

ACTION_SYSTEM_PROMPT = """Ты — модуль распознавания команд портала «ФП-Контроль». Определи, \
является ли сообщение пользователя запросом на ДЕЙСТВИЕ (изменить данные), а не обычным \
вопросом/просьбой что-то посмотреть. Если не уверен — это не действие.

Ответь СТРОГО одним JSON-объектом, без пояснений и без markdown-разметки (без ```), по одной \
из схем:

Обычный вопрос: {"action": "none"}

Поставить/снять отметку риска по статье:
{"action": "set_risk", "row_query": "<клиент/проект/номер договора из сообщения>", "risk_level": 0}
risk_level: 0 — снять отметку, 1 — низкий, 2 — средний, 3 — высокий.

Добавить обязательство / назначить ответственного по статье:
{"action": "add_obligation", "row_query": "...", "title": "<суть обязательства>", \
"responsible_name": "<ФИО ответственного, как в сообщении>", "due_date": "YYYY-MM-DD или null"}

Изменить статус обязательства по статье:
{"action": "set_obligation_status", "row_query": "...", "title_hint": "<часть текста обязательства или пусто>", \
"status": "not_started|in_progress|done|blocked"}

Изменить внутренний комментарий по статье:
{"action": "set_comment", "row_query": "...", "comment": "<новый текст комментария>"}

row_query — то, по чему можно найти статью ФП: название клиента, номер договора/проекта/ЕЗ, \
название проекта — бери из сообщения как можно точнее. Верни только JSON, ничего больше."""

ACTION_DETECT_TIMEOUT = min(OLLAMA_TIMEOUT, 20)

RISK_LEVEL_WORDS = {"высокий": 3, "средний": 2, "низкий": 1, "нет": 0, "снять": 0, "убрать": 0, "сбросить": 0}


def _normalize_risk_level(value):
    if isinstance(value, bool):
        return 0
    if isinstance(value, int):
        return value if value in (0, 1, 2, 3) else 0
    if isinstance(value, str):
        v = value.strip().lower()
        if v.isdigit() and int(v) in (0, 1, 2, 3):
            return int(v)
        for word, lvl in RISK_LEVEL_WORDS.items():
            if word in v:
                return lvl
    return 0


OBLIGATION_STATUS_WORDS = {
    "не начато": "not_started", "в работе": "in_progress", "выполнено": "done",
    "готово": "done", "сделано": "done", "блокировано": "blocked", "блок": "blocked",
}


def _normalize_obligation_status(value):
    if not value:
        return None
    v = str(value).strip().lower()
    if v in ("not_started", "in_progress", "done", "blocked"):
        return v
    for word, key in OBLIGATION_STATUS_WORDS.items():
        if word in v:
            return key
    return None


def _match_responsible_name(name_text):
    """Сопоставляет текст из ответа модели (ФИО ответственного) со списком допустимых имён --
    точное совпадение, затем по подстроке, затем по фамилии (первое слово имени)."""
    if not name_text:
        return None
    text_lower = str(name_text).strip().lower()
    if not text_lower:
        return None
    for n in ALL_RESPONSIBLE:
        if n.lower() == text_lower:
            return n
    for n in ALL_RESPONSIBLE:
        if n.lower() in text_lower or text_lower in n.lower():
            return n
    for n in ALL_RESPONSIBLE:
        surname = n.split()[0].lower()
        if len(surname) >= 3 and surname in text_lower:
            return n
    return None


def _resolve_rows_for_action(conn, query_text):
    """Находит активные строки ФП по тексту из действия чата: код договора/проекта/ЕЗ -- самое
    точное совпадение, иначе клиент, иначе подстрока в названии проекта. Если кандидатов больше
    одного, дополнительно сужает по разделу/портфелю/месяцу, если они упомянуты в тексте -- это
    покрывает большинство реальных формулировок без отдельного диалога выбора."""
    query_text = (query_text or "").strip()
    if not query_text:
        return []

    code_matches = _find_code_matches(conn, query_text)
    if code_matches:
        return code_matches

    all_rows = conn.execute("SELECT * FROM fp_rows WHERE is_active = 1").fetchall()
    all_clients = sorted(set(r["client_name"] for r in all_rows))
    matched_clients = _find_matching_clients(query_text, all_clients)
    if matched_clients:
        candidates = [r for r in all_rows if r["client_name"] in matched_clients]
    else:
        q_lower = query_text.lower()
        candidates = [r for r in all_rows if q_lower in (r["project_name"] or "").lower()]
        if not candidates:
            return []

    if len(candidates) > 1:
        sections = _find_keywords(query_text, SECTION_KEYWORDS)
        if sections:
            narrowed = [r for r in candidates if r["section"] in sections]
            if narrowed:
                candidates = narrowed
    if len(candidates) > 1:
        portfolios = _find_keywords(query_text, PORTFOLIO_KEYWORDS)
        if portfolios:
            narrowed = [r for r in candidates if r["portfolio"] in portfolios]
            if narrowed:
                candidates = narrowed
    if len(candidates) > 1:
        q_lower = query_text.lower()
        for m in MONTH_ORDER:
            if m.lower() in q_lower:
                narrowed = [r for r in candidates if r["month"] == m]
                if narrowed:
                    candidates = narrowed
                break
    return candidates


def _format_row_candidates(rows, limit=10):
    # Чаще всего несколько строк совпадают по клиенту/проекту, потому что один и тот же проект
    # учитывается отдельной строкой ФП на каждый месяц и на каждый портфель (Факт/0-100/
    # Возможности) -- поэтому именно месяц и портфель показываем, а не раздел (он почти всегда
    # одинаковый внутри одного проекта и для уточнения бесполезен).
    header = f"{'Клиент':<18} {'Проект':<22} {'№ проекта':<10} {'Месяц':<9} {'Портфель':<12} {'Сумма':>12}"
    lines = [header, "-" * len(header)]
    for r in rows[:limit]:
        lines.append(
            f"{_truncate(r['client_name'], 18):<18} {_truncate(r['project_name'], 22):<22} "
            f"{_truncate(r['project_num'] or '-', 10):<10} {_truncate(r['month'] or '-', 9):<9} "
            f"{_truncate(r['portfolio'] or '-', 12):<12} {fmt_money(r['amount_0_100']):>12}"
        )
    if len(rows) > limit:
        lines.append(f"... и ещё {len(rows) - limit}")
    return "\n".join(lines)


def _row_label(row):
    return f"{row['client_name']} / {row['project_name']}" + (f" (№ {row['project_num']})" if row["project_num"] else "")


def _exec_set_risk(conn, actor, row, risk_level, dry_run=False):
    if (actor or {}).get("role") != "owner":
        return False, "Менять отметку риска может только руководитель продукта (Owner)."
    label = RISK_LEVELS[risk_level]["label"] if risk_level else "отметка снята"
    if dry_run:
        action_phrase = f"поставить риск «{label}»" if risk_level else "снять отметку риска"
        return True, f"{action_phrase} по статье «{_row_label(row)}»"
    conn.execute(
        "UPDATE fp_rows SET risk_level = ?, is_risk = ? WHERE id = ?",
        (risk_level, 1 if risk_level else 0, row["id"]),
    )
    conn.commit()
    return True, f"Риск по статье «{_row_label(row)}»: {label}."


def _exec_add_obligation(conn, actor, row, title, responsible_name, due_date, dry_run=False):
    if not (actor or {}).get("user_id"):
        return False, "Действие доступно только пользователю с привязанным аккаунтом портала."
    title = (title or "").strip()
    if not title:
        return False, "Не понял, какое обязательство добавить — нужен текст обязательства."
    resolved_name = _match_responsible_name(responsible_name)
    if not resolved_name:
        return False, f"Не понял, кто ответственный («{responsible_name}»). Варианты: {', '.join(ALL_RESPONSIBLE)}."
    if row["section"] in TEAM_LEAD_SECTIONS and resolved_name not in TEAM_LEADS:
        return False, "Для этого раздела ответственным может быть только один из руководителей команд."
    if due_date and not re.match(r"^\d{4}-\d{2}-\d{2}$", str(due_date)):
        due_date = None
    due_part = f", срок {due_date}" if due_date else ""
    if dry_run:
        return True, f"добавить обязательство «{title}» по статье «{_row_label(row)}», ответственный: {resolved_name}{due_part}"
    responsible_type = "owner" if resolved_name == OWNER_LABEL else "team_lead"
    conn.execute("""
        INSERT INTO obligations (fp_row_id, title, description, responsible_type, responsible_name, due_date, created_by)
        VALUES (?,?,?,?,?,?,?)
    """, (row["id"], title, "", responsible_type, resolved_name, due_date, actor.get("user_id")))
    obl_id = conn.execute("SELECT last_insert_rowid() id").fetchone()["id"]
    conn.execute(
        "INSERT INTO obligation_history (obligation_id, user_id, action, details) VALUES (?,?,?,?)",
        (obl_id, actor.get("user_id"), "created", title),
    )
    conn.commit()
    return True, f"Добавлено обязательство «{title}» по статье «{_row_label(row)}», ответственный: {resolved_name}{due_part}."


def _exec_set_obligation_status(conn, actor, row, title_hint, status, dry_run=False):
    status = _normalize_obligation_status(status)
    if not status:
        return False, "Не понял, какой статус нужно поставить."
    obls = conn.execute("SELECT * FROM obligations WHERE fp_row_id = ?", (row["id"],)).fetchall()
    if not obls:
        return False, f"По статье «{_row_label(row)}» нет обязательств."
    if title_hint:
        th = title_hint.strip().lower()
        narrowed = [o for o in obls if th in (o["title"] or "").lower()]
        if narrowed:
            obls = narrowed
    if len(obls) > 1:
        listing = "\n".join(f"- {o['title']} ({OBLIGATION_STATUS_LABELS.get(o['status'], o['status'])})" for o in obls)
        return False, f"По статье «{_row_label(row)}» несколько обязательств, уточните, какое именно:\n{listing}"
    obl = obls[0]
    role = (actor or {}).get("role")
    can_edit = role == "owner" or (role == "team_lead" and obl["responsible_name"] == (actor or {}).get("team_lead_name"))
    if not can_edit:
        return False, "Нет прав менять статус этого обязательства — вы не ответственный."
    if dry_run:
        return True, f"поставить статус «{OBLIGATION_STATUS_LABELS.get(status, status)}» обязательству «{obl['title']}» по статье «{_row_label(row)}»"
    completed_at = "CURRENT_TIMESTAMP" if status == "done" else "NULL"
    conn.execute(
        f"UPDATE obligations SET status = ?, updated_at = CURRENT_TIMESTAMP, completed_at = {completed_at} WHERE id = ?",
        (status, obl["id"]),
    )
    conn.execute(
        "INSERT INTO obligation_history (obligation_id, user_id, action, details) VALUES (?,?,?,?)",
        (obl["id"], (actor or {}).get("user_id"), "status_changed", status),
    )
    conn.commit()
    return True, f"Статус обязательства «{obl['title']}»: {OBLIGATION_STATUS_LABELS.get(status, status)}."


def _exec_set_comment(conn, actor, row, comment, dry_run=False):
    if (actor or {}).get("role") != "owner":
        return False, "Менять внутренний комментарий может только руководитель продукта (Owner)."
    comment = (comment or "").strip()
    if dry_run:
        if not comment:
            return True, f"очистить комментарий по статье «{_row_label(row)}»"
        preview = comment[:60] + ("…" if len(comment) > 60 else "")
        return True, f"изменить комментарий по статье «{_row_label(row)}» на «{preview}»"
    conn.execute("UPDATE fp_rows SET internal_comment = ? WHERE id = ?", (comment or None, row["id"]))
    conn.commit()
    return True, f"Комментарий по статье «{_row_label(row)}» обновлён."


def _detect_chat_action(question):
    """Отдельный строгий JSON-запрос к Ollama: действие это или обычный вопрос. При любой
    проблеме (Ollama недоступна, модель ответила не JSON-ом и т.п.) считаем, что это не
    действие, и чат идёт по обычной ветке вопрос-ответ -- этот шаг никогда не должен сам
    становиться причиной, что чат не отвечает. Таймаут короче обычного (ACTION_DETECT_TIMEOUT),
    чтобы зависшая Ollama не удваивала время ожидания ответа на обычные вопросы."""
    try:
        raw = _call_ollama([
            {"role": "system", "content": ACTION_SYSTEM_PROMPT},
            {"role": "user", "content": question},
        ], timeout=ACTION_DETECT_TIMEOUT)
    except Exception:
        return None
    raw = re.sub(r"^```(?:json)?|```$", "", raw.strip(), flags=re.IGNORECASE | re.MULTILINE).strip()
    try:
        data = json.loads(raw)
    except (ValueError, TypeError):
        return None
    if not isinstance(data, dict) or data.get("action") in (None, "none"):
        return None
    return data


CONFIRM_NO_WORDS = ("нет", "отмен", "не надо", "не нужно", "стоп", "cancel", "no")
CONFIRM_YES_WORDS = ("да", "подтвер", "ага", "угу", "ок", "окей", "выполн", "сделай", "согласен", "верно", "норм", "go", "yes")


def _is_confirmation_reply(text):
    """Пытается понять, является ли сообщение ответом «да»/«нет» на предложенное ранее
    действие. Возвращает True/False, либо None, если сообщение не похоже ни на то, ни на
    другое (тогда вызывающий код считает, что пользователь передумал и пишет что-то новое)."""
    t = (text or "").strip().lower().rstrip(".!")
    if not t:
        return None
    for w in CONFIRM_NO_WORDS:
        if t == w or t.startswith(w):
            return False
    for w in CONFIRM_YES_WORDS:
        if t == w or t.startswith(w):
            return True
    return None


def _run_chat_action(conn, actor, row, action, data, dry_run=False):
    """Общая точка входа к исполнителям действий -- одна и та же функция используется и для
    «черновой» проверки (dry_run=True, ничего не пишет в БД, только проверяет права и
    параметры и возвращает текстовое описание, что будет сделано), и для реального
    выполнения после подтверждения (dry_run=False)."""
    if action == "set_risk":
        return _exec_set_risk(conn, actor, row, _normalize_risk_level(data.get("risk_level")), dry_run=dry_run)
    if action == "add_obligation":
        return _exec_add_obligation(
            conn, actor, row, data.get("title"), data.get("responsible_name"), data.get("due_date"), dry_run=dry_run,
        )
    if action == "set_obligation_status":
        return _exec_set_obligation_status(
            conn, actor, row, data.get("title_hint") or "", data.get("status"), dry_run=dry_run,
        )
    return _exec_set_comment(conn, actor, row, data.get("comment") or "", dry_run=dry_run)


def _apply_pending_action(pending, actor):
    """Выполняет ранее предложенное и подтверждённое действие. Строку ФП и права перепроверяем
    заново (а не доверяем тому, что было на момент предложения) -- на случай, если между
    предложением и подтверждением что-то успело измениться."""
    conn = get_conn()
    try:
        row = conn.execute("SELECT * FROM fp_rows WHERE id = ?", (pending["row_id"],)).fetchone()
        if not row:
            return "⚠ Статья ФП больше не найдена (возможно, изменилась выгрузка)."
        ok, msg = _run_chat_action(conn, actor, row, pending["action"], pending["data"], dry_run=False)
        return ("✅ " if ok else "⚠ ") + msg
    finally:
        conn.close()


def _try_handle_chat_action(question, actor, pending_action):
    """Обрабатывает либо подтверждение/отмену ранее предложенного действия, либо новый запрос
    на действие. Ничего не пишет в БД сразу -- сначала проверяет права и параметры и
    предлагает пользователю точную формулировку того, что будет сделано, и только после
    явного «да» в следующем сообщении выполняет его.

    Возвращает (answer_text, new_pending_action):
    - answer_text is None, если это обычный вопрос (вызывающий код идёт дальше как раньше);
    - new_pending_action нужно сохранить (сессия на сайте / память бота по chat_id) и передать
      в следующий вызов этой же функции, иначе подтверждение «да» не к чему будет привязать."""
    q = (question or "").strip()

    if pending_action:
        confirm = _is_confirmation_reply(q)
        if confirm is True:
            return _apply_pending_action(pending_action, actor), None
        if confirm is False:
            return "Отменено, изменений не вносил.", None
        # Сообщение не похоже ни на «да», ни на «нет» -- считаем, что пользователь передумал
        # и пишет что-то новое; старое предложение просто сгорает, дальше обрабатываем как
        # обычное сообщение (вопрос или новое действие).
        pending_action = None

    data = _detect_chat_action(q)
    if not data:
        return None, None

    action = data.get("action")
    if action not in ("set_risk", "add_obligation", "set_obligation_status", "set_comment"):
        return None, None

    if not actor:
        return ("Чтобы выполнять действия через чат (а не только спрашивать), нужен привязанный "
                "аккаунт портала. Привяжите его командой /link КОД."), None

    conn = get_conn()
    try:
        row_query = (data.get("row_query") or "").strip()
        rows = _resolve_rows_for_action(conn, row_query) if row_query else []
        if not rows:
            return f"⚠ Не нашёл статью ФП по запросу «{row_query}». Уточните клиента или номер договора/проекта.", None
        if len(rows) > 1:
            return (
                f"⚠ По запросу «{row_query}» нашлось {len(rows)} статей, уточните, какая нужна "
                f"(добавьте номер договора/проекта, раздел или месяц):\n\n" + _format_row_candidates(rows)
            ), None
        row = rows[0]

        ok, msg = _run_chat_action(conn, actor, row, action, data, dry_run=True)
        if not ok:
            return "⚠ " + msg, None

        pending = {"action": action, "row_id": row["id"], "data": data}
        return f"Готов выполнить: {msg}.\n\nПодтвердите: «да» — сделаю, «нет» — отменю.", pending
    finally:
        conn.close()


def answer_portal_query(question, actor=None, pending_action=None):
    """Отвечает на вопрос по данным портала через локальную LLM в Ollama: собираем релевантный
    контекст из БД (см. _build_chat_context) и просим модель ответить по нему обычным русским
    языком. Если Ollama недоступна (не запущена, модель не скачана, таймаут) -- откатываемся
    на локальный поиск по ключевым словам (_answer_portal_query_local), чтобы чат не падал
    целиком, и явно сообщаем об этом в ответе.

    actor -- словарь с данными пользователя, который спрашивает (role, team_lead_name,
    user_id, full_name), нужен только для проверки прав при действиях через чат (см.
    _try_handle_chat_action ниже). Если actor=None -- чат отвечает на вопросы, но не
    выполняет действия (так и должно быть для неизвестного/непривязанного пользователя).

    pending_action -- ранее предложенное, но ещё не подтверждённое действие (см.
    _try_handle_chat_action); вызывающий код должен сохранять его между сообщениями (Flask
    session на сайте, словарь по chat_id в боте) и передавать сюда на каждый следующий вызов.

    Возвращает (answer_text, new_pending_action) -- new_pending_action нужно сохранить и
    передать в следующий вызов вместо pending_action (или удалить сохранённое, если None)."""
    q = (question or "").strip()
    if not q or _any_word(q, HELP_WORDS):
        return HELP_TEXT, None

    action_answer, new_pending = _try_handle_chat_action(q, actor, pending_action)
    if action_answer is not None:
        return action_answer, new_pending

    conn = get_conn()
    try:
        context = _build_chat_context(conn, q)
    finally:
        conn.close()

    system_prompt = (
        "Ты — помощник внутреннего портала «ФП-Контроль» (финансовое планирование компании). "
        "Отвечай на вопросы пользователя только на основе данных из раздела «Данные портала» "
        "ниже. Не выдумывай цифры и факты, которых там нет — если данных не хватает, прямо "
        "скажи об этом, не утверждай, что статьи вообще нет в системе, если явно не видишь "
        "подтверждения этому. Если есть блок «Найдено по коду из вопроса» — это точное "
        "совпадение по номеру договора/проекта/ЕЗ, используй его как основной источник ответа. "
        "Учти: в данных два разных понятия риска — «риск (вручную)» (поле, которое сам "
        "проставляет пользователь: Низкий/Средний/Высокий) и статьи «без оформленного "
        "обязательства» (автоматический список по сумме ≥300 000 ₽ без ответственного) — это "
        "не одно и то же, не путай их. Отвечай по-русски, кратко и по существу, обычным "
        "связным языком (не нужно копировать таблицы целиком — пересказывай суть и ключевые "
        "цифры)."
    )
    user_prompt = f"Данные портала:\n{context}\n\nВопрос пользователя: {q}"

    try:
        return _call_ollama([
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ]), None
    except Exception:
        fallback = _answer_portal_query_local(q)
        return (
            "⚠ Локальная модель (Ollama) сейчас недоступна — показываю результат обычного "
            "поиска по ключевым словам:\n\n" + fallback
        ), None


def get_default_view_quarter(conn):
    """Квартал по умолчанию: последний квартал с активными строками.
    Если строк нет — возвращает current_quarter_label из настроек или вычисляет по дате."""
    row = conn.execute("""
        SELECT quarter_label FROM fp_rows WHERE is_active=1 AND quarter_label IS NOT NULL
        ORDER BY quarter_label DESC LIMIT 1
    """).fetchone()
    if row:
        return row["quarter_label"]
    return get_setting(conn, "current_quarter_label") or compute_quarter_label(datetime.date.today())


def get_view_context(conn):
    """Возвращает (quarter_labels: list, mode: str) для текущего сеанса.

    mode='quarter' — смотрим один квартал (quarter_labels = ['2026-Q2'])
    mode='fy'      — смотрим весь финансовый год (quarter_labels = ['2026-Q2',...,'2027-Q1'])
    """
    mode = session.get("view_mode", "quarter")
    if mode == "fy":
        fy_key = session.get("view_fy", "")
        if fy_key and fy_key.startswith("FY"):
            return fiscal_year_quarters(fy_key), "fy"
    # quarter mode (default)
    vq = session.get("view_quarter")
    if not vq:
        vq = get_default_view_quarter(conn)
    return [vq], "quarter"


def get_view_quarter(conn):
    """Обратная совместимость: возвращает одиночный квартал (первый из списка)."""
    vqls, _ = get_view_context(conn)
    return vqls[0] if vqls else get_default_view_quarter(conn)


def get_period_info(quarter_labels=None, conn=None):
    """Определяет рабочий период по месяцам, реально присутствующим в активных строках.
    quarter_labels: str или list — фильтр по кварталу(ам). None = все активные строки.
    conn: опциональная открытая коннекция — используется как есть, без закрытия.
    Возвращает (months, qnum, outlier_months)."""
    _own_conn = conn is None
    if _own_conn:
        conn = get_conn()
    if quarter_labels:
        if isinstance(quarter_labels, str):
            quarter_labels = [quarter_labels]
        ph = ",".join("?" * len(quarter_labels))
        rows = conn.execute(
            f"SELECT DISTINCT month FROM fp_rows WHERE is_active = 1 AND quarter_label IN ({ph})",
            quarter_labels,
        ).fetchall()
    else:
        rows = conn.execute("SELECT DISTINCT month FROM fp_rows WHERE is_active = 1").fetchall()
    if _own_conn:
        conn.close()
    months_in_data = [r["month"] for r in rows if r["month"] in MONTH_INDEX]
    months_in_data.sort(key=lambda m: MONTH_INDEX[m])

    if not months_in_data:
        months, qnum = QUARTER_MONTHS[(datetime.date.today().month - 1) // 3 + 1], None
        return months, qnum, set()

    # "Основной" квартал — тот, в который попадает большинство присутствующих месяцев
    quarter_votes = Counter()
    for m in months_in_data:
        idx = MONTH_INDEX[m]
        q = idx // 3 + 1
        quarter_votes[q] += 1
    main_q = quarter_votes.most_common(1)[0][0]
    main_q_months = set(QUARTER_MONTHS[main_q])

    outlier_months = set(m for m in months_in_data if m not in main_q_months)

    return months_in_data, main_q, outlier_months


# ---------- Auth helpers ----------

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login", next=request.path))
        return f(*args, **kwargs)
    return wrapper


def owner_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if session.get("role") != "owner":
            flash("Недостаточно прав для этого действия", "error")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return wrapper


def _get_portal_users(conn):
    """Возвращает (all_names, tl_names, name_to_role) из таблицы users.
    Используется вместо захардкоженных TEAM_LEADS / ALL_RESPONSIBLE."""
    rows = conn.execute(
        "SELECT full_name, role FROM users WHERE role IN ('owner','team_lead') ORDER BY role DESC, full_name"
    ).fetchall()
    all_names = [r["full_name"] for r in rows]
    tl_names = [r["full_name"] for r in rows if r["role"] == "team_lead"]
    name_to_role = {r["full_name"]: r["role"] for r in rows}
    return all_names, tl_names, name_to_role


def can_edit_obligation(obligation_row):
    role = session.get("role")
    if role == "owner":
        return True
    if role == "team_lead" and obligation_row["responsible_name"] == session.get("team_lead_name"):
        return True
    return False


@app.context_processor
def inject_user():
    pending_changes = {"new": 0, "closed": 0, "other": 0}
    available_quarters = []
    available_fiscal_years = []
    view_quarter = None
    view_mode = "quarter"
    view_fy = ""
    _all_resp = [OWNER_LABEL]
    _tl_names = []
    if session.get("user_id"):
        conn = get_conn()
        if session.get("role") == "owner":
            for key, types in TAB_EVENT_TYPES.items():
                ph = ",".join("?" * len(types))
                pending_changes[key] = conn.execute(
                    f"SELECT COUNT(*) c FROM row_events WHERE event_type IN ({ph}) AND reviewed_at IS NULL",
                    list(types),
                ).fetchone()["c"]
        available_quarters = get_available_quarters(conn)
        available_fiscal_years = get_available_fiscal_years(conn)
        vqls, view_mode = get_view_context(conn)
        view_quarter = vqls[0] if vqls else None
        view_fy = session.get("view_fy", "")
        _all_resp, _tl_names, _ = _get_portal_users(conn)
        conn.close()
    fin_labels = {q["label"] for q in available_quarters if q["is_finalized"]}

    # Для дашборда: финансовый квартал и год текущего view_quarter
    view_fq_num = None
    view_fy_label = ""
    if view_quarter:
        try:
            _fk, view_fy_label, view_fq_num = calendar_to_fiscal(view_quarter)
        except Exception:
            pass

    return dict(
        current_user_name=session.get("full_name"),
        current_user_role=session.get("role"),
        team_leads=_tl_names,
        all_responsible=_all_resp,
        pending_changes=pending_changes,
        pending_changes_total=sum(pending_changes.values()),
        asset_version=ASSET_VERSION,
        available_quarters=available_quarters,
        available_fiscal_years=available_fiscal_years,
        view_quarter=view_quarter,
        view_mode=view_mode,
        view_fy=view_fy,
        view_fq_num=view_fq_num,
        view_fy_label=view_fy_label,
        view_quarter_is_finalized=view_quarter in fin_labels if view_quarter else False,
        view_quarter_next=next_quarter_label(view_quarter) if view_quarter else "",
    )


# ---------- Auth routes ----------

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "")
        conn = get_conn()
        user = conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
        conn.close()
        if user and check_password_hash(user["password_hash"], password):
            session["user_id"] = user["id"]
            session["username"] = user["username"]
            session["full_name"] = user["full_name"]
            session["role"] = user["role"]
            session["team_lead_name"] = user["team_lead_name"]
            next_url = request.args.get("next") or url_for("dashboard")
            return redirect(next_url)
        flash("Неверный логин или пароль", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ---------- Dashboard ----------

@app.route("/")
@login_required
def dashboard():
    conn = get_conn()
    vqls, _mode = get_view_context(conn)
    ql_ph = ",".join("?" * len(vqls))
    months, qnum, outlier_months = get_period_info(vqls, conn=conn)
    m_ph = ",".join("?" * len(months)) if months else "''"
    rows = conn.execute(f"""
        SELECT * FROM fp_rows
        WHERE is_active = 1 AND quarter_label IN ({ql_ph}) AND month IN ({m_ph})
    """, vqls + list(months)).fetchall()
    today = datetime.date.today()

    filter_client = request.args.get("client", "")
    filter_responsible = request.args.get("responsible", "")
    clients = conn.execute(
        f"SELECT DISTINCT client_name FROM fp_rows WHERE is_active=1 AND quarter_label IN ({ql_ph}) ORDER BY client_name",
        vqls,
    ).fetchall()
    clients = [c["client_name"] for c in clients]

    summary = {"Факт": 0.0, "0-100": 0.0, "Возможности": 0.0}
    by_section = {}
    focus_summary = {"0-100": 0.0, "Возможности": 0.0}
    support_by_client = {}
    risk_sums = {1: 0.0, 2: 0.0, 3: 0.0}
    risk_counts = {1: 0, 2: 0, 3: 0}
    for r in rows:
        p = r["portfolio"] or "Прочее"
        summary[p] = summary.get(p, 0.0) + (r["amount_0_100"] or 0)
        sec = r["section"] or "Без раздела"
        by_section.setdefault(sec, {"Факт": 0.0, "0-100": 0.0, "Возможности": 0.0})
        by_section[sec][p] = by_section[sec].get(p, 0.0) + (r["amount_0_100"] or 0)

        # Фокус работы: 0‑100 и Возможности по Проекты/Заказные доработки/Лицензии
        if sec in FOCUS_SECTIONS and p in focus_summary:
            focus_summary[p] += (r["amount_0_100"] or 0)

        # Сопровождение и TaaS — отдельный блок только для отслеживания, без ответственного
        # (Факт показываем отдельно как справочный столбец, в "Итого" по клиенту он не входит)
        if sec in TRACKING_ONLY_SECTIONS and p in ("Факт", "0-100", "Возможности"):
            if not filter_client or r["client_name"] == filter_client:
                d = support_by_client.setdefault(r["client_name"], {"Факт": 0.0, "0-100": 0.0, "Возможности": 0.0})
                d[p] += (r["amount_0_100"] or 0)

        # Сумма под риском: сумма 0‑100 по статьям с вручную выставленным уровнем риска
        # (Низкий/Средний/Высокий) — отдельная отметка, не связана со статусом ответственного.
        rl = r["risk_level"] or 0
        if rl in risk_sums:
            risk_sums[rl] += (r["amount_0_100"] or 0)
            risk_counts[rl] += 1

    section_totals = {
        "Факт": sum(v.get("Факт", 0) for v in by_section.values()),
        "0-100": sum(v.get("0-100", 0) for v in by_section.values()),
        "Возможности": sum(v.get("Возможности", 0) for v in by_section.values()),
    }

    support_by_client = dict(sorted(support_by_client.items()))
    support_totals = {
        "Факт": sum(v["Факт"] for v in support_by_client.values()),
        "0-100": sum(v["0-100"] for v in support_by_client.values()),
        "Возможности": sum(v["Возможности"] for v in support_by_client.values()),
    }

    # «К получению в ближайшие 4 дня» -- календарный разрез по дате ДПА, не привязан к месяцам
    # текущего периода (см. комментарий у DPA_DUE_WINDOW_DAYS выше).
    dpa_window_start = today
    dpa_window_end = today + datetime.timedelta(days=DPA_DUE_WINDOW_DAYS)
    dpa_due_rows = conn.execute(f"""
        SELECT * FROM fp_rows
        WHERE is_active = 1 AND quarter_label IN ({ql_ph}) AND portfolio IN ('0-100', 'Возможности')
        AND dpa_date IS NOT NULL AND dpa_date >= ? AND dpa_date <= ?
    """, vqls + [dpa_window_start.isoformat(), dpa_window_end.isoformat()]).fetchall()

    dpa_section_to_category = {s: label for label, secs in DPA_DUE_CATEGORIES for s in secs}
    dpa_due_by_category = {label: 0.0 for label, _ in DPA_DUE_CATEGORIES}
    dpa_due_counts = {label: 0 for label, _ in DPA_DUE_CATEGORIES}
    for r in dpa_due_rows:
        cat_label = dpa_section_to_category.get(r["section"])
        if cat_label is None:
            continue
        dpa_due_by_category[cat_label] += (r["amount_0_100"] or 0)
        dpa_due_counts[cat_label] += 1
    dpa_due_total = sum(dpa_due_by_category.values())
    dpa_due_count_total = sum(dpa_due_counts.values())
    dpa_due_all_sections = [s for _, secs in DPA_DUE_CATEGORIES for s in secs]

    # «Просроченные деньги» — 0-100/Возможности, у которых дата ДПА уже прошла
    dpa_overdue_rows = conn.execute(f"""
        SELECT * FROM fp_rows
        WHERE is_active = 1 AND quarter_label IN ({ql_ph}) AND portfolio IN ('0-100', 'Возможности')
        AND dpa_date IS NOT NULL AND dpa_date < ?
    """, vqls + [today.isoformat()]).fetchall()

    dpa_overdue_by_category = {label: 0.0 for label, _ in DPA_DUE_CATEGORIES}
    dpa_overdue_counts = {label: 0 for label, _ in DPA_DUE_CATEGORIES}
    for r in dpa_overdue_rows:
        cat_label = dpa_section_to_category.get(r["section"])
        if cat_label is None:
            continue
        dpa_overdue_by_category[cat_label] += (r["amount_0_100"] or 0)
        dpa_overdue_counts[cat_label] += 1
    dpa_overdue_total = sum(dpa_overdue_by_category.values())
    dpa_overdue_count_total = sum(dpa_overdue_counts.values())

    row_ids = [r["id"] for r in rows]
    obligations = []
    if row_ids:
        ph = ",".join("?" * len(row_ids))
        obligations = conn.execute(f"""
            SELECT o.*, f.client_name, f.project_name, f.section, f.amount_0_100, f.portfolio, f.dpa_date
            FROM obligations o JOIN fp_rows f ON f.id = o.fp_row_id
            WHERE o.fp_row_id IN ({ph})
        """, row_ids).fetchall()

    overdue, due_soon = [], []
    for o in obligations:
        if filter_client and o["client_name"] != filter_client:
            continue
        if filter_responsible and o["responsible_name"] != filter_responsible:
            continue
        if o["status"] == "done":
            continue
        if o["due_date"]:
            due = datetime.date.fromisoformat(o["due_date"])
            if due < today:
                overdue.append(o)
            elif (due - today).days <= 3:
                due_soon.append(o)

    tl_ph = ",".join("?" * len(TEAM_LEAD_SECTIONS))
    own_ph = ",".join("?" * len(OWNER_SECTIONS))
    risky_rows = conn.execute(f"""
        SELECT f.* FROM fp_rows f
        WHERE f.is_active = 1
        AND f.quarter_label IN ({ql_ph})
        AND f.month IN ({m_ph})
        AND (
            (f.section IN ({tl_ph}) AND NOT EXISTS (
                SELECT 1 FROM obligations o WHERE o.fp_row_id = f.id AND o.responsible_type = 'team_lead'
            ))
            OR
            (f.section IN ({own_ph}) AND NOT EXISTS (
                SELECT 1 FROM obligations o WHERE o.fp_row_id = f.id AND o.responsible_type = 'owner'
            ))
        )
    """, vqls + list(months) + list(TEAM_LEAD_SECTIONS) + list(OWNER_SECTIONS)).fetchall()
    if filter_client:
        risky_rows = [r for r in risky_rows if r["client_name"] == filter_client]
    # Мелкие суммы не показываем -- чтобы не перегружать список статьями, которые не критичны
    risky_rows = [r for r in risky_rows if (r["amount_0_100"] or 0) >= 300000]

    period_label = "-".join(months) if months else "период"
    target_row = conn.execute(
        "SELECT * FROM quarter_targets WHERE period_label = ?", (period_label,)
    ).fetchone()
    target_amount = target_row["target_amount"] if target_row else None

    # Ручная корректировка факта (последний день периода без нового файла).
    # Суммируем по всем кварталам текущего вида (обычно 1, но при просмотре ФГ — до 4).
    # Корректировка прибавляется к Факту ДО вычисления производных показателей.
    fact_correction = 0.0
    fact_correction_notes = {}  # {quarter_label: note}
    for ql in vqls:
        corr_str = get_setting(conn, f"fact_correction:{ql}")
        if corr_str:
            try:
                fact_correction += float(corr_str)
            except (ValueError, TypeError):
                pass
        note_str = get_setting(conn, f"fact_correction_note:{ql}") or ""
        if note_str:
            fact_correction_notes[ql] = note_str
    # Применяем корректировку к отображаемой сумме Факта
    summary["Факт"] = summary.get("Факт", 0.0) + fact_correction
    # Текущая корректировка для одного квартала (для формы редактирования)
    current_q_correction = 0.0
    current_q_correction_note = ""
    if len(vqls) == 1:
        try:
            current_q_correction = float(get_setting(conn, f"fact_correction:{vqls[0]}") or "0")
        except (ValueError, TypeError):
            pass
        current_q_correction_note = get_setting(conn, f"fact_correction_note:{vqls[0]}") or ""

    achieved = summary.get("Факт", 0) + summary.get("0-100", 0) + summary.get("Возможности", 0)
    target_pct = (achieved / target_amount * 100) if target_amount else None

    gap_after_fact = (target_amount - summary.get("Факт", 0)) if target_amount else None
    gap_after_0100 = (target_amount - summary.get("Факт", 0) - summary.get("0-100", 0)) if target_amount else None
    gap_after_opportunities = (target_amount - achieved) if target_amount else None

    conn.close()
    return render_template(
        "dashboard.html",
        summary=summary, by_section=by_section, section_totals=section_totals,
        overdue=overdue, due_soon=due_soon,
        risky_rows=risky_rows, qnum=qnum, months=months, outlier_months=outlier_months,
        total_rows=len(rows), period_label=period_label,
        target_amount=target_amount, target_pct=target_pct, achieved=achieved,
        gap_after_fact=gap_after_fact, gap_after_0100=gap_after_0100, gap_after_opportunities=gap_after_opportunities,
        clients=clients, filter_client=filter_client, filter_responsible=filter_responsible,
        team_lead_sections=TEAM_LEAD_SECTIONS, owner_sections=OWNER_SECTIONS,
        focus_summary=focus_summary, support_by_client=support_by_client, support_totals=support_totals,
        focus_sections=list(FOCUS_SECTIONS), tracking_only_sections=list(TRACKING_ONLY_SECTIONS),
        risk_sums=risk_sums, risk_counts=risk_counts, risk_levels=RISK_LEVELS,
        fact_correction=fact_correction, fact_correction_notes=fact_correction_notes,
        current_q_correction=current_q_correction, current_q_correction_note=current_q_correction_note,
        correction_quarter=vqls[0] if len(vqls) == 1 else "",
        dpa_due_categories=DPA_DUE_CATEGORIES, dpa_due_by_category=dpa_due_by_category,
        dpa_due_counts=dpa_due_counts, dpa_due_total=dpa_due_total, dpa_due_count_total=dpa_due_count_total,
        dpa_due_all_sections=dpa_due_all_sections, dpa_window_start=dpa_window_start.isoformat(),
        dpa_window_end=dpa_window_end.isoformat(), dpa_due_window_days=DPA_DUE_WINDOW_DAYS,
        dpa_overdue_by_category=dpa_overdue_by_category, dpa_overdue_counts=dpa_overdue_counts,
        dpa_overdue_total=dpa_overdue_total, dpa_overdue_count_total=dpa_overdue_count_total,
        today_iso=today.isoformat(),
    )


# ── Заказные доработки: отдельный срез дашборда ──────────────────────────────
@app.route("/zd")
@login_required
def zd_dashboard():
    ZD_SECTION = "Заказные доработки"
    conn = get_conn()
    vqls, _mode = get_view_context(conn)
    ql_ph = ",".join("?" * len(vqls))
    months, qnum, outlier_months = get_period_info(vqls, conn=conn)
    m_ph = ",".join("?" * len(months)) if months else "''"

    rows = conn.execute(f"""
        SELECT * FROM fp_rows
        WHERE is_active = 1 AND quarter_label IN ({ql_ph}) AND month IN ({m_ph})
        AND section = ?
    """, vqls + list(months) + [ZD_SECTION]).fetchall()
    today = datetime.date.today()

    filter_client = request.args.get("client", "")
    filter_responsible = request.args.get("responsible", "")
    clients_q = conn.execute(
        f"SELECT DISTINCT client_name FROM fp_rows WHERE is_active=1 AND quarter_label IN ({ql_ph}) AND section = ? ORDER BY client_name",
        vqls + [ZD_SECTION],
    ).fetchall()
    clients = [c["client_name"] for c in clients_q]

    summary = {"Факт": 0.0, "0-100": 0.0, "Возможности": 0.0}
    by_client = {}
    risk_sums = {1: 0.0, 2: 0.0, 3: 0.0}
    risk_counts = {1: 0, 2: 0, 3: 0}
    for r in rows:
        p = r["portfolio"] or "Прочее"
        summary[p] = summary.get(p, 0.0) + (r["amount_0_100"] or 0)
        cli = r["client_name"]
        by_client.setdefault(cli, {"Факт": 0.0, "0-100": 0.0, "Возможности": 0.0})
        by_client[cli][p] = by_client[cli].get(p, 0.0) + (r["amount_0_100"] or 0)
        rl = r["risk_level"] or 0
        if rl in risk_sums:
            risk_sums[rl] += (r["amount_0_100"] or 0)
            risk_counts[rl] += 1

    by_client_sorted = dict(sorted(by_client.items()))
    client_totals = {
        "Факт": sum(v.get("Факт", 0) for v in by_client.values()),
        "0-100": sum(v.get("0-100", 0) for v in by_client.values()),
        "Возможности": sum(v.get("Возможности", 0) for v in by_client.values()),
    }

    # DPA — к получению в ближайшие N дней
    dpa_window_start = today
    dpa_window_end = today + datetime.timedelta(days=DPA_DUE_WINDOW_DAYS)
    dpa_due_rows = conn.execute(f"""
        SELECT * FROM fp_rows
        WHERE is_active = 1 AND quarter_label IN ({ql_ph}) AND portfolio IN ('0-100', 'Возможности')
        AND dpa_date IS NOT NULL AND dpa_date >= ? AND dpa_date <= ?
        AND section = ?
    """, vqls + [dpa_window_start.isoformat(), dpa_window_end.isoformat(), ZD_SECTION]).fetchall()
    dpa_due_total = sum((r["amount_0_100"] or 0) for r in dpa_due_rows)
    dpa_due_count_total = len(dpa_due_rows)

    # DPA overdue
    dpa_overdue_rows = conn.execute(f"""
        SELECT * FROM fp_rows
        WHERE is_active = 1 AND quarter_label IN ({ql_ph}) AND portfolio IN ('0-100', 'Возможности')
        AND dpa_date IS NOT NULL AND dpa_date < ?
        AND section = ?
    """, vqls + [today.isoformat(), ZD_SECTION]).fetchall()
    dpa_overdue_total = sum((r["amount_0_100"] or 0) for r in dpa_overdue_rows)
    dpa_overdue_count_total = len(dpa_overdue_rows)

    # Obligations
    row_ids = [r["id"] for r in rows]
    obligations_list = []
    if row_ids:
        ph = ",".join("?" * len(row_ids))
        obligations_list = conn.execute(f"""
            SELECT o.*, f.client_name, f.project_name, f.section, f.amount_0_100, f.portfolio, f.dpa_date
            FROM obligations o JOIN fp_rows f ON f.id = o.fp_row_id
            WHERE o.fp_row_id IN ({ph})
        """, row_ids).fetchall()

    overdue, due_soon = [], []
    for o in obligations_list:
        if filter_client and o["client_name"] != filter_client:
            continue
        if filter_responsible and o["responsible_name"] != filter_responsible:
            continue
        if o["status"] == "done":
            continue
        if o["due_date"]:
            due = datetime.date.fromisoformat(o["due_date"])
            if due < today:
                overdue.append(o)
            elif (due - today).days <= 3:
                due_soon.append(o)

    # Статьи без обязательства (ЗД требует team_lead)
    risky_rows = conn.execute(f"""
        SELECT f.* FROM fp_rows f
        WHERE f.is_active = 1
        AND f.quarter_label IN ({ql_ph})
        AND f.month IN ({m_ph})
        AND f.section = ?
        AND NOT EXISTS (
            SELECT 1 FROM obligations o WHERE o.fp_row_id = f.id AND o.responsible_type = 'team_lead'
        )
    """, vqls + list(months) + [ZD_SECTION]).fetchall()
    if filter_client:
        risky_rows = [r for r in risky_rows if r["client_name"] == filter_client]
    risky_rows = [r for r in risky_rows if (r["amount_0_100"] or 0) >= 300000]

    conn.close()
    return render_template(
        "zd.html",
        summary=summary, by_client=by_client_sorted, client_totals=client_totals,
        overdue=overdue, due_soon=due_soon,
        risky_rows=risky_rows, qnum=qnum, months=months, outlier_months=outlier_months,
        total_rows=len(rows),
        clients=clients, filter_client=filter_client, filter_responsible=filter_responsible,
        team_lead_sections=TEAM_LEAD_SECTIONS,
        risk_sums=risk_sums, risk_counts=risk_counts, risk_levels=RISK_LEVELS,
        dpa_due_rows=dpa_due_rows, dpa_due_total=dpa_due_total, dpa_due_count_total=dpa_due_count_total,
        dpa_window_start=dpa_window_start.isoformat(), dpa_window_end=dpa_window_end.isoformat(),
        dpa_due_window_days=DPA_DUE_WINDOW_DAYS,
        dpa_overdue_rows=dpa_overdue_rows, dpa_overdue_total=dpa_overdue_total,
        dpa_overdue_count_total=dpa_overdue_count_total,
        today_iso=today.isoformat(),
    )


@app.route("/save-fact-correction", methods=["POST"])
@login_required
@owner_required
def save_fact_correction():
    """Сохраняет ручную корректировку итогового факта для квартала.
    Используется, когда на последний день периода факт изменился, но новый файл
    ещё недоступен для загрузки. Корректировка хранится в app_settings и прибавляется
    к сумме Факт при отображении.
    """
    q = request.form.get("quarter_label", "").strip()
    if not q or "-Q" not in q:
        flash("Не указан квартал для корректировки", "error")
        return redirect(url_for("dashboard"))
    amount_str = request.form.get("correction_amount", "0").strip().replace(" ", "").replace(",", ".")
    try:
        amount = float(amount_str)
    except ValueError:
        flash("Некорректная сумма корректировки", "error")
        return redirect(url_for("dashboard"))
    note = request.form.get("correction_note", "").strip()
    conn = get_conn()
    set_setting(conn, f"fact_correction:{q}", str(amount))
    set_setting(conn, f"fact_correction_note:{q}", note)
    conn.commit()
    conn.close()
    sign = "+" if amount >= 0 else ""
    flash(f"Корректировка факта {q}: {sign}{amount:,.0f} ₽ сохранена.", "success")
    return redirect(url_for("dashboard"))


@app.route("/target", methods=["POST"])
@login_required
@owner_required
def set_target():
    conn = get_conn()
    vqls, _ = get_view_context(conn)
    conn.close()
    months, qnum, outlier_months = get_period_info(vqls)
    period_label = "-".join(months) if months else "период"
    amount = request.form.get("target_amount", "").strip().replace(" ", "").replace(",", ".")
    try:
        amount_val = float(amount)
    except ValueError:
        flash("Введите корректную сумму", "error")
        return redirect(url_for("dashboard"))

    conn.execute("""
        INSERT INTO quarter_targets (period_label, target_amount, set_by, updated_at)
        VALUES (?,?,?,CURRENT_TIMESTAMP)
        ON CONFLICT(period_label) DO UPDATE SET target_amount=excluded.target_amount,
            set_by=excluded.set_by, updated_at=CURRENT_TIMESTAMP
    """, (period_label, amount_val, session["user_id"]))
    conn.commit()
    conn.close()
    flash("Целевая сумма квартала обновлена", "success")
    return redirect(url_for("dashboard"))


@app.route("/chat", methods=["POST"])
@login_required
def chat_query():
    question = (request.form.get("question") or "").strip()
    if not question:
        return jsonify({"error": "Введите вопрос"}), 400

    actor = {
        "role": session.get("role"),
        "team_lead_name": session.get("team_lead_name"),
        "user_id": session.get("user_id"),
        "full_name": session.get("full_name"),
    }
    try:
        answer, pending = answer_portal_query(question, actor=actor, pending_action=session.get("pending_chat_action"))
    except Exception as e:
        return jsonify({"error": f"Ошибка поиска: {e}"}), 500

    if pending:
        session["pending_chat_action"] = pending
    else:
        session.pop("pending_chat_action", None)

    return jsonify({"answer": answer})


@app.route("/chat/voice", methods=["POST"])
@login_required
def chat_voice_query():
    """Голосовой вопрос с дашборда: браузер записывает аудио (MediaRecorder) и шлёт сюда
    multipart-файлом. Распознаём локально (app/stt.py, Whisper) и дальше прогоняем ровно
    через тот же answer_portal_query, что и обычный текстовый /chat -- с тем же
    pending_action в сессии, так что подтверждение действий «да»/«нет» работает одинаково
    независимо от того, текстом или голосом задан вопрос."""
    audio = request.files.get("audio")
    if not audio:
        return jsonify({"error": "Не получил аудио"}), 400

    suffix = Path(secure_filename(audio.filename or "voice.webm")).suffix or ".webm"
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            audio.save(tmp)
            tmp_path = tmp.name
        question = transcribe(tmp_path)
    except SttError as e:
        return jsonify({"error": str(e)}), 503
    except Exception as e:
        return jsonify({"error": f"Ошибка обработки аудио: {e}"}), 500
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.remove(tmp_path)

    actor = {
        "role": session.get("role"),
        "team_lead_name": session.get("team_lead_name"),
        "user_id": session.get("user_id"),
        "full_name": session.get("full_name"),
    }
    try:
        answer, pending = answer_portal_query(question, actor=actor, pending_action=session.get("pending_chat_action"))
    except Exception as e:
        return jsonify({"error": f"Ошибка поиска: {e}"}), 500

    if pending:
        session["pending_chat_action"] = pending
    else:
        session.pop("pending_chat_action", None)

    return jsonify({"question": question, "answer": answer})


# ---------- Import ----------

@app.route("/import", methods=["GET", "POST"])
@login_required
@owner_required
def import_view():
    if request.method == "POST":
        file = request.files.get("file")
        if not file or not file.filename.endswith(".xlsx"):
            flash("Пожалуйста, выберите файл .xlsx", "error")
            return redirect(url_for("import_view"))
        upload_quarter = request.form.get("upload_quarter", "").strip()
        fname = secure_filename(file.filename)
        path = UPLOAD_DIR / fname
        file.save(path)
        try:
            _c = get_conn(); _fp_rp = get_setting(_c, "fp_rp_filter") or None; _c.close()
            result = import_excel(str(path), fname, session["user_id"],
                                  quarter_label=upload_quarter or None,
                                  rp_filter=_fp_rp or None)
        except Exception as e:
            flash(f"Ошибка импорта: {e}", "error")
            return redirect(url_for("import_view"))

        effective_quarter = result.get("quarter_label", upload_quarter)
        if effective_quarter:
            session["view_quarter"] = effective_quarter

        msg = (f"Импорт завершён ({effective_quarter or '?'}): "
               f"всего строк {result['rows_total']}, новых {result['rows_new']}, "
               f"изменено {result['rows_updated']}, без изменений {result['rows_unchanged']}, "
               f"деактивировано {result['rows_deactivated']}")
        if result["header_mismatches"]:
            msg += f". ВНИМАНИЕ: несовпадение заголовков: {result['header_mismatches']}"
        flash(msg, "success" if not result["header_mismatches"] else "warning")

        ro = result.get("rollover")
        if ro:
            flash(
                f"Первая загрузка {ro['new_label']}: перенос из {ro['old_label']} завершён. "
                f"По {ro['matched']} из {ro['candidates']} статей (0‑100 / Возможности) "
                f"перенесены обязательства ({ro['obligations_copied']}) и примечания "
                f"({ro['comments_copied']}). "
                f"Уровень риска не переносится — нужно оценить заново.",
                "success",
            )
        return redirect(url_for("import_view", last=1))

    conn = get_conn()
    history = conn.execute("""
        SELECT il.*, u.full_name FROM import_log il LEFT JOIN users u ON u.id = il.imported_by
        ORDER BY il.imported_at DESC LIMIT 20
    """).fetchall()

    last_diff = None
    if request.args.get("last") and history:
        try:
            last_diff = json.loads(history[0]["diff_json"] or "{}")
        except (json.JSONDecodeError, TypeError):
            last_diff = None

    current_q = get_setting(conn, "current_quarter_label") or compute_quarter_label(datetime.date.today())
    next_q = next_quarter_label(current_q)
    q_options = quarter_options(current_q)
    fp_rp_filter  = get_setting(conn, "fp_rp_filter") or ""
    ts_rp_filter = get_setting(conn, "ts_rp_filter") or ""
    ts_pc_filter = get_setting(conn, "ts_pc_filter") or ""
    smb_server    = get_setting(conn, "smb_server") or ""
    smb_share     = get_setting(conn, "smb_share") or ""
    smb_subfolder = get_setting(conn, "smb_subfolder") or ""
    smb_username  = get_setting(conn, "smb_username") or ""
    smb_has_password = bool(get_setting(conn, "smb_password"))
    smb_filename  = get_setting(conn, "smb_filename") or ""
    conn.close()
    return render_template(
        "import.html",
        history=history, last_diff=last_diff,
        current_quarter_label=current_q,
        next_quarter_label=next_q,
        quarter_option_list=q_options,
        ts_rp_filter=ts_rp_filter,
        ts_pc_filter=ts_pc_filter,
        fp_rp_filter=fp_rp_filter,
        smb_server=smb_server, smb_share=smb_share, smb_subfolder=smb_subfolder,
        smb_username=smb_username, smb_has_password=smb_has_password,
        smb_filename=smb_filename,
    )


@app.route("/import/<int:log_id>/diff")
@login_required
@owner_required
def import_diff(log_id):
    conn = get_conn()
    log = conn.execute("SELECT * FROM import_log WHERE id = ?", (log_id,)).fetchone()
    conn.close()
    if not log:
        flash("Запись об импорте не найдена", "error")
        return redirect(url_for("import_view"))
    try:
        diff = json.loads(log["diff_json"] or "{}")
    except (json.JSONDecodeError, TypeError):
        diff = {}
    return render_template("import_diff.html", log=log, diff=diff)


@app.route("/import/fp-rp-filter", methods=["POST"])
@login_required
@owner_required
def fp_rp_filter_save():
    """Сохраняет фильтр РП для импорта финансового плана."""
    val = (request.form.get("fp_rp_filter") or "").strip()
    conn = get_conn()
    set_setting(conn, "fp_rp_filter", val)
    conn.commit()
    conn.close()
    if val:
        flash(f"Фильтр ФП по РП установлен: «{val}». Будут загружены только строки с этим РП.", "success")
    else:
        flash("Фильтр ФП по РП снят — загружаются все строки.", "success")
    return redirect(url_for("import_view"))


@app.route("/import/smb-settings", methods=["POST"])
@login_required
@owner_required
def smb_settings_save():
    """Сохраняет параметры подключения к SMB-шаре в app_settings."""
    conn = get_conn()
    for key in ("smb_server", "smb_share", "smb_subfolder", "smb_username", "smb_filename"):
        set_setting(conn, key, (request.form.get(key) or "").strip())
    # Пароль не перезаписываем, если поле оставлено пустым
    pwd = (request.form.get("smb_password") or "").strip()
    if pwd:
        set_setting(conn, "smb_password", pwd)
    conn.commit()
    conn.close()
    flash("Параметры сетевой папки сохранены.", "success")
    return redirect(url_for("import_view"))


@app.route("/import/smb-fetch", methods=["POST"])
@login_required
@owner_required
def smb_fetch():
    """Монтирует SMB-шару через mount_smbfs (macOS), копирует последний .xlsx и
    импортирует его так же, как при ручной загрузке через форму /import."""
    import subprocess as _sp

    conn = get_conn()
    smb_server    = get_setting(conn, "smb_server") or ""
    smb_share     = get_setting(conn, "smb_share") or ""
    smb_subfolder = get_setting(conn, "smb_subfolder") or ""
    smb_username  = get_setting(conn, "smb_username") or ""
    smb_password  = get_setting(conn, "smb_password") or ""
    smb_filename  = get_setting(conn, "smb_filename") or ""
    fp_rp_filter_smb = get_setting(conn, "fp_rp_filter") or None
    upload_quarter = get_setting(conn, "current_quarter_label") or ""
    conn.close()

    if not smb_server or not smb_share or not smb_username:
        flash("Укажите сервер, шару и логин в настройках сетевой папки.", "error")
        return redirect(url_for("import_view"))

    # mount_smbfs URL: //[DOMAIN;]user:password@server/share
    # Разделитель домена и пользователя — «;», а не «\» (как в Windows-нотации).
    # Все компоненты кроме сервера percent-кодируются, чтобы спецсимволы не ломали URL.
    if "\\" in smb_username:
        domain, user = smb_username.split("\\", 1)
        safe_user = urllib.parse.quote(domain, safe="") + ";" + urllib.parse.quote(user, safe="")
    else:
        safe_user = urllib.parse.quote(smb_username, safe="")
    safe_pwd   = urllib.parse.quote(smb_password, safe="")
    safe_share = urllib.parse.quote(smb_share, safe="")
    smb_url    = f"//{safe_user}:{safe_pwd}@{smb_server}/{safe_share}"

    mnt = tempfile.mkdtemp(prefix="fp_smb_")
    mounted = False
    dest = None
    try:
        res = _sp.run(["mount_smbfs", smb_url, mnt],
                      capture_output=True, text=True, timeout=20)
        if res.returncode != 0:
            err = (res.stderr or res.stdout or "неизвестная ошибка").strip()
            flash(f"Не удалось подключиться к шаре: {err}", "error")
            return redirect(url_for("import_view"))
        mounted = True

        search_dir = os.path.join(mnt, smb_subfolder) if smb_subfolder else mnt

        if smb_filename:
            src = os.path.join(search_dir, smb_filename)
            if not os.path.exists(src):
                flash(f"Файл «{smb_filename}» не найден в {smb_subfolder or '/'}.", "error")
                return redirect(url_for("import_view"))
        else:
            xlsx_files = glob.glob(os.path.join(search_dir, "*.xlsx"))
            if not xlsx_files:
                flash("В сетевой папке не найдено файлов .xlsx.", "error")
                return redirect(url_for("import_view"))
            src = max(xlsx_files, key=os.path.getmtime)

        fname = os.path.basename(src)
        dest  = UPLOAD_DIR / secure_filename(fname)
        shutil.copy2(src, dest)

    except _sp.TimeoutExpired:
        flash("Таймаут при подключении к сетевой папке. Проверьте доступность сервера.", "error")
        return redirect(url_for("import_view"))
    except Exception as e:
        flash(f"Ошибка при работе с сетевой папкой: {e}", "error")
        return redirect(url_for("import_view"))
    finally:
        if mounted:
            try:
                _sp.run(["umount", mnt], capture_output=True, timeout=10)
            except Exception:
                pass
        try:
            os.rmdir(mnt)
        except Exception:
            pass

    if dest is None:
        return redirect(url_for("import_view"))

    # Импортируем полученный файл
    try:
        result = import_excel(str(dest), fname, session["user_id"],
                              quarter_label=upload_quarter or None,
                              rp_filter=fp_rp_filter_smb or None)
    except Exception as e:
        flash(f"Файл получен ({fname}), но ошибка при импорте: {e}", "error")
        return redirect(url_for("import_view"))

    effective_quarter = result.get("quarter_label", upload_quarter)
    if effective_quarter:
        session["view_quarter"] = effective_quarter

    msg = (f"Файл «{fname}» получен с сетевой папки и импортирован"
           f" ({effective_quarter or '?'}): "
           f"всего {result['rows_total']}, новых {result['rows_new']}, "
           f"изменено {result['rows_updated']}, без изменений {result['rows_unchanged']}, "
           f"деактивировано {result['rows_deactivated']}")
    if result["header_mismatches"]:
        msg += f". Несовпадение заголовков: {result['header_mismatches']}"
    flash(msg, "success" if not result["header_mismatches"] else "warning")

    ro = result.get("rollover")
    if ro:
        flash(
            f"Первая загрузка {ro['new_label']}: перенос из {ro['old_label']} завершён. "
            f"По {ro['matched']} из {ro['candidates']} статей перенесены обязательства "
            f"({ro['obligations_copied']}) и примечания ({ro['comments_copied']}).",
            "success",
        )
    return redirect(url_for("import_view", last=1))


@app.route("/set-quarter-view", methods=["POST"])
@login_required
def set_quarter_view():
    """Переключает, какие данные сейчас смотрит пользователь.
    mode=quarter: один квартал (quarter=2026-Q2)
    mode=fy: весь финансовый год (fy=FY2026)
    """
    mode = request.form.get("mode", "quarter")
    if mode == "fy":
        fy = request.form.get("fy", "").strip()
        if fy and fy.startswith("FY"):
            session["view_mode"] = "fy"
            session["view_fy"] = fy
    else:
        q = request.form.get("quarter", "").strip()
        if q and "-Q" in q:
            session["view_mode"] = "quarter"
            session["view_quarter"] = q
    next_url = request.form.get("next") or url_for("dashboard")
    return redirect(next_url)


@app.route("/unfix-quarter", methods=["POST"])
@login_required
@owner_required
def unfix_quarter():
    """Снимает фиксацию квартала: убирает его из finalized_quarters.
    Если current_quarter_label был автоматически сдвинут на следующий квартал
    при фиксации, откатывает его обратно.
    """
    q = request.form.get("quarter", "").strip()
    if not q or "-Q" not in q:
        flash("Некорректная метка квартала", "error")
        return redirect(url_for("dashboard"))
    conn = get_conn()
    fin_str = get_setting(conn, "finalized_quarters") or ""
    finalized = {x.strip() for x in fin_str.split(",") if x.strip()}
    finalized.discard(q)
    set_setting(conn, "finalized_quarters", ",".join(sorted(finalized)))

    # Если current_quarter_label = следующий квартал после q — откатываем
    current_q = get_setting(conn, "current_quarter_label")
    rolled_back = False
    if current_q == next_quarter_label(q):
        set_setting(conn, "current_quarter_label", q)
        rolled_back = True
        session["view_mode"] = "quarter"
        session["view_quarter"] = q

    conn.commit()
    conn.close()

    msg = f"Фиксация квартала {q} снята."
    if rolled_back:
        msg += f" Портал возвращён к {q}."
    flash(msg, "success")
    return redirect(url_for("dashboard"))


@app.route("/fix-quarter", methods=["POST"])
@login_required
@owner_required
def fix_quarter():
    """Фиксирует текущий квартал как завершённый и переключает портал на следующий квартал.
    Добавляет старый квартал в finalized_quarters. Перенос обязательств/примечаний
    происходит автоматически при первой загрузке нового квартала (см. importer.py).
    """
    conn = get_conn()
    old_q = get_setting(conn, "current_quarter_label") or compute_quarter_label(datetime.date.today())
    new_q = next_quarter_label(old_q)

    # Отмечаем старый квартал как завершённый
    fin_str = get_setting(conn, "finalized_quarters") or ""
    finalized = {q.strip() for q in fin_str.split(",") if q.strip()}
    finalized.add(old_q)
    set_setting(conn, "finalized_quarters", ",".join(sorted(finalized)))

    # Переключаем текущий квартал на следующий
    set_setting(conn, "current_quarter_label", new_q)
    conn.commit()
    conn.close()

    # Переключаем просмотр пользователя на новый квартал
    session["view_quarter"] = new_q

    flash(
        f"Квартал {old_q} зафиксирован. Данные {old_q} сохранены и доступны в «Архиве». "
        f"Портал переключён на {new_q}. "
        f"При загрузке первого файла {new_q} обязательства и примечания "
        f"из открытых статей {old_q} будут перенесены автоматически.",
        "success",
    )
    return redirect(url_for("dashboard"))


# ---------- Архив кварталов ----------
# Только просмотр (без редактирования) -- доступен всем ролям, аналогично «Детализации» и
# «Моим обязательствам». При переходе на новый квартал (см. app/importer.py) строки прошлого
# квартала никуда не удаляются и не перемещаются -- они просто перестают быть is_active=1 и
# остаются доступны здесь под своей исходной меткой quarter_label.

@app.route("/archive")
@login_required
def archive_view():
    conn = get_conn()
    current_label = get_setting(conn, "current_quarter_label")
    quarters = [r["quarter_label"] for r in conn.execute("""
        SELECT DISTINCT quarter_label FROM fp_rows
        WHERE quarter_label IS NOT NULL AND quarter_label != ''
        ORDER BY quarter_label DESC
    """).fetchall()]

    selected = request.args.get("q") or ""
    if selected not in quarters:
        past = [q for q in quarters if q != current_label]
        selected = past[0] if past else (quarters[0] if quarters else "")

    rows = []
    obl_by_row = {}
    total_amount = 0.0
    if selected:
        rows = conn.execute("""
            SELECT fr.*, u.full_name AS responsible_full_name
            FROM fp_rows fr LEFT JOIN users u ON u.id = fr.responsible_user_id
            WHERE fr.quarter_label = ?
            ORDER BY fr.section, fr.client_name, fr.project_name
        """, (selected,)).fetchall()
        row_ids = [r["id"] for r in rows]
        if row_ids:
            ph = ",".join("?" * len(row_ids))
            for o in conn.execute(
                f"SELECT * FROM obligations WHERE fp_row_id IN ({ph}) ORDER BY id", row_ids
            ).fetchall():
                obl_by_row.setdefault(o["fp_row_id"], []).append(o)
        total_amount = sum((r["amount_0_100"] or 0.0) for r in rows)

    conn.close()
    return render_template(
        "archive.html", quarters=quarters, selected=selected, current_label=current_label,
        rows=rows, obl_by_row=obl_by_row, total_amount=total_amount, risk_levels=RISK_LEVELS,
    )


# ---------- Изменения: постоянная очередь на просмотр (row_events) ----------

@app.route("/changes")
@login_required
@owner_required
def changes_view():
    tab = request.args.get("tab", "new")
    if tab not in TAB_EVENT_TYPES:
        tab = "new"
    show_all = request.args.get("show") == "all"
    import_log_id = request.args.get("import_log_id", "")

    conn = get_conn()

    # Список загрузок для фильтра «Дата загрузки» — только те загрузки, по которым
    # реально есть события в журнале изменений (иначе в списке появлялись бы пустые даты).
    import_logs = conn.execute("""
        SELECT DISTINCT il.id, il.filename, il.imported_at
        FROM import_log il
        JOIN row_events e ON e.import_log_id = il.id
        ORDER BY il.imported_at DESC
    """).fetchall()

    types = TAB_EVENT_TYPES[tab]
    ph = ",".join("?" * len(types))
    query = f"""
        SELECT e.*, f.is_active AS row_active, f.amount_0_100 AS current_amount,
               ru.full_name AS reviewed_by_name
        FROM row_events e
        JOIN fp_rows f ON f.id = e.fp_row_id
        LEFT JOIN users ru ON ru.id = e.reviewed_by
        WHERE e.event_type IN ({ph})
    """
    params = list(types)
    if not show_all:
        query += " AND e.reviewed_at IS NULL"
    if import_log_id:
        query += " AND e.import_log_id = ?"
        params.append(import_log_id)
    query += " ORDER BY e.created_at DESC LIMIT 500"
    events = conn.execute(query, params).fetchall()

    counts = {}
    for key, t in TAB_EVENT_TYPES.items():
        ph2 = ",".join("?" * len(t))
        count_query = f"SELECT COUNT(*) c FROM row_events WHERE event_type IN ({ph2}) AND reviewed_at IS NULL"
        count_params = list(t)
        if import_log_id:
            count_query += " AND import_log_id = ?"
            count_params.append(import_log_id)
        counts[key] = conn.execute(count_query, count_params).fetchone()["c"]
    conn.close()
    return render_template(
        "changes.html", events=events, tab=tab, show_all=show_all,
        counts=counts, tab_labels=TAB_LABELS,
        import_logs=import_logs, filter_import_log_id=import_log_id,
    )


@app.route("/changes/<int:event_id>/review", methods=["POST"])
@login_required
@owner_required
def review_change(event_id):
    conn = get_conn()
    conn.execute(
        "UPDATE row_events SET reviewed_at = CURRENT_TIMESTAMP, reviewed_by = ? WHERE id = ?",
        (session["user_id"], event_id),
    )
    conn.commit()
    conn.close()
    tab = request.form.get("tab", "new")
    show_all = request.form.get("show", "")
    import_log_id = request.form.get("import_log_id", "")
    return redirect(url_for("changes_view", tab=tab, show=(show_all or None), import_log_id=(import_log_id or None)))


@app.route("/changes/review_all", methods=["POST"])
@login_required
@owner_required
def review_all_changes():
    tab = request.form.get("tab", "new")
    if tab not in TAB_EVENT_TYPES:
        tab = "new"
    import_log_id = request.form.get("import_log_id", "")
    types = TAB_EVENT_TYPES[tab]
    conn = get_conn()
    ph = ",".join("?" * len(types))
    update_query = f"UPDATE row_events SET reviewed_at = CURRENT_TIMESTAMP, reviewed_by = ? WHERE event_type IN ({ph}) AND reviewed_at IS NULL"
    update_params = [session["user_id"]] + list(types)
    if import_log_id:
        update_query += " AND import_log_id = ?"
        update_params.append(import_log_id)
    conn.execute(update_query, update_params)
    conn.commit()
    conn.close()
    flash("Изменения отмечены как просмотренные", "success")
    return redirect(url_for("changes_view", tab=tab, import_log_id=(import_log_id or None)))


# ---------- FP rows list + obligations (с группировкой, сортировкой и раскрытием) ----------

SORTABLE_FIELDS = {
    "client_name": "Клиент", "project_name": "Проект", "project_num": "Номер проекта",
    "section": "Раздел", "project_manager": "Менеджер", "contract_num": "Договор",
    "dpa_date": "ДПА", "amount_0_100": "Сумма 0-100", "portfolio": "Портфель", "month": "Месяц",
}
GROUPABLE_FIELDS = {
    "": "Без группировки", "section": "Раздел ФП", "client_name": "Клиент",
    "project_manager": "Менеджер проекта", "portfolio": "Портфель", "month": "Месяц",
}
# Столбцы режима «Агрегировано по клиентам» — отдельное пространство сортировки (asort/adir),
# чтобы не пересекаться с sort/dir детального вида.
AGG_SORT_FIELDS = {
    "client": "Клиент", "category": "Категория", "managers": "Менеджер(ы)",
    "amount_0_100": "0-100", "amount_opportunities": "Возможности", "total": "Итого",
    "amount_fact": "Факт",
    "has_obligations": "Обязательства", "risk_level": "Риск",
}


@app.route("/rows")
@login_required
def rows_list():
    conn = get_conn()
    section = request.args.getlist("section")
    portfolio = request.args.getlist("portfolio")
    manager = request.args.getlist("manager")
    client = request.args.getlist("client")
    search = request.args.get("q", "")
    only_unassigned = request.args.get("unassigned") == "1"
    agg_sort = request.args.get("asort", "total")
    if agg_sort not in AGG_SORT_FIELDS:
        agg_sort = "total"
    agg_dir = request.args.get("adir", "desc")
    if agg_dir not in ("asc", "desc"):
        agg_dir = "desc"
    # Программные фильтры -- не показаны как отдельные галочки в панели фильтров, но
    # используются ссылками-суммами с дашборда (риск под суммой, "к получению за 4 дня" по ДПА).
    risk_level = []
    for v in request.args.getlist("risk_level"):
        try:
            iv = int(v)
        except (TypeError, ValueError):
            continue
        if iv in (1, 2, 3):
            risk_level.append(iv)
    dpa_from = request.args.get("dpa_from", "")
    dpa_to = request.args.get("dpa_to", "")
    current_query_string = request.query_string.decode("utf-8")

    vqls, _mode = get_view_context(conn)
    ql_ph = ",".join("?" * len(vqls))
    months, qnum, outlier_months = get_period_info(vqls, conn=conn)
    m_ph = ",".join("?" * len(months)) if months else "''"
    query = f"""
        SELECT f.*
        FROM fp_rows f
        WHERE f.is_active = 1 AND f.quarter_label IN ({ql_ph}) AND f.month IN ({m_ph})
    """
    params = vqls + list(months)

    if section:
        query += f" AND f.section IN ({','.join('?' * len(section))})"
        params += section
    if portfolio:
        query += f" AND f.portfolio IN ({','.join('?' * len(portfolio))})"
        params += portfolio
    if manager:
        query += f" AND f.project_manager IN ({','.join('?' * len(manager))})"
        params += manager
    if client:
        query += f" AND f.client_name IN ({','.join('?' * len(client))})"
        params += client
    if risk_level:
        query += f" AND f.risk_level IN ({','.join('?' * len(risk_level))})"
        params += risk_level
    if dpa_from:
        query += " AND f.dpa_date >= ?"
        params.append(dpa_from)
    if dpa_to:
        query += " AND f.dpa_date <= ?"
        params.append(dpa_to)
    if search:
        query += " AND (LOWER(f.client_name) LIKE LOWER(?) OR LOWER(f.project_name) LIKE LOWER(?) OR LOWER(f.contract_num) LIKE LOWER(?))"
        params += [f"%{search}%"] * 3

    query += " ORDER BY f.client_name, f.dpa_date IS NULL, f.dpa_date"
    rows = conn.execute(query, params).fetchall()

    row_ids = [r["id"] for r in rows]
    obl_by_row = {}   # {row_id: count} — только для счётчика на кнопке
    if row_ids:
        ph = ",".join("?" * len(row_ids))
        if only_unassigned:
            # Полные данные нужны для фильтрации по типу ответственного
            obls_full = conn.execute(
                f"SELECT fp_row_id, responsible_type FROM obligations WHERE fp_row_id IN ({ph})",
                row_ids,
            ).fetchall()
            _obl_full_map = {}  # {row_id: [list of obls]}
            for o in obls_full:
                _obl_full_map.setdefault(o["fp_row_id"], []).append(o)
            rows = [r for r in rows if (
                (r["section"] in TEAM_LEAD_SECTIONS and not any(
                    o["responsible_type"] == "team_lead" for o in _obl_full_map.get(r["id"], [])
                ))
                or
                (r["section"] in OWNER_SECTIONS and not any(
                    o["responsible_type"] == "owner" for o in _obl_full_map.get(r["id"], [])
                ))
            )]
            # Пересчитываем row_ids после фильтрации
            row_ids = [r["id"] for r in rows]
            ph = ",".join("?" * len(row_ids)) if row_ids else None

        if row_ids:
            ph = ph or ",".join("?" * len(row_ids))
            for cnt_row in conn.execute(
                f"SELECT fp_row_id, COUNT(*) as cnt FROM obligations WHERE fp_row_id IN ({ph}) GROUP BY fp_row_id",
                row_ids,
            ).fetchall():
                obl_by_row[cnt_row["fp_row_id"]] = cnt_row["cnt"]

    # Дропдауны фильтров — берём из уже загруженных строк (без дополнительных запросов к БД).
    # При активном section/client/manager-фильтре дополнительно подгружаем все доступные значения.
    _all_rows_q = f"SELECT section, client_name, project_manager FROM fp_rows WHERE is_active=1 AND quarter_label IN ({ql_ph})"
    _all_rows = conn.execute(_all_rows_q, vqls).fetchall()
    _sections_set  = sorted({r["section"]         for r in _all_rows if r["section"]},         key=str)
    _clients_set   = sorted({r["client_name"]      for r in _all_rows if r["client_name"]},     key=str)
    _managers_set  = sorted({r["project_manager"]  for r in _all_rows if r["project_manager"]}, key=str)

    # Агрегированный вид: аккордеон по клиентам (client → sections → rows)
    client_accordion = None
    if True:
        cli_map   = {}
        cli_order = []
        for r in rows:
            cli = r["client_name"] or "—"
            sec = r["section"] or "—"
            if cli not in cli_map:
                cli_map[cli] = {
                    "client": cli,
                    "amount_0100": 0.0, "amount_opp": 0.0, "fact_total": 0.0,
                    "row_count": 0, "max_risk": 0,
                    "sec_map": {}, "sec_order": [],
                }
                cli_order.append(cli)
            cm = cli_map[cli]
            if sec not in cm["sec_map"]:
                cm["sec_map"][sec] = {
                    "section": sec,
                    "amount_0100": 0.0, "amount_opp": 0.0, "fact_total": 0.0,
                    "row_count": 0, "max_risk": 0,
                    "rows": [],
                }
                cm["sec_order"].append(sec)
            sm = cm["sec_map"][sec]

            p   = r["portfolio"]
            amt = r["amount_0_100"] or 0
            if p == "0-100":
                cm["amount_0100"] += amt
                sm["amount_0100"] += amt
            elif p == "Возможности":
                cm["amount_opp"] += amt
                sm["amount_opp"] += amt
            elif p == "Факт":
                cm["fact_total"] += amt
                sm["fact_total"] += amt
            cm["row_count"] += 1
            sm["row_count"] += 1
            cm["max_risk"] = max(cm["max_risk"], r["risk_level"] or 0)
            sm["max_risk"] = max(sm["max_risk"], r["risk_level"] or 0)
            sm["rows"].append(r)

        client_accordion = []
        for cli in cli_order:
            cm = cli_map[cli]
            sections = [cm["sec_map"][s] for s in cm["sec_order"]]
            sections.sort(key=lambda s: -(s["amount_0100"] + s["amount_opp"] + s["fact_total"]))
            plan  = cm["amount_0100"] + cm["amount_opp"]
            total = plan + cm["fact_total"]
            client_accordion.append({
                "client":        cm["client"],
                "amount_0100":   cm["amount_0100"],
                "amount_opp":    cm["amount_opp"],
                "fact_total":    cm["fact_total"],
                "plan_total":    plan,
                "total_revenue": total,   # для шкалы бара: 0-100 + Возм. + Факт
                "row_count":     cm["row_count"],
                "max_risk":      cm["max_risk"],
                "sections":      sections,
            })
        client_accordion.sort(key=lambda c: -(c["amount_0100"] + c["amount_opp"] + c["fact_total"]))


    saved_filters = conn.execute(
        "SELECT * FROM saved_filters WHERE user_id = ? ORDER BY created_at DESC", (session["user_id"],)
    ).fetchall()

    # Целевой план квартала для блока сравнения
    agg_target_amount = None
    if months:
        period_label_agg = "-".join(months)
        trow = conn.execute(
            "SELECT target_amount FROM quarter_targets WHERE period_label = ?",
            (period_label_agg,),
        ).fetchone()
        if trow:
            agg_target_amount = trow["target_amount"]

    conn.close()
    return render_template(
        "rows_list.html", rows=rows, obl_by_row=obl_by_row,
        sections=_sections_set,
        clients=_clients_set,
        managers=_managers_set,
        saved_filters=saved_filters, current_query_string=current_query_string,
        filters=dict(section=section, portfolio=portfolio, manager=manager, q=search,
                     unassigned=only_unassigned, client=client,
                     risk_level=risk_level, dpa_from=dpa_from, dpa_to=dpa_to),
        owner_sections=OWNER_SECTIONS, team_lead_sections=TEAM_LEAD_SECTIONS,
        today=datetime.date.today().isoformat(),
        outlier_months=outlier_months, client_accordion=client_accordion,
        agg_sort=agg_sort, agg_dir=agg_dir, agg_sortable_fields=AGG_SORT_FIELDS,
        risk_levels=RISK_LEVELS, agg_target_amount=agg_target_amount,
    )


@app.route("/rows/filters/save", methods=["POST"])
@login_required
def save_filter():
    name = request.form.get("name", "").strip()
    query_string = request.form.get("query_string", "")
    if not name:
        flash("Укажите название набора фильтров", "error")
    else:
        conn = get_conn()
        conn.execute(
            "INSERT INTO saved_filters (user_id, name, query_string) VALUES (?,?,?)",
            (session["user_id"], name, query_string),
        )
        conn.commit()
        conn.close()
        flash(f"Набор фильтров «{name}» сохранён", "success")
    target = url_for("rows_list")
    if query_string:
        target += "?" + query_string
    return redirect(target)


@app.route("/rows/filters/<int:filter_id>/delete", methods=["POST"])
@login_required
def delete_saved_filter(filter_id):
    conn = get_conn()
    conn.execute("DELETE FROM saved_filters WHERE id = ? AND user_id = ?", (filter_id, session["user_id"]))
    conn.commit()
    conn.close()
    flash("Набор фильтров удалён", "success")
    return redirect(url_for("rows_list"))


def render_obligation_fragment(row_id):
    # Рендерит ту же панель обязательств, что обычно скрыта внутри страницы «Детализация»
    # (см. _obligations_fragment.html), но как отдельный HTML-фрагмент. Используется роутами
    # ниже при AJAX-запросе (заголовок X-Requested-With: fetch), чтобы обновить в браузере
    # только эту панель, без перезагрузки всей страницы и без потери остальных открытых строк.
    conn = get_conn()
    row = conn.execute("SELECT * FROM fp_rows WHERE id = ?", (row_id,)).fetchone()
    if not row:
        conn.close()
        return None
    obls = conn.execute(
        "SELECT * FROM obligations WHERE fp_row_id = ? ORDER BY due_date IS NULL, due_date ASC",
        (row_id,),
    ).fetchall()
    conn.close()
    return render_template(
        "_obligations_fragment.html",
        row=row, obl_by_row={row_id: obls},
        today=datetime.date.today().isoformat(),
        risk_levels=RISK_LEVELS,
        owner_sections=OWNER_SECTIONS, team_lead_sections=TEAM_LEAD_SECTIONS,
    )


def _is_fetch_request():
    return request.headers.get("X-Requested-With") == "fetch"


@app.route("/rows/<int:row_id>/obl-panel")
@login_required
def row_obl_panel(row_id):
    """Ленивая загрузка панели обязательств по строке (AJAX, GET).
    Возвращает HTML-фрагмент <tr class='obl-row'> для вставки в DOM."""
    frag = render_obligation_fragment(row_id)
    if frag is None:
        return "", 404
    return frag


@app.route("/rows/<int:row_id>/comment", methods=["POST"])
@login_required
@owner_required
def set_row_comment(row_id):
    conn = get_conn()
    row = conn.execute("SELECT id FROM fp_rows WHERE id = ?", (row_id,)).fetchone()
    if not row:
        conn.close()
        if _is_fetch_request():
            return jsonify({"error": "Строка не найдена"}), 404
        flash("Строка не найдена", "error")
        return redirect(url_for("rows_list"))
    comment = request.form.get("comment", "").strip()
    conn.execute("UPDATE fp_rows SET internal_comment = ? WHERE id = ?", (comment or None, row_id))
    conn.commit()
    conn.close()
    if _is_fetch_request():
        return render_obligation_fragment(row_id)
    flash("Комментарий сохранён", "success")
    return redirect(url_for("rows_list", _anchor=f"row-{row_id}", view="detail"))


@app.route("/rows/<int:row_id>/risk", methods=["POST"])
@login_required
@owner_required
def set_row_risk(row_id):
    conn = get_conn()
    row = conn.execute("SELECT id FROM fp_rows WHERE id = ?", (row_id,)).fetchone()
    if not row:
        conn.close()
        if _is_fetch_request():
            return jsonify({"error": "Строка не найдена"}), 404
        flash("Строка не найдена", "error")
        return redirect(url_for("rows_list"))

    try:
        risk_level = int(request.form.get("risk_level", "0"))
    except ValueError:
        risk_level = 0
    if risk_level not in (0, 1, 2, 3):
        risk_level = 0

    conn.execute("UPDATE fp_rows SET risk_level = ?, is_risk = ? WHERE id = ?",
                 (risk_level, 1 if risk_level else 0, row_id))
    conn.commit()
    conn.close()
    if _is_fetch_request():
        return render_obligation_fragment(row_id)
    if risk_level:
        flash(f"Отмечено: «{RISK_LEVELS[risk_level]['label']}»", "success")
    else:
        flash("Отметка риска снята", "success")
    return redirect(url_for("rows_list", _anchor=f"row-{row_id}", view="detail"))


@app.route("/rows/<int:row_id>/obligations/add", methods=["POST"])
@login_required
def add_obligation(row_id):
    # "next" -- необязательный путь для редиректа назад (например, с дашборда, чтобы не уносить
    # пользователя на страницу детализации после быстрого добавления ответственного/обязательства).
    next_url = request.form.get("next", "").strip()
    is_fetch = _is_fetch_request()

    def _redirect_back():
        if next_url and next_url.startswith("/"):
            return redirect(next_url)
        return redirect(url_for("rows_list", _anchor=f"row-{row_id}", view="detail"))

    def _fail(message, status=400):
        if is_fetch:
            return jsonify({"error": message}), status
        flash(message, "error")
        return _redirect_back()

    conn = get_conn()
    row = conn.execute("SELECT * FROM fp_rows WHERE id = ?", (row_id,)).fetchone()
    if not row:
        conn.close()
        if is_fetch:
            return jsonify({"error": "Строка не найдена"}), 404
        flash("Строка не найдена", "error")
        return redirect(url_for("rows_list"))

    title = request.form.get("title", "").strip()
    description = request.form.get("description", "").strip()
    responsible_name = request.form.get("responsible_name", "").strip()
    due_date = request.form.get("due_date") or None

    if not title:
        conn.close()
        return _fail("Укажите текст обязательства")

    if not responsible_name:
        conn.close()
        return _fail("Выберите ответственного")

    _all_names, _tl_names_route, _name_to_role = _get_portal_users(conn)
    if row["section"] in TEAM_LEAD_SECTIONS and responsible_name not in _tl_names_route:
        conn.close()
        return _fail("Для проектных статей выберите одного из руководителей команд")

    if responsible_name not in _all_names:
        conn.close()
        return _fail("Выберите ответственного из списка")

    responsible_type = "owner" if _name_to_role.get(responsible_name) == "owner" else "team_lead"

    conn.execute("""
        INSERT INTO obligations (fp_row_id, title, description, responsible_type, responsible_name, due_date, created_by)
        VALUES (?,?,?,?,?,?,?)
    """, (row_id, title, description, responsible_type, responsible_name, due_date, session["user_id"]))
    obl_id = conn.execute("SELECT last_insert_rowid() id").fetchone()["id"]
    conn.execute(
        "INSERT INTO obligation_history (obligation_id, user_id, action, details) VALUES (?,?,?,?)",
        (obl_id, session["user_id"], "created", title),
    )
    conn.commit()
    conn.close()
    if is_fetch:
        return render_obligation_fragment(row_id)
    flash("Обязательство добавлено", "success")
    return _redirect_back()


@app.route("/obligations/<int:obl_id>/reassign", methods=["POST"])
@login_required
@owner_required
def reassign_obligation(obl_id):
    is_fetch = _is_fetch_request()
    conn = get_conn()
    obl = conn.execute("SELECT * FROM obligations WHERE id = ?", (obl_id,)).fetchone()
    if not obl:
        conn.close()
        if is_fetch:
            return jsonify({"error": "Обязательство не найдено"}), 404
        flash("Обязательство не найдено", "error")
        return redirect(url_for("rows_list"))

    new_name = request.form.get("responsible_name", "").strip()
    _all_names_r, _, _name_to_role_r = _get_portal_users(conn)
    if new_name not in _all_names_r:
        conn.close()
        if is_fetch:
            return jsonify({"error": "Выберите ответственного из списка"}), 400
        flash("Выберите ответственного из списка", "error")
        return redirect(url_for("rows_list", _anchor=f"row-{obl['fp_row_id']}", view="detail"))

    new_type = "owner" if _name_to_role_r.get(new_name) == "owner" else "team_lead"

    conn.execute(
        "UPDATE obligations SET responsible_name = ?, responsible_type = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
        (new_name, new_type, obl_id),
    )
    conn.execute(
        "INSERT INTO obligation_history (obligation_id, user_id, action, details) VALUES (?,?,?,?)",
        (obl_id, session["user_id"], "reassigned", f"{obl['responsible_name']} → {new_name}"),
    )
    conn.commit()
    row_id = obl["fp_row_id"]
    conn.close()
    if is_fetch:
        return render_obligation_fragment(row_id)
    flash("Ответственный изменён", "success")
    return redirect(url_for("rows_list", _anchor=f"row-{row_id}", view="detail"))


@app.route("/obligations/<int:obl_id>/status", methods=["POST"])
@login_required
def update_obligation_status(obl_id):
    conn = get_conn()
    obl = conn.execute("SELECT * FROM obligations WHERE id = ?", (obl_id,)).fetchone()
    if not obl:
        conn.close()
        return jsonify({"error": "not found"}), 404
    if not can_edit_obligation(obl):
        conn.close()
        return jsonify({"error": "forbidden"}), 403

    new_status = request.form.get("status")
    if new_status not in ("not_started", "in_progress", "done", "blocked"):
        conn.close()
        return jsonify({"error": "bad status"}), 400

    completed_at = "CURRENT_TIMESTAMP" if new_status == "done" else "NULL"
    conn.execute(f"""
        UPDATE obligations SET status = ?, updated_at = CURRENT_TIMESTAMP, completed_at = {completed_at}
        WHERE id = ?
    """, (new_status, obl_id))
    conn.execute(
        "INSERT INTO obligation_history (obligation_id, user_id, action, details) VALUES (?,?,?,?)",
        (obl_id, session["user_id"], "status_changed", new_status),
    )
    conn.commit()
    row_id = obl["fp_row_id"]
    conn.close()
    if _is_fetch_request():
        return render_obligation_fragment(row_id)
    return redirect(url_for("rows_list", _anchor=f"row-{row_id}", view="detail"))


@app.route("/obligations/<int:obl_id>/delete", methods=["POST"])
@login_required
def delete_obligation(obl_id):
    is_fetch = _is_fetch_request()
    conn = get_conn()
    obl = conn.execute("SELECT * FROM obligations WHERE id = ?", (obl_id,)).fetchone()
    if not obl:
        conn.close()
        if is_fetch:
            return jsonify({"error": "Обязательство не найдено"}), 404
        flash("Обязательство не найдено", "error")
        return redirect(url_for("rows_list"))
    if session.get("role") != "owner":
        conn.close()
        if is_fetch:
            return jsonify({"error": "Удалять обязательства может только руководитель продукта"}), 403
        flash("Удалять обязательства может только руководитель продукта", "error")
        return redirect(url_for("rows_list", _anchor=f"row-{obl['fp_row_id']}", view="detail"))
    row_id = obl["fp_row_id"]
    conn.execute("DELETE FROM obligations WHERE id = ?", (obl_id,))
    conn.commit()
    conn.close()
    if is_fetch:
        return render_obligation_fragment(row_id)
    flash("Обязательство удалено", "success")
    return redirect(url_for("rows_list", _anchor=f"row-{row_id}", view="detail"))


# ---------- My obligations (per team lead) ----------

@app.route("/my")
@login_required
def my_obligations():
    conn = get_conn()
    if session["role"] == "team_lead":
        name = session["team_lead_name"]
        obls = conn.execute("""
            SELECT o.*, f.client_name, f.project_name, f.section, f.amount_0_100, f.contract_num
            FROM obligations o JOIN fp_rows f ON f.id = o.fp_row_id
            WHERE o.responsible_name = ? ORDER BY o.due_date IS NULL, o.due_date ASC
        """, (name,)).fetchall()
    else:
        obls = conn.execute("""
            SELECT o.*, f.client_name, f.project_name, f.section, f.amount_0_100, f.contract_num
            FROM obligations o JOIN fp_rows f ON f.id = o.fp_row_id
            ORDER BY o.due_date IS NULL, o.due_date ASC
        """).fetchall()
    conn.close()
    return render_template("my_obligations.html", obligations=obls, today=datetime.date.today().isoformat())


# ---------- Admin: users management ----------

@app.route("/admin/users", methods=["GET", "POST"])
@login_required
@owner_required
def admin_users():
    conn = get_conn()
    if request.method == "POST":
        username = request.form.get("username", "").strip().lower()
        full_name = request.form.get("full_name", "").strip()
        password = request.form.get("password", "")
        role = request.form.get("role")
        team_lead_name = request.form.get("team_lead_name") or None
        if role != "team_lead":
            team_lead_name = None
        if not (username and full_name and password and role):
            flash("Заполните все поля", "error")
        else:
            try:
                conn.execute(
                    "INSERT INTO users (username, full_name, password_hash, role, team_lead_name) VALUES (?,?,?,?,?)",
                    (username, full_name, generate_password_hash(password, method="pbkdf2:sha256"), role, team_lead_name),
                )
                conn.commit()
                flash("Пользователь создан", "success")
            except Exception as e:
                flash(f"Ошибка: {e}", "error")
    users = conn.execute("SELECT * FROM users ORDER BY role, full_name").fetchall()
    conn.close()
    return render_template("admin_users.html", users=users, team_leads=TEAM_LEADS)


@app.route("/admin/users/<int:user_id>/edit", methods=["POST"])
@login_required
@owner_required
def edit_user(user_id):
    conn = get_conn()
    user = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not user:
        conn.close()
        flash("Пользователь не найден", "error")
        return redirect(url_for("admin_users"))

    username = request.form.get("username", "").strip().lower()
    full_name = request.form.get("full_name", "").strip()
    password = request.form.get("password", "")
    role = request.form.get("role")
    team_lead_name = request.form.get("team_lead_name") or None
    if role != "team_lead":
        team_lead_name = None

    if not (username and full_name and role):
        flash("Заполните логин, имя и роль", "error")
        conn.close()
        return redirect(url_for("admin_users"))

    if role not in ("owner", "team_lead", "viewer"):
        flash("Некорректная роль", "error")
        conn.close()
        return redirect(url_for("admin_users"))

    if user["role"] == "owner" and role != "owner":
        remaining_owners = conn.execute(
            "SELECT COUNT(*) c FROM users WHERE role = 'owner' AND id != ?", (user_id,)
        ).fetchone()["c"]
        if remaining_owners == 0:
            flash("Нельзя снять роль Owner с последнего владельца — сначала назначьте другого Owner", "error")
            conn.close()
            return redirect(url_for("admin_users"))

    try:
        if password:
            conn.execute(
                "UPDATE users SET username=?, full_name=?, password_hash=?, role=?, team_lead_name=? WHERE id=?",
                (username, full_name, generate_password_hash(password, method="pbkdf2:sha256"), role, team_lead_name, user_id),
            )
        else:
            conn.execute(
                "UPDATE users SET username=?, full_name=?, role=?, team_lead_name=? WHERE id=?",
                (username, full_name, role, team_lead_name, user_id),
            )
        conn.commit()
        if session.get("user_id") == user_id:
            session["username"] = username
            session["full_name"] = full_name
            session["role"] = role
            session["team_lead_name"] = team_lead_name
        flash(f"Пользователь «{full_name}» обновлён", "success")
    except Exception as e:
        flash(f"Ошибка: {e}", "error")
    conn.close()
    return redirect(url_for("admin_users"))


@app.route("/admin/users/<int:user_id>/delete", methods=["POST"])
@login_required
@owner_required
def delete_user(user_id):
    conn = get_conn()
    conn.execute("DELETE FROM users WHERE id = ? AND role != 'owner'", (user_id,))
    conn.commit()
    conn.close()
    return redirect(url_for("admin_users"))


@app.route("/admin/users/<int:user_id>/telegram_code", methods=["POST"])
@login_required
@owner_required
def generate_telegram_code(user_id):
    from app.db import generate_link_code
    conn = get_conn()
    user = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not user:
        conn.close()
        flash("Пользователь не найден", "error")
        return redirect(url_for("admin_users"))
    code = generate_link_code()
    conn.execute("UPDATE users SET telegram_link_code = ?, telegram_chat_id = NULL WHERE id = ?", (code, user_id))
    conn.commit()
    conn.close()
    flash(f"Код привязки для {user['full_name']}: {code} — передайте его пользователю, он должен отправить его боту командой /link {code}", "success")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/<int:user_id>/telegram_unlink", methods=["POST"])
@login_required
@owner_required
def unlink_telegram(user_id):
    conn = get_conn()
    conn.execute("UPDATE users SET telegram_chat_id = NULL, telegram_link_code = NULL WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()
    flash("Telegram отвязан", "success")
    return redirect(url_for("admin_users"))


# ═══════════════════════════════════════════════════════════════════════════════
#  ТРУДОЗАТРАТЫ
# ═══════════════════════════════════════════════════════════════════════════════

_RE_PROJ_NUM = re.compile(r'^(\d{6,8})\b')

# Слова, которые убираем при нормализации имён клиентов (правовые формы и мусор)
_CLIENT_LEGAL_RE = re.compile(
    r'\b(общество|общества|ответственностью|ограниченной|акционерное|акционерный|'
    r'акционерного|акционерной|публичное|закрытое|открытое|коммерческий|коммерческого|'
    r'государственный|специализированный|российский|экспортно|импортный|'
    r'инвестиционно|финансовая|корпорация|компания|'
    r'ооо|ао|пао|оао|зао|акб|кб|нко|банк|bank)\b',
    re.IGNORECASE | re.UNICODE
)


def _client_tokens(name: str) -> frozenset:
    """Возвращает набор характерных слов из названия клиента."""
    if not name:
        return frozenset()
    s = re.sub(r'["\'\(\)«»,\.\-/]', ' ', name.upper())
    s = _CLIENT_LEGAL_RE.sub(' ', s)
    words = re.findall(r'[А-ЯЁA-Z]{3,}', s)
    return frozenset(words)


def _client_similarity(name_a: str, tokens_b: frozenset):
    """
    Возвращает (score, shared_tokens).
    score: Jaccard по характерным токенам (>= 0.10 считаем возможным совпадением).
    tokens_b — уже нормализованный frozenset (из _client_tokens).
    """
    ta = _client_tokens(name_a)
    if not ta or not tokens_b:
        return 0.0, []
    inter = ta & tokens_b
    if not inter:
        return 0.0, []
    union = ta | tokens_b
    return len(inter) / len(union), sorted(inter)


def _ts_build_compare(import_id, conn, emp_names=None, fp_quarter_labels=None):
    """
    Сопоставляет трудозатраты (ts_rows) с финпланом (fp_rows).
    fp_quarter_labels: список quarter_label для фильтрации ФП; None = все кварталы.
    Порядок матча:
      1) По числовому номеру проекта (в начале названия ts_rows.project)
      2) По точному совпадению названия проекта
      3) По нечёткому совпадению имени клиента (Jaccard по характерным словам)

    match_type:
      'num'         — по номеру
      'name'        — точное имя
      'client_high' — клиент, Jaccard >= 0.5  (надёжно)
      'client_med'  — клиент, Jaccard 0.25–0.5 (предупреждение)
      'client_low'  — клиент, Jaccard 0.1–0.25 (высокая вероятность ошибки)

    Возвращает (hierarchy, total_hours, totals,
                hierarchy_no_mine, hours_no_mine,
                hierarchy_less_mine, hours_less_mine).
    """
    # ── Все строки для построения иерархии ───────────────────────────────────
    all_rows = conn.execute(
        "SELECT * FROM ts_rows WHERE import_id=?", (import_id,)
    ).fetchall()

    # ── Агрегируем по проекту для матчинга ───────────────────────────────────
    ts_agg = {}   # project_name → {hours, client}
    for r in all_rows:
        proj = r["project"] or "—"
        if proj not in ts_agg:
            ts_agg[proj] = {"hours": 0.0, "client": r["client"] or ""}
        ts_agg[proj]["hours"] += r["hours"] or 0.0

    # ── Индекс финплана ───────────────────────────────────────────────────────
    fp_by_num    = {}   # cleaned_num   → fp_rec
    fp_by_name   = {}   # lower_name    → fp_rec
    fp_by_client = {}   # frozenset(tokens) → list[fp_rec]  (для fuzzy-клиента)

    # Суммируем раздельно по portfolio: Факт / 0-100 / Возможности
    _fp_case = (
        "SUM(CASE WHEN portfolio='Факт'        THEN amount_0_100 ELSE 0 END) as fp_fact, "
        "SUM(CASE WHEN portfolio='0-100'       THEN amount_0_100 ELSE 0 END) as fp_plan, "
        "SUM(CASE WHEN portfolio='Возможности' THEN amount_0_100 ELSE 0 END) as fp_opp "
    )
    if fp_quarter_labels:
        ql_ph = ",".join("?" * len(fp_quarter_labels))
        fp_sql = (
            f"SELECT project_num, project_name, section, client_name, {_fp_case}"
            f"FROM fp_rows WHERE is_active=1 AND quarter_label IN ({ql_ph}) "
            f"GROUP BY project_num"
        )
        fp_params = list(fp_quarter_labels)
    else:
        fp_sql = (
            f"SELECT project_num, project_name, section, client_name, {_fp_case}"
            f"FROM fp_rows WHERE is_active=1 GROUP BY project_num"
        )
        fp_params = []

    for r in conn.execute(fp_sql, fp_params).fetchall():
        num_clean = re.sub(r"[\s\xa0]+", "", r["project_num"] or "")
        fp_rec = {
            "fact":    r["fp_fact"] or 0.0,
            "plan":    r["fp_plan"] or 0.0,
            "opp":     r["fp_opp"]  or 0.0,
            "section": r["section"] or "",
            "client":  r["client_name"] or "",
            "fp_name": r["project_name"] or "",
        }
        if num_clean and num_clean != "0":
            fp_by_num[num_clean] = fp_rec
        name_key = re.sub(r"[\s\xa0]+", " ", r["project_name"] or "").strip().lower()
        if name_key:
            fp_by_name[name_key] = fp_rec
        # Клиентский индекс: группируем по набору токенов клиента
        tokens = _client_tokens(r["client_name"] or "")
        if tokens:
            fp_by_client.setdefault(tokens, []).append(fp_rec)

    # Предвычисляем список (tokens, recs) один раз
    fp_client_index = list(fp_by_client.items())

    # ── Матчинг → proj_finance ────────────────────────────────────────────────
    proj_finance = {}   # lower(project_name) → финансовые данные

    for proj_name, data in ts_agg.items():
        ts_cli = data["client"]
        fp_rec = None
        match_type = None
        match_score = 1.0
        match_tokens = []

        # 1) По числу
        m = _RE_PROJ_NUM.match(proj_name.strip())
        if m and m.group(1) in fp_by_num:
            fp_rec = fp_by_num[m.group(1)]
            match_type = "num"

        # 2) По точному имени
        if fp_rec is None:
            key = re.sub(r"[\s\xa0]+", " ", proj_name).strip().lower()
            if key in fp_by_name:
                fp_rec = fp_by_name[key]
                match_type = "name"

        # 3) По клиенту (fuzzy)
        if fp_rec is None and ts_cli and ts_cli != "**DIASOFT":
            best_score, best_rec, best_tokens = 0.0, None, []
            for fp_tokens, recs in fp_client_index:
                score, shared = _client_similarity(ts_cli, fp_tokens)
                if score > best_score:
                    best_score, best_rec, best_tokens = score, recs[0], shared
            if best_score >= 0.10 and best_rec:
                fp_rec = best_rec
                match_score = best_score
                match_tokens = best_tokens
                if best_score >= 0.50:
                    match_type = "client_high"
                elif best_score >= 0.25:
                    match_type = "client_med"
                else:
                    match_type = "client_low"

        if fp_rec:
            proj_finance[proj_name.lower()] = {
                "fp_fact":      fp_rec["fact"],
                "fp_plan":      fp_rec["plan"],
                "fp_opp":       fp_rec["opp"],
                "section":      fp_rec["section"],
                "client_fp":    fp_rec["client"] or ts_cli,
                "ts_client":    ts_cli,
                "match_type":   match_type,
                "match_score":  round(match_score, 2),
                "match_tokens": match_tokens,
            }

    # ── Строим иерархию с финансовыми данными ────────────────────────────────
    hierarchy, total_hours = _ts_build_compare_hier(all_rows, emp_names=emp_names, proj_finance=proj_finance)

    matched_hours = sum(
        ts_agg[p]["hours"] for p in ts_agg if p.lower() in proj_finance
    )
    totals = {
        "total_hours":   total_hours,
        "matched_hours": matched_hours,
        "fp_fact_sum":   sum(f["fp_fact"] for f in proj_finance.values()),
        "fp_plan_sum":   sum(f["fp_plan"] for f in proj_finance.values()),
        "fp_opp_sum":    sum(f["fp_opp"]  for f in proj_finance.values()),
    }

    # ── Дополнительные блоки (Без моих / Мои меньше чужих) ───────────────────
    hierarchy_no_mine   = []
    hours_no_mine       = 0.0
    hierarchy_less_mine = []
    hours_less_mine     = 0.0

    if emp_names:
        proj_mine   = {}
        proj_others = {}
        for r in all_rows:
            emp_up = (r["employee"] or "—").upper()
            proj   = r["project"] or "—"
            proj_mine.setdefault(proj, 0.0)
            proj_others.setdefault(proj, 0.0)
            if emp_up in emp_names:
                proj_mine[proj]   += r["hours"] or 0
            else:
                proj_others[proj] += r["hours"] or 0

        no_mine_projs   = {p for p in proj_others if proj_mine.get(p, 0) == 0 and proj_others[p] > 0}
        less_mine_projs = {p for p in proj_mine   if proj_mine[p] > 0 and proj_mine[p] < proj_others.get(p, 0)}

        rows_no_mine   = [r for r in all_rows if (r["project"] or "—") in no_mine_projs]
        rows_less_mine = [r for r in all_rows if (r["project"] or "—") in less_mine_projs]

        hierarchy_no_mine,   hours_no_mine   = _ts_build_compare_hier(rows_no_mine,   emp_names, proj_finance=proj_finance)
        hierarchy_less_mine, hours_less_mine = _ts_build_compare_hier(rows_less_mine, emp_names, proj_finance=proj_finance)

    return hierarchy, total_hours, totals, hierarchy_no_mine, hours_no_mine, hierarchy_less_mine, hours_less_mine


def _ts_build_compare_hier(rows, emp_names=None, proj_finance=None):
    """
    Иерархия для страницы «Расходы vs Доходы»:
      dept (ПЦ) → project (с финансовыми данными) → employees (с задачами по проекту)
    """
    dept_map   = {}
    dept_order = []

    for r in rows:
        dept    = r["dept"] or "—"
        proj    = r["project"] or "—"
        emp     = r["employee"] or "—"
        task    = r["task"] or ""
        hours   = r["hours"] or 0.0
        is_mine = bool(emp_names and emp.upper() in emp_names)

        dot_pos = task.find(".")
        if dot_pos > 0 and task[:dot_pos].strip().isdigit():
            t_num  = task[:dot_pos].strip()
            t_name = task[dot_pos + 1:].strip()
        else:
            t_num  = ""
            t_name = task

        pm_match = _RE_PROJ_NUM.match(proj.strip())
        p_num = re.sub(r"[\s\xa0]+", "", pm_match.group(1)) if pm_match else ""

        if dept not in dept_map:
            dept_map[dept] = {}
            dept_order.append(dept)

        if proj not in dept_map[dept]:
            fin = (proj_finance or {}).get(proj.lower(), {})
            dept_map[dept][proj] = {
                "hours": 0.0, "my_hours": 0.0, "others_hours": 0.0,
                "client": r["client"] or "", "work_type": r["work_type"] or "",
                "fp_fact":      fin.get("fp_fact", 0.0),
                "fp_plan":      fin.get("fp_plan", 0.0),
                "fp_opp":       fin.get("fp_opp",  0.0),
                "section":      fin.get("section", ""),
                "client_fp":    fin.get("client_fp", ""),
                "ts_client":    fin.get("ts_client", ""),
                "match_type":   fin.get("match_type"),
                "match_score":  fin.get("match_score", 0.0),
                "match_tokens": fin.get("match_tokens", []),
                "emps": {},
            }

        pd = dept_map[dept][proj]
        pd["hours"]        += hours
        pd["my_hours"]     += hours if is_mine else 0.0
        pd["others_hours"] += hours if not is_mine else 0.0

        if emp not in pd["emps"]:
            pd["emps"][emp] = {"hours": 0.0, "tasks": {}}
        pd["emps"][emp]["hours"] += hours

        task_key = (proj, task)
        if task_key not in pd["emps"][emp]["tasks"]:
            pd["emps"][emp]["tasks"][task_key] = {
                "hours": 0.0, "task_num": t_num, "task_name": t_name, "proj_num": p_num,
            }
        pd["emps"][emp]["tasks"][task_key]["hours"] += hours

    hierarchy = []
    total = 0.0
    for dept in dept_order:
        proj_list = sorted(
            [
                {
                    "name":         proj_name,
                    "hours":        pd["hours"],
                    "my_hours":     pd["my_hours"],
                    "others_hours": pd["others_hours"],
                    "client":       pd["client"],
                    "fp_fact":      pd["fp_fact"],
                    "fp_plan":      pd["fp_plan"],
                    "fp_opp":       pd["fp_opp"],
                    "section":      pd["section"],
                    "client_fp":    pd["client_fp"],
                    "ts_client":    pd["ts_client"],
                    "match_type":   pd["match_type"],
                    "match_score":  pd["match_score"],
                    "match_tokens": pd["match_tokens"],
                    "employees": sorted(
                        [
                            {
                                "name":    en,
                                "hours":   ed["hours"],
                                "is_mine": bool(emp_names and en.upper() in emp_names),
                                "tasks":   sorted(
                                    [{"task_num": td["task_num"], "task_name": td["task_name"],
                                      "proj_num": td["proj_num"], "hours": td["hours"]}
                                     for td in ed["tasks"].values()],
                                    key=lambda x: -x["hours"],
                                ),
                            }
                            for en, ed in pd["emps"].items()
                        ],
                        key=lambda x: -x["hours"],
                    ),
                }
                for proj_name, pd in dept_map[dept].items()
            ],
            key=lambda x: -x["hours"],
        )
        dept_hours = sum(p["hours"] for p in proj_list)
        total += dept_hours
        hierarchy.append({"dept": dept, "dept_hours": dept_hours, "projects": proj_list})

    hierarchy.sort(key=lambda x: -x["dept_hours"])
    return hierarchy, total


def _ts_build_hierarchy(rows, emp_names=None, proj_finance=None):
    """
    Строит иерархию из плоских строк трудозатрат:
      dept → division → employees + projects
    emp_names:    set of uppercased my-employee names (для подсчёта my_hours/others_hours).
    proj_finance: dict {lower(project_name): {plan, fact, section, client_fp, ts_client,
                        match_type, match_score, match_tokens}} — для страницы сравнения.
    Возвращает:
      hierarchy: list[dict] dept, dept_hours, divisions
        divisions: list[dict] name, hours, employees, projects
          employees: list[dict] name, hours, tasks: list[{project, task, hours}]
          projects:  list[dict] name, client, hours, my_hours, others_hours, work_type,
                                plan, fact, section, client_fp, ts_client,
                                match_type, match_score, match_tokens
      total_hours: float
    """
    dept_map   = {}
    dept_order = []

    for r in rows:
        dept    = r["dept"] or "—"
        div     = r["division"] or "—"
        proj    = r["project"] or "—"
        emp     = r["employee"] or "—"
        task    = r["task"] or ""
        is_mine = bool(emp_names and emp.upper() in emp_names)

        # Парсим задачу: формат «НОМЕР.НАЗВАНИЕ...»
        dot_pos = task.find(".")
        if dot_pos > 0 and task[:dot_pos].strip().isdigit():
            t_num  = task[:dot_pos].strip()
            t_name = task[dot_pos + 1:].strip()
        else:
            t_num  = ""
            t_name = task
        # Номер проекта из начала поля project
        pm_match = _RE_PROJ_NUM.match(proj.strip())
        p_num = re.sub(r"[\s\xa0]+", "", pm_match.group(1)) if pm_match else ""

        if dept not in dept_map:
            dept_map[dept] = {}
            dept_order.append(dept)
        if div not in dept_map[dept]:
            dept_map[dept][div] = {"proj": {}, "emp": {}}

        # проекты
        pm = dept_map[dept][div]["proj"]
        if proj not in pm:
            fin = (proj_finance or {}).get(proj.lower(), {})
            pm[proj] = {
                "hours": 0.0, "my_hours": 0.0, "others_hours": 0.0,
                "client": r["client"] or "", "work_type": r["work_type"] or "",
                "fp_fact":      fin.get("fp_fact", 0.0),
                "fp_plan":      fin.get("fp_plan", 0.0),
                "fp_opp":       fin.get("fp_opp",  0.0),
                "section":      fin.get("section", ""),
                "client_fp":    fin.get("client_fp", ""),
                "ts_client":    fin.get("ts_client", ""),
                "match_type":   fin.get("match_type"),
                "match_score":  fin.get("match_score", 0.0),
                "match_tokens": fin.get("match_tokens", []),
            }
        pm[proj]["hours"] += r["hours"]
        if is_mine:
            pm[proj]["my_hours"] += r["hours"]
        else:
            pm[proj]["others_hours"] += r["hours"]

        # сотрудники (с задачами)
        em = dept_map[dept][div]["emp"]
        if emp not in em:
            em[emp] = {"hours": 0.0, "tasks": {}}
        em[emp]["hours"] += r["hours"]
        task_key = (proj, task)
        if task_key not in em[emp]["tasks"]:
            em[emp]["tasks"][task_key] = {
                "hours": 0.0,
                "task_num": t_num,
                "task_name": t_name,
                "proj": proj,
                "proj_num": p_num,
            }
        em[emp]["tasks"][task_key]["hours"] += r["hours"]

    hierarchy = []
    total = 0.0
    for dept in dept_order:
        div_list = []
        for div_name, data in dept_map[dept].items():
            proj_list = sorted(
                [{"name": k, "hours": v["hours"],
                  "my_hours": v["my_hours"], "others_hours": v["others_hours"],
                  "client": v["client"], "work_type": v["work_type"],
                  "fp_fact": v["fp_fact"], "fp_plan": v["fp_plan"], "fp_opp": v["fp_opp"],
                  "section": v["section"],
                  "client_fp": v["client_fp"], "ts_client": v["ts_client"],
                  "match_type": v["match_type"], "match_score": v["match_score"],
                  "match_tokens": v["match_tokens"]}
                 for k, v in data["proj"].items()],
                key=lambda x: -x["hours"]
            )
            emp_list = sorted(
                [{"name": k, "hours": v["hours"],
                  "tasks": sorted(
                      [{"project":  td["proj"],
                        "proj_num": td["proj_num"],
                        "task_num": td["task_num"],
                        "task_name": td["task_name"],
                        "hours":    td["hours"]}
                       for td in v["tasks"].values()],
                      key=lambda x: -x["hours"]
                  )}
                 for k, v in data["emp"].items()],
                key=lambda x: -x["hours"]
            )
            div_hours = sum(p["hours"] for p in proj_list)
            div_list.append({
                "name": div_name,
                "hours": div_hours,
                "employees": emp_list,
                "projects": proj_list,
            })

        div_list.sort(key=lambda x: -x["hours"])
        dept_hours = sum(d["hours"] for d in div_list)
        total += dept_hours
        hierarchy.append({"dept": dept, "dept_hours": dept_hours, "divisions": div_list})

    hierarchy.sort(key=lambda x: -x["dept_hours"])
    return hierarchy, total


@app.route("/timesheets")
@login_required
def timesheets():
    conn = get_conn()

    ts_rp_filter = get_setting(conn, "ts_rp_filter") or ""
    ts_pc_filter = get_setting(conn, "ts_pc_filter") or ""

    imports = conn.execute(
        "SELECT * FROM ts_imports ORDER BY imported_at DESC LIMIT 20"
    ).fetchall()

    selected_import_id = request.args.get("import_id", type=int)
    if not selected_import_id and imports:
        selected_import_id = imports[0]["id"]

    hierarchy        = []
    total_hours      = 0.0
    hierarchy_no_mine   = []
    hours_no_mine       = 0.0
    hierarchy_less_mine = []
    hours_less_mine     = 0.0
    current_import   = None
    import_emp_names = set()   # сотрудники в текущем импорте (upper)

    # Справочник «мои сотрудники» (полный список для отображения/редактирования)
    employees = conn.execute(
        "SELECT id, full_name FROM my_employees ORDER BY full_name"
    ).fetchall()
    emp_names = {e["full_name"].upper() for e in employees}

    hierarchy_product   = []
    hours_product       = 0.0

    if selected_import_id:
        current_import = conn.execute(
            "SELECT * FROM ts_imports WHERE id = ?", (selected_import_id,)
        ).fetchone()
        rows = conn.execute(
            "SELECT * FROM ts_rows WHERE import_id = ? ORDER BY id",
            (selected_import_id,),
        ).fetchall()

        # ── Разделяем строки: "по команде" vs "по продукту (другая команда)" ──
        rp_filter_val = (ts_rp_filter or "").strip()
        if rp_filter_val:
            rows_product_only = [
                r for r in rows
                if r["rp_product"] is not None
                and (r["rp_product"] or "").strip() == rp_filter_val
                and (r["rp"] or "").strip() != rp_filter_val
            ]
            rows_main = [r for r in rows if r not in rows_product_only]
        else:
            rows_main         = rows
            rows_product_only = []

        hierarchy, total_hours = _ts_build_hierarchy(rows_main, emp_names)
        if rows_product_only:
            hierarchy_product, hours_product = _ts_build_hierarchy(rows_product_only, emp_names)

        import_emp_names = {r["employee"].upper() for r in rows if r["employee"]}

        # Классифицируем проекты: мои vs чужие (только из rows_main)
        if emp_names:
            proj_mine   = {}
            proj_others = {}
            for r in rows_main:
                emp_up = (r["employee"] or "—").upper()
                proj   = r["project"] or "—"
                proj_mine.setdefault(proj, 0.0)
                proj_others.setdefault(proj, 0.0)
                if emp_up in emp_names:
                    proj_mine[proj]   += r["hours"]
                else:
                    proj_others[proj] += r["hours"]

            # Блок 2: проекты без моих сотрудников
            no_mine_projs = {p for p in proj_others if proj_mine.get(p, 0) == 0 and proj_others[p] > 0}
            # Блок 3: мои > 0, но мои < чужих
            less_mine_projs = {
                p for p in proj_mine
                if proj_mine[p] > 0 and proj_mine[p] < proj_others.get(p, 0)
            }

            rows_no_mine   = [r for r in rows_main if (r["project"] or "—") in no_mine_projs]
            rows_less_mine = [r for r in rows_main if (r["project"] or "—") in less_mine_projs]
            hierarchy_no_mine,   hours_no_mine   = _ts_build_hierarchy(rows_no_mine,   emp_names)
            hierarchy_less_mine, hours_less_mine = _ts_build_hierarchy(rows_less_mine, emp_names)

    conn.close()
    return render_template(
        "timesheets.html",
        imports=imports,
        selected_import_id=selected_import_id,
        current_import=current_import,
        hierarchy=hierarchy,
        total_hours=total_hours,
        hierarchy_no_mine=hierarchy_no_mine,
        hours_no_mine=hours_no_mine,
        hierarchy_less_mine=hierarchy_less_mine,
        hours_less_mine=hours_less_mine,
        hierarchy_product=hierarchy_product,
        hours_product=hours_product,
        employees=employees,
        emp_names=emp_names,
        import_emp_names=import_emp_names,
        ts_rp_filter=ts_rp_filter,
        ts_pc_filter=ts_pc_filter,
    )


@app.route("/timesheets/import", methods=["POST"])
@login_required
@owner_required
def ts_import():
    f = request.files.get("ts_file")
    if not f or not f.filename:
        flash("Выберите файл", "error")
        return redirect(url_for("timesheets"))

    filename = secure_filename(f.filename)
    tmp_path = UPLOAD_DIR / filename
    f.save(tmp_path)

    conn_cfg = get_conn()
    ts_rp_filter = get_setting(conn_cfg, "ts_rp_filter") or None
    ts_pc_filter = get_setting(conn_cfg, "ts_pc_filter") or None
    conn_cfg.close()

    try:
        rows, file_type = parse_ts_file(str(tmp_path), rp_filter=ts_rp_filter, pc_filter=ts_pc_filter)
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        app.logger.error("ts_import error:\n%s", tb)
        flash(f"Ошибка разбора файла: {e} | {tb[-300:]}", "error")
        return redirect(url_for("timesheets"))

    # Извлекаем метку периода из имени файла (часть между # и предпоследним словом)
    period_label = None
    import re as _re
    m = _re.search(r"#\d+\s+(.+?)(?:\s+(?:предварительный|ДЛЯ\s+РП|итог))", f.filename, _re.IGNORECASE)
    if m:
        period_label = m.group(1).strip()
    if not period_label:
        # Попытка 2: вытащить что-то похожее на даты
        m2 = _re.search(r"(\d{2}_\w+_\d{4}\s*-\s*\d{2}_\w+_\d{4})", f.filename)
        if m2:
            period_label = m2.group(1).strip()
    if not period_label:
        period_label = filename

    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO ts_imports (filename, period_label, imported_by, rows_total, rp_filter, pc_filter, file_type) VALUES (?,?,?,?,?,?,?)",
        (filename, period_label, session.get("user_id"), len(rows), ts_rp_filter, ts_pc_filter, file_type),
    )
    import_id = cur.lastrowid

    for r in rows:
        cur.execute(
            """INSERT INTO ts_rows
               (import_id, rp, rp_product, dept, division, employee, project, task,
                client, project_type, work_type, hours)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (import_id, r["rp"], r.get("rp_product"), r["dept"], r["division"], r["employee"],
             r["project"], r["task"], r["client"],
             r["project_type"], r["work_type"], r["hours"]),
        )

    conn.commit()
    conn.close()
    type_label = "Детализация до задачи" if file_type == "detail" else "Отчёт проектного офиса"
    flash(f"Импортировано {len(rows)} строк за период «{period_label}» [{type_label}]", "success")
    return redirect(url_for("timesheets", import_id=import_id))


@app.route("/timesheets/set-filters", methods=["POST"])
@login_required
@owner_required
def ts_set_filters():
    rp = request.form.get("ts_rp_filter", "").strip()
    pc = request.form.get("ts_pc_filter", "").strip()
    conn = get_conn()
    set_setting(conn, "ts_rp_filter", rp)
    set_setting(conn, "ts_pc_filter", pc)
    conn.commit()
    conn.close()
    parts = []
    if rp:
        parts.append(f"РП: {rp}")
    if pc:
        parts.append(f"ПЦ: {pc}")
    msg = "Фильтры сохранены" + (": " + ", ".join(parts) if parts else " (пусто — загружается всё)")
    flash(msg, "success")
    return redirect(url_for("import_view"))


@app.route("/timesheets/employees", methods=["POST"])
@login_required
@owner_required
def ts_import_employees():
    f = request.files.get("emp_file")
    if not f or not f.filename:
        flash("Выберите файл", "error")
        return redirect(url_for("timesheets"))

    filename = secure_filename(f.filename)
    tmp_path = UPLOAD_DIR / filename
    f.save(tmp_path)

    try:
        names = parse_employees_file(str(tmp_path))
    except Exception as e:
        flash(f"Ошибка разбора файла: {e}", "error")
        return redirect(url_for("timesheets"))

    conn = get_conn()
    added = 0
    for name in names:
        try:
            conn.execute("INSERT INTO my_employees (full_name) VALUES (?)", (name,))
            added += 1
        except Exception:
            pass  # дубликат
    conn.commit()
    conn.close()
    flash(f"Загружено сотрудников: {added} из {len(names)}", "success")
    return redirect(url_for("timesheets"))


@app.route("/timesheets/employees/add", methods=["POST"])
@login_required
@owner_required
def ts_add_employee():
    name = (request.form.get("full_name") or "").strip()
    if not name:
        flash("Введите имя сотрудника", "error")
        return redirect(url_for("timesheets"))
    conn = get_conn()
    try:
        conn.execute("INSERT INTO my_employees (full_name) VALUES (?)", (name,))
        conn.commit()
        flash(f"Сотрудник «{name}» добавлен", "success")
    except Exception:
        flash(f"Сотрудник «{name}» уже есть в списке", "error")
    conn.close()
    return redirect(url_for("timesheets"))


@app.route("/timesheets/employees/<int:emp_id>/delete", methods=["POST"])
@login_required
@owner_required
def ts_delete_employee(emp_id):
    conn = get_conn()
    emp = conn.execute("SELECT full_name FROM my_employees WHERE id = ?", (emp_id,)).fetchone()
    if emp:
        conn.execute("DELETE FROM my_employees WHERE id = ?", (emp_id,))
        conn.commit()
        flash(f"Сотрудник «{emp['full_name']}» удалён", "success")
    conn.close()
    return redirect(url_for("timesheets"))


@app.route("/timesheets/employees/clear", methods=["POST"])
@login_required
@owner_required
def ts_clear_employees():
    conn = get_conn()
    conn.execute("DELETE FROM my_employees")
    conn.commit()
    conn.close()
    flash("Список сотрудников очищен", "success")
    return redirect(url_for("timesheets"))


@app.route("/timesheets/import/<int:import_id>/delete", methods=["POST"])
@login_required
@owner_required
def ts_delete_import(import_id):
    conn = get_conn()
    conn.execute("DELETE FROM ts_imports WHERE id = ?", (import_id,))
    conn.commit()
    conn.close()
    flash("Импорт удалён", "success")
    return redirect(url_for("timesheets"))


@app.route("/timesheets/compare")
@login_required
def ts_compare():
    conn = get_conn()
    imports = conn.execute(
        "SELECT * FROM ts_imports ORDER BY imported_at DESC LIMIT 20"
    ).fetchall()

    selected_import_id = request.args.get("import_id", type=int)
    if not selected_import_id and imports:
        selected_import_id = imports[0]["id"]

    employees = conn.execute("SELECT * FROM my_employees ORDER BY full_name").fetchall()
    emp_names = {e["full_name"].upper() for e in employees}

    # Квартал ФП для сравнения: из GET-параметра fp_quarter, иначе текущий view_quarter
    fp_quarter = request.args.get("fp_quarter", "").strip()
    if not fp_quarter:
        fp_quarter = get_view_quarter(conn)

    # Квартал "все" позволяет сравнивать без фильтра
    if fp_quarter == "__all__":
        fp_quarter_labels = None
    else:
        fp_quarter_labels = [fp_quarter] if fp_quarter else None

    hierarchy = []
    total_hours = 0.0
    totals = {"total_hours": 0, "matched_hours": 0, "fp_fact_sum": 0, "fp_plan_sum": 0, "fp_opp_sum": 0}
    hierarchy_no_mine   = []
    hours_no_mine       = 0.0
    hierarchy_less_mine = []
    hours_less_mine     = 0.0
    current_import = None
    if selected_import_id:
        current_import = conn.execute(
            "SELECT * FROM ts_imports WHERE id=?", (selected_import_id,)
        ).fetchone()
        (hierarchy, total_hours, totals,
         hierarchy_no_mine, hours_no_mine,
         hierarchy_less_mine, hours_less_mine) = _ts_build_compare(
            selected_import_id, conn, emp_names,
            fp_quarter_labels=fp_quarter_labels,
        )

    daily_rate = float(get_setting(conn, "ts_daily_rate") or 0)
    cmp_rp_filter = get_setting(conn, "ts_rp_filter") or ""
    fp_quarters = get_available_quarters(conn)
    conn.close()
    return render_template(
        "ts_compare.html",
        imports=imports,
        selected_import_id=selected_import_id,
        current_import=current_import,
        hierarchy=hierarchy,
        total_hours=total_hours,
        totals=totals,
        daily_rate=daily_rate,
        emp_names=emp_names,
        hierarchy_no_mine=hierarchy_no_mine,
        hours_no_mine=hours_no_mine,
        hierarchy_less_mine=hierarchy_less_mine,
        hours_less_mine=hours_less_mine,
        rp_filter=cmp_rp_filter,
        fp_quarter=fp_quarter,
        fp_quarters=fp_quarters,
    )


@app.route("/timesheets/compare/set-rate", methods=["POST"])
@login_required
@owner_required
def ts_set_rate():
    raw = (request.form.get("rate") or "").strip().replace(",", ".").replace(" ", "")
    try:
        rate = float(raw)
        if rate < 0:
            raise ValueError
        conn = get_conn()
        set_setting(conn, "ts_daily_rate", str(rate))
        conn.commit()
        conn.close()
        flash(f"Ставка обновлена: {rate:,.0f} ₽/день".replace(",", " "), "success")
    except ValueError:
        flash("Некорректное значение ставки", "error")
    return redirect(url_for("ts_compare",
                            import_id=request.form.get("import_id", type=int)))


# ─── Трудозатраты: аналитика по времени ──────────────────────────────────────

@app.route("/timesheets/analytics")
@login_required
def ts_analytics():
    conn = get_conn()
    imports = conn.execute(
        "SELECT * FROM ts_imports ORDER BY imported_at ASC LIMIT 50"
    ).fetchall()
    employees = conn.execute("SELECT full_name FROM my_employees").fetchall()
    emp_names = {e["full_name"].upper() for e in employees}

    periods = []
    for imp in imports:
        rows = conn.execute(
            "SELECT employee, SUM(hours) h FROM ts_rows WHERE import_id = ? GROUP BY employee",
            (imp["id"],),
        ).fetchall()
        total_h = sum(r["h"] or 0 for r in rows)
        my_h = sum(r["h"] or 0 for r in rows if (r["employee"] or "").upper() in emp_names)
        periods.append({
            "id": imp["id"],
            "label": imp["period_label"] or imp["filename"],
            "total_hours": total_h,
            "my_hours": my_h,
            "others_hours": total_h - my_h,
            "rows_total": imp["rows_total"] or 0,
        })
    conn.close()
    max_h = max((p["total_hours"] for p in periods), default=1) or 1
    return render_template("ts_analytics.html", periods=periods, emp_names=emp_names, max_h=max_h)


# ─── Трудозатраты: сравнение двух периодов ───────────────────────────────────

@app.route("/timesheets/diff")
@login_required
def ts_diff():
    conn = get_conn()
    imports = conn.execute(
        "SELECT * FROM ts_imports ORDER BY imported_at DESC LIMIT 50"
    ).fetchall()
    id_a = request.args.get("a", type=int)
    id_b = request.args.get("b", type=int)

    employees = conn.execute("SELECT full_name FROM my_employees").fetchall()
    emp_names = {e["full_name"].upper() for e in employees}

    import_a = import_b = None
    emp_diff = []
    proj_diff = []

    if id_a and id_b and id_a != id_b:
        import_a = conn.execute("SELECT * FROM ts_imports WHERE id = ?", (id_a,)).fetchone()
        import_b = conn.execute("SELECT * FROM ts_imports WHERE id = ?", (id_b,)).fetchone()

        def _agg(import_id, field):
            return {
                (r[field] or "—"): (r["h"] or 0)
                for r in conn.execute(
                    f"SELECT {field}, SUM(hours) h FROM ts_rows WHERE import_id = ? GROUP BY {field}",
                    (import_id,),
                ).fetchall()
            }

        emp_a = _agg(id_a, "employee")
        emp_b = _agg(id_b, "employee")
        for name in sorted(set(emp_a) | set(emp_b)):
            ha, hb = emp_a.get(name, 0), emp_b.get(name, 0)
            emp_diff.append({
                "name": name, "hours_a": ha, "hours_b": hb,
                "delta": hb - ha, "is_mine": name.upper() in emp_names,
            })
        emp_diff.sort(key=lambda x: -abs(x["delta"]))

        proj_a = _agg(id_a, "project")
        proj_b = _agg(id_b, "project")
        for name in sorted(set(proj_a) | set(proj_b)):
            ha, hb = proj_a.get(name, 0), proj_b.get(name, 0)
            proj_diff.append({"name": name, "hours_a": ha, "hours_b": hb, "delta": hb - ha})
        proj_diff.sort(key=lambda x: -abs(x["delta"]))

    conn.close()
    return render_template(
        "ts_diff.html",
        imports=imports, id_a=id_a, id_b=id_b,
        import_a=import_a, import_b=import_b,
        emp_diff=emp_diff, proj_diff=proj_diff,
        emp_names=emp_names,
    )


# ─── Трудозатраты: поиск по сотруднику ───────────────────────────────────────

@app.route("/timesheets/employee")
@login_required
def ts_employee():
    conn = get_conn()
    q = (request.args.get("q") or "").strip()

    all_emps = [
        r["employee"] for r in conn.execute(
            "SELECT DISTINCT employee FROM ts_rows "
            "WHERE employee IS NOT NULL AND employee != '' ORDER BY employee"
        ).fetchall()
    ]
    my_employees = conn.execute("SELECT full_name FROM my_employees ORDER BY full_name").fetchall()
    emp_names = {e["full_name"].upper() for e in my_employees}

    matched_name = ""
    timeline = []

    if q:
        q_up = q.upper()
        matched_name = next((e for e in all_emps if e.upper() == q_up), None)
        if not matched_name:
            matched_name = next((e for e in all_emps if q_up in e.upper()), None)
        if matched_name:
            imports = conn.execute(
                "SELECT i.* FROM ts_imports i "
                "WHERE EXISTS (SELECT 1 FROM ts_rows r WHERE r.import_id = i.id AND r.employee = ?) "
                "ORDER BY i.imported_at ASC",
                (matched_name,),
            ).fetchall()
            for imp in imports:
                rows = conn.execute(
                    "SELECT * FROM ts_rows WHERE import_id = ? AND employee = ? ORDER BY project, task",
                    (imp["id"], matched_name),
                ).fetchall()
                timeline.append({
                    "import": imp,
                    "rows": rows,
                    "total_hours": sum(r["hours"] or 0 for r in rows),
                })

    conn.close()
    return render_template(
        "ts_employee.html",
        q=q, all_emps=all_emps, matched_name=matched_name,
        timeline=timeline, emp_names=emp_names,
    )


# ─── Трудозатраты: выгрузка в Excel ──────────────────────────────────────────

@app.route("/timesheets/export")
@login_required
def ts_export():
    import_id = request.args.get("import_id", type=int)
    if not import_id:
        flash("Выберите период для выгрузки", "error")
        return redirect(url_for("timesheets"))

    conn = get_conn()
    imp = conn.execute("SELECT * FROM ts_imports WHERE id = ?", (import_id,)).fetchone()
    if not imp:
        conn.close()
        flash("Импорт не найден", "error")
        return redirect(url_for("timesheets"))

    rows = conn.execute(
        "SELECT * FROM ts_rows WHERE import_id = ? ORDER BY dept, division, employee, project",
        (import_id,),
    ).fetchall()
    conn.close()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Трудозатраты"

    period = imp["period_label"] or imp["filename"]
    headers = ["Период", "РП", "Департамент", "Подразделение", "Сотрудник",
               "Проект", "Задача", "Клиент", "Тип проекта", "Вид работ", "Часы"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="2563A8")

    for i, r in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=period)
        ws.cell(row=i, column=2, value=r["rp"])
        ws.cell(row=i, column=3, value=r["dept"])
        ws.cell(row=i, column=4, value=r["division"])
        ws.cell(row=i, column=5, value=r["employee"])
        ws.cell(row=i, column=6, value=r["project"])
        ws.cell(row=i, column=7, value=r["task"])
        ws.cell(row=i, column=8, value=r["client"])
        ws.cell(row=i, column=9, value=r["project_type"])
        ws.cell(row=i, column=10, value=r["work_type"])
        ws.cell(row=i, column=11, value=r["hours"])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 55)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe = re.sub(r"[^\w\-]", "_", period)
    from flask import send_file
    return send_file(
        buf, as_attachment=True,
        download_name=f"ts_{safe}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    init_db()
    app.run(host="127.0.0.1", port=5000, debug=True)
