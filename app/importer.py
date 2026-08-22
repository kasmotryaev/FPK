import openpyxl
import datetime
import json
from collections import defaultdict
from app.db import get_conn, row_key, get_setting, set_setting, compute_quarter_label

# Обязательные колонки файла (проверяются по имени, не по позиции).
# Новый формат файла (23 кол.) включает дополнительные столбцы относительно старого (20 кол.),
# поэтому парсер использует поиск по имени заголовка, а не фиксированные индексы.
REQUIRED_HEADERS = {
    "Месяц", "ПЦ", "Раздел ФП", "Наименование клиента",
    "Сумма по 0-100, руб",          # новый формат (без точки)
    "Портфель",
}

# Список заголовков нового формата — для справки и документации
EXPECTED_HEADERS = [
    "Месяц", "Стратегическое решение", "ПЦ", "РП", "Раздел ФП",
    "Наименование клиента", "Номер проекта", "Наименование проекта", "Менеджер проекта",
    "Номер договора", "Номер ЭЗ", "СДЗ", "ДПА", "ФДЗ", "Способ учета",
    "0-100 от МП", "Комментарий МП к СДЗ", "Сумма CRM, руб", "Сумма по 0-100, руб",
    "Примечание", "Портфель", "Колодец", "Внешний ID клиента",
]

# Поля, изменения которых стоит показывать пользователю в сводке загрузки
TRACKED_FIELDS = [
    ("amount_0_100", "Сумма по 0-100"),
    ("portfolio", "Портфель"),
    ("dpa_date", "ДПА"),
    ("sdz_date", "СДЗ"),
    ("fdz_date", "фДЗ"),
    ("crm_amount", "Сумма из CRM"),
    ("project_manager", "Менеджер проекта"),
    ("strategic_solution", "Стратегическое решение"),
]


def _parse_date(val):
    if val is None or val == "":
        return None
    if isinstance(val, datetime.datetime):
        return val.date().isoformat()
    if isinstance(val, datetime.date):
        return val.isoformat()
    s = str(val).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return s or None


def _num(val):
    if val is None or val == "":
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _normalized_identity_part(value):
    """Normalize human-entered identity fields used for event reconciliation."""
    return " ".join(str(value or "").split()).casefold()


def _project_label(event):
    """Build a readable project value for the change history."""
    number = str(event.get("project_num") or "").strip()
    name = str(event.get("project_name") or "").strip()
    if number and name and number not in name:
        return f"{number} - {name}"
    return name or number or "-"


def _collapse_portfolio_changes(events):
    """Collapse a forecast/new pair even when its month changed at the same time."""
    forecast_portfolios = {"0-100", "\u0412\u043e\u0437\u043c\u043e\u0436\u043d\u043e\u0441\u0442\u0438"}
    fact_portfolio = "\u0424\u0430\u043a\u0442"
    new_events = [
        event for event in events
        if event["event_type"] == "new" and (event.get("portfolio") or "") == fact_portfolio
    ]
    deactivated_events = [
        event for event in events
        if event["event_type"] == "deactivated"
        and (event.get("portfolio") or "") in forecast_portfolios
    ]
    candidates = []

    for old_event in deactivated_events:
        for new_event in new_events:
            same_identity = (
                _normalized_identity_part(old_event.get("client_name"))
                == _normalized_identity_part(new_event.get("client_name"))
                and _normalized_identity_part(_project_label(old_event))
                == _normalized_identity_part(_project_label(new_event))
                and _normalized_identity_part(old_event.get("section"))
                == _normalized_identity_part(new_event.get("section"))
            )
            if not same_identity:
                continue

            amount_delta = abs(
                (old_event.get("amount_before") or 0)
                - (new_event.get("amount_after") or 0)
            )
            if amount_delta > 1.0:
                continue

            old_contract = _normalized_identity_part(old_event.get("contract_num"))
            new_contract = _normalized_identity_part(new_event.get("contract_num"))
            contract_mismatch = int(not old_contract or old_contract != new_contract)
            candidates.append((contract_mismatch, amount_delta, old_event, new_event))

    # Prefer exact contracts and closest amounts; consume each row only once.
    candidates.sort(key=lambda item: (item[0], item[1], item[2]["fp_row_id"], item[3]["fp_row_id"]))
    consumed_ids = set()
    portfolio_changed_events = []
    for _, _, old_event, new_event in candidates:
        if id(old_event) in consumed_ids or id(new_event) in consumed_ids:
            continue
        consumed_ids.update((id(old_event), id(new_event)))
        portfolio_changed_events.append(dict(
            fp_row_id=new_event["fp_row_id"],
            event_type="portfolio_changed",
            field_label="portfolio",
            old_value=old_event.get("portfolio"),
            new_value=new_event.get("portfolio"),
            amount_before=old_event.get("amount_before"),
            amount_after=new_event.get("amount_after"),
            month=new_event.get("month"),
            client_name=new_event.get("client_name"),
            project_name=new_event.get("project_name"),
            section=new_event.get("section"),
            portfolio=new_event.get("portfolio"),
            contract_num=new_event.get("contract_num"),
        ))

    if not consumed_ids:
        return events
    return [event for event in events if id(event) not in consumed_ids] + portfolio_changed_events


def _collapse_project_moves(events):
    """Replace matching new/deactivated pairs with one project-change event."""
    new_events = [event for event in events if event["event_type"] == "new"]
    deactivated_events = [event for event in events if event["event_type"] == "deactivated"]
    candidates = []

    for old_event in deactivated_events:
        for new_event in new_events:
            same_context = (
                _normalized_identity_part(old_event.get("client_name"))
                == _normalized_identity_part(new_event.get("client_name"))
                and _normalized_identity_part(old_event.get("month"))
                == _normalized_identity_part(new_event.get("month"))
                and _normalized_identity_part(old_event.get("section"))
                == _normalized_identity_part(new_event.get("section"))
                and _normalized_identity_part(old_event.get("portfolio"))
                == _normalized_identity_part(new_event.get("portfolio"))
            )
            if not same_context:
                continue

            old_project = _project_label(old_event)
            new_project = _project_label(new_event)
            if _normalized_identity_part(old_project) == _normalized_identity_part(new_project):
                continue

            amount_delta = abs(
                (old_event.get("amount_before") or 0)
                - (new_event.get("amount_after") or 0)
            )
            if amount_delta > 1.0:
                continue

            old_contract = _normalized_identity_part(old_event.get("contract_num"))
            new_contract = _normalized_identity_part(new_event.get("contract_num"))
            contract_mismatch = int(not old_contract or old_contract != new_contract)
            candidates.append((contract_mismatch, amount_delta, old_event, new_event))

    # Prefer exact contracts and closest amounts; consume each row only once.
    candidates.sort(key=lambda item: (item[0], item[1], item[2]["fp_row_id"], item[3]["fp_row_id"]))
    consumed_ids = set()
    project_changed_events = []
    for _, _, old_event, new_event in candidates:
        if id(old_event) in consumed_ids or id(new_event) in consumed_ids:
            continue
        consumed_ids.update((id(old_event), id(new_event)))
        project_changed_events.append(dict(
            fp_row_id=new_event["fp_row_id"], event_type="field_changed",
            field_label="\u041f\u0440\u043e\u0435\u043a\u0442 \u0438\u0437\u043c\u0435\u043d\u0438\u043b\u0441\u044f",
            old_value=_project_label(old_event), new_value=_project_label(new_event),
            amount_before=old_event["amount_before"], amount_after=new_event["amount_after"],
            month=new_event.get("month"), client_name=new_event.get("client_name"),
            project_name=new_event.get("project_name"), section=new_event.get("section"),
            portfolio=new_event.get("portfolio"), contract_num=new_event.get("contract_num"),
        ))

    if not consumed_ids:
        return events
    return [event for event in events if id(event) not in consumed_ids] + project_changed_events


def _build_col_map(ws):
    """Строит маппинг «имя заголовка → номер столбца (1-based)» по первой строке файла."""
    col_map = {}
    for c in range(1, ws.max_column + 1):
        hdr = ws.cell(row=1, column=c).value
        if hdr is not None:
            col_map[str(hdr).strip()] = c
    return col_map


def _get(ws, row, col_map, *names):
    """Возвращает значение первой найденной колонки из списка имён (поддержка старых/новых названий)."""
    for name in names:
        c = col_map.get(name)
        if c is not None:
            return ws.cell(row=row, column=c).value
    return None


def detect_header_mismatch(ws):
    """Проверяет наличие обязательных колонок по имени (не по позиции).
    Возвращает список отсутствующих заголовков; пустой список — файл корректен."""
    col_map = _build_col_map(ws)
    # «Сумма по 0-100» может называться по-разному в старом и новом форматах
    actual_names = set(col_map.keys())
    missing = []
    for req in REQUIRED_HEADERS:
        # для суммы проверяем оба варианта написания
        if req == "Сумма по 0-100, руб":
            if "Сумма по 0-100, руб" not in actual_names and "Сумма по 0-100, руб." not in actual_names:
                missing.append(req)
        elif req not in actual_names:
            missing.append(req)
    return missing


def import_excel(filepath, filename, user_id, quarter_label=None, rp_filter=None):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    mismatches = detect_header_mismatch(ws)
    col_map = _build_col_map(ws)

    conn = get_conn()
    cur = conn.cursor()

    seen_keys = set()
    rows_new = 0
    rows_updated = 0
    rows_unchanged = 0

    diff_new = []        # новые строки
    diff_changed = []     # изменённые поля по существующим строкам
    diff_deactivated = []  # строки, пропавшие из файла

    # Постоянный журнал событий (row_events) — в отличие от diff_*, не привязан к одной
    # загрузке и копится между ними; пишется после того, как известен import_log_id.
    pending_events = []

    # occurrence-счётчик для устойчивости к содержательным дублям внутри одной загрузки
    occurrence_counter = defaultdict(int)

    # ---------- Определяем квартал загрузки ----------
    # Если вызывающий явно передал quarter_label (из формы загрузки), используем его.
    # Иначе берём current_quarter_label из настроек (обратная совместимость).
    stored_quarter_label = get_setting(conn, "current_quarter_label")
    if stored_quarter_label is None:
        stored_quarter_label = compute_quarter_label(datetime.date.today())
        set_setting(conn, "current_quarter_label", stored_quarter_label)
        conn.commit()

    upload_ql = (quarter_label or "").strip() or stored_quarter_label

    # ---------- Перенос обязательств/примечаний при первой загрузке нового квартала ----------
    # Если для upload_ql ещё нет активных строк — это первый файл нового квартала.
    # Тогда ищем строки 0-100/Возможности из ПРЕДЫДУЩЕГО квартала и копируем на совпадающие
    # (project_num, section) в новом файле. Это срабатывает ровно один раз — пока
    # пользователь не загрузит хотя бы одну строку с upload_ql.
    is_new_quarter = not cur.execute(
        "SELECT 1 FROM fp_rows WHERE is_active=1 AND quarter_label=? LIMIT 1",
        (upload_ql,),
    ).fetchone()

    carry_over_map = {}
    rollover_summary = None
    if is_new_quarter:
        # Находим самый свежий другой квартал с активными строками
        prev_q_row = cur.execute(
            """SELECT quarter_label FROM fp_rows
               WHERE is_active=1 AND quarter_label IS NOT NULL AND quarter_label != ''
                 AND quarter_label != ?
               ORDER BY quarter_label DESC LIMIT 1""",
            (upload_ql,),
        ).fetchone()
        if prev_q_row:
            prev_ql = prev_q_row["quarter_label"]
            cur.execute(
                """SELECT * FROM fp_rows
                   WHERE is_active = 1 AND quarter_label = ?
                     AND portfolio IN ('0-100', 'Возможности')
                     AND project_num IS NOT NULL AND project_num != ''
                     AND section IS NOT NULL AND section != ''""",
                (prev_ql,),
            )
            old_rows = cur.fetchall()
            old_row_ids = [r["id"] for r in old_rows]
            old_obls_by_row = defaultdict(list)
            if old_row_ids:
                ph = ",".join("?" * len(old_row_ids))
                cur.execute(f"SELECT * FROM obligations WHERE fp_row_id IN ({ph})", old_row_ids)
                for o in cur.fetchall():
                    old_obls_by_row[o["fp_row_id"]].append(o)
            for r in old_rows:
                entry = carry_over_map.setdefault(
                    (r["project_num"], r["section"]),
                    {"comments": [], "obligations": [], "responsible_user_id": None},
                )
                if r["internal_comment"]:
                    entry["comments"].append(r["internal_comment"])
                entry["obligations"].extend(old_obls_by_row.get(r["id"], []))
                if entry["responsible_user_id"] is None and r["responsible_user_id"] is not None:
                    entry["responsible_user_id"] = r["responsible_user_id"]
            rollover_summary = {
                "old_label": prev_ql, "new_label": upload_ql,
                "candidates": len(carry_over_map), "matched": 0,
                "obligations_copied": 0, "comments_copied": 0,
            }

    # ---------- Предварительный снимок пользовательских данных текущего квартала ----------
    # Если квартал уже существует (is_new_quarter=False), строки могут поменять row_key между
    # загрузками (незначительное изменение ключевого поля в файле). Тогда старая строка
    # деактивируется и создаётся новая — без комментариев/рисков/обязательств.
    # Чтобы этого не происходило, делаем снимок по (project_num, section) ДО основного цикла.
    same_q_user_data = {}  # (project_num, section) → {internal_comment, risk_level, ...}
    if not is_new_quarter:
        cur.execute("""
            SELECT id, project_num, section, internal_comment, risk_level, responsible_user_id
            FROM fp_rows
            WHERE is_active = 1 AND quarter_label = ?
        """, (upload_ql,))
        _sq_rows = cur.fetchall()
        _sq_ids = [r_["id"] for r_ in _sq_rows]
        _sq_obls = defaultdict(list)
        if _sq_ids:
            _ph = ",".join("?" * len(_sq_ids))
            cur.execute(f"SELECT * FROM obligations WHERE fp_row_id IN ({_ph})", _sq_ids)
            for _o in cur.fetchall():
                _sq_obls[_o["fp_row_id"]].append(_o)
        for _r in _sq_rows:
            pn = str(_r["project_num"] or "").strip()
            sc = str(_r["section"] or "").strip()
            if pn and sc:
                same_q_user_data.setdefault(
                    (pn, sc),
                    {
                        "internal_comment": _r["internal_comment"],
                        "risk_level": _r["risk_level"],
                        "responsible_user_id": _r["responsible_user_id"],
                        "obligations": _sq_obls.get(_r["id"], []),
                    },
                )

    for r in range(2, ws.max_row + 1):
        # Читаем по именам заголовков — работает и со старым (20 кол.), и с новым (23 кол.) форматом
        g = lambda *names: _get(ws, r, col_map, *names)

        month        = g("Месяц")
        pc           = g("ПЦ")
        section      = g("Раздел ФП")
        client       = g("Наименование клиента")
        proj_num     = g("Номер проекта")
        proj_name    = g("Наименование проекта")
        manager      = g("Менеджер проекта")
        contract     = g("Номер договора")
        ez           = g("Номер ЭЗ")
        sdz          = g("СДЗ")
        dpa          = g("ДПА")
        fdz          = g("ФДЗ", "фДЗ")           # оба регистра
        accounting   = g("Способ учета")
        mp_0_100     = g("0-100 от МП")
        mp_comment   = g("Комментарий МП к СДЗ")
        crm_amount   = g("Сумма CRM, руб", "Сумма из СRM, руб.")  # старый и новый варианты
        amount_0_100 = g("Сумма по 0-100, руб", "Сумма по 0-100, руб.")
        note         = g("Примечание")
        portfolio    = g("Портфель")
        kolodec      = g("Колодец")
        # «Стратегическое решение» — блок стратегии, к которому отнесена статья ФП
        # (например «Решение А»). В старом формате файла колонки нет — тогда None.
        strategic    = g("Стратегическое решение")
        # Новые колонки (пока не используются в БД, но не вызывают ошибок)
        # g("РП"), g("Внешний ID клиента")

        if all(v is None or v == "" for v in (month, client, proj_num, section, amount_0_100, portfolio)):
            continue

        # Фильтр по РП из колонки «РП» нового формата файла.
        # Если фильтр задан и колонка присутствует — пропускаем строки других РП.
        # Если колонки «РП» нет в файле (старый формат) — фильтр не применяется.
        if rp_filter and "РП" in col_map:
            row_rp = g("РП")
            if row_rp is None or str(row_rp).strip() != rp_filter:
                continue

        base_key_tuple = (month, client, proj_num, section, contract, ez, portfolio, accounting)
        occurrence = occurrence_counter[base_key_tuple]
        occurrence_counter[base_key_tuple] += 1

        # quarter_label включён в хэш (row_key v2): ключ уникален в рамках одного квартала,
        # строки разных кварталов никогда не коллидируют по row_key.
        key = row_key(month, client, proj_num, section, contract, ez, portfolio, accounting, occurrence, upload_ql)
        seen_keys.add(key)

        cur.execute("SELECT * FROM fp_rows WHERE row_key = ?", (key,))
        existing = cur.fetchone()

        data = dict(
            row_key=key, month=month, pc=pc, section=section, client_name=client,
            project_num=proj_num, project_name=proj_name, project_manager=manager,
            contract_num=contract, ez_num=ez,
            sdz_date=_parse_date(sdz), dpa_date=_parse_date(dpa), fdz_date=_parse_date(fdz),
            accounting_entity=accounting, mp_0_100=int(mp_0_100) if mp_0_100 not in (None, "") else None,
            mp_comment=mp_comment, crm_amount=_num(crm_amount), amount_0_100=_num(amount_0_100),
            note=note, portfolio=portfolio, kolodec=kolodec, strategic_solution=strategic,
            is_active=1, quarter_label=upload_ql,
        )

        if existing:
            changed_fields = []
            for field, label in TRACKED_FIELDS:
                old_val = existing[field]
                new_val = data[field]
                # сравнение чисел с допуском на погрешность float
                if isinstance(new_val, float) and isinstance(old_val, (int, float)):
                    is_diff = abs((old_val or 0) - new_val) > 0.01
                else:
                    is_diff = old_val != new_val
                if is_diff:
                    changed_fields.append({"field": label, "key": field, "old": old_val, "new": new_val})

            was_reactivated = (existing["is_active"] or 0) == 0

            def _snapshot(extra=None):
                base = dict(month=month, client_name=client, project_name=proj_name,
                            section=section, portfolio=portfolio, contract_num=contract)
                if extra:
                    base.update(extra)
                return base

            for cf in changed_fields:
                if cf["key"] == "amount_0_100":
                    old_amt = existing["amount_0_100"] or 0.0
                    new_amt = data["amount_0_100"] or 0.0
                    if abs(new_amt) <= 0.01 and old_amt > 0.01:
                        # Сумма обнулилась — всегда показываем
                        ev_type = "zeroed"
                    elif round(old_amt) == round(new_amt):
                        # Погрешность в пределах рубля — событие не создаём,
                        # но значение в БД всё равно обновится (UPDATE выполняется ниже)
                        continue
                    else:
                        ev_type = "amount_changed"
                    pending_events.append(dict(
                        fp_row_id=existing["id"], event_type=ev_type,
                        field_label=None, old_value=None, new_value=None,
                        amount_before=old_amt, amount_after=new_amt,
                        **_snapshot(),
                    ))
                else:
                    pending_events.append(dict(
                        fp_row_id=existing["id"], event_type="field_changed",
                        field_label=cf["field"],
                        old_value=None if cf["old"] is None else str(cf["old"]),
                        new_value=None if cf["new"] is None else str(cf["new"]),
                        amount_before=None, amount_after=None,
                        **_snapshot(),
                    ))

            if was_reactivated:
                pending_events.append(dict(
                    fp_row_id=existing["id"], event_type="reactivated",
                    field_label=None, old_value=None, new_value=None,
                    amount_before=existing["amount_0_100"] or 0.0, amount_after=data["amount_0_100"] or 0.0,
                    **_snapshot(),
                ))

            cur.execute("""
                UPDATE fp_rows SET month=:month, pc=:pc, section=:section, client_name=:client_name,
                    project_num=:project_num, project_name=:project_name, project_manager=:project_manager,
                    contract_num=:contract_num, ez_num=:ez_num, sdz_date=:sdz_date, dpa_date=:dpa_date,
                    fdz_date=:fdz_date, accounting_entity=:accounting_entity, mp_0_100=:mp_0_100,
                    mp_comment=:mp_comment, crm_amount=:crm_amount, amount_0_100=:amount_0_100,
                    note=:note, portfolio=:portfolio, kolodec=:kolodec,
                    strategic_solution=:strategic_solution, is_active=1,
                    quarter_label=:quarter_label, last_seen_at=CURRENT_TIMESTAMP
                WHERE row_key=:row_key
            """, data)

            if changed_fields:
                rows_updated += 1
                diff_changed.append({
                    "client": client, "project": proj_name, "section": section,
                    "changes": changed_fields,
                })
            else:
                rows_unchanged += 1
        else:
            cur.execute("""
                INSERT INTO fp_rows (row_key, month, pc, section, client_name, project_num, project_name,
                    project_manager, contract_num, ez_num, sdz_date, dpa_date, fdz_date, accounting_entity,
                    mp_0_100, mp_comment, crm_amount, amount_0_100, note, portfolio, kolodec,
                    strategic_solution, is_active, quarter_label)
                VALUES (:row_key, :month, :pc, :section, :client_name, :project_num, :project_name,
                    :project_manager, :contract_num, :ez_num, :sdz_date, :dpa_date, :fdz_date,
                    :accounting_entity, :mp_0_100, :mp_comment, :crm_amount, :amount_0_100, :note,
                    :portfolio, :kolodec, :strategic_solution, 1, :quarter_label)
            """, data)
            new_id = cur.lastrowid
            # Автозаполнение комментария из полей файла МП, если пользователь ещё не добавил своё.
            # Объединяем «Комментарий МП к СДЗ» и «Примечание» через разделитель.
            _auto_parts = [p for p in (str(mp_comment or "").strip(), str(note or "").strip()) if p]
            if _auto_parts:
                _auto_comment = " · ".join(_auto_parts)
                cur.execute(
                    "UPDATE fp_rows SET internal_comment = ? WHERE id = ? AND (internal_comment IS NULL OR internal_comment = '')",
                    (_auto_comment, new_id),
                )
            rows_new += 1
            diff_new.append({
                "client": client, "project": proj_name, "section": section,
                "portfolio": portfolio, "amount": amount_0_100,
            })
            pending_events.append(dict(
                fp_row_id=new_id, event_type="new",
                field_label=None, old_value=None, new_value=None,
                amount_before=None, amount_after=data["amount_0_100"] or 0.0,
                month=month, client_name=client, project_num=proj_num, project_name=proj_name,
                section=section, portfolio=portfolio, contract_num=contract,
            ))

            # Переход на новый квартал: если для этой новой строки находится совпадение по
            # (Номер проекта, Раздел ФП) среди строк прошлого квартала, ещё не ставших "Факт" --
            # переносим на неё обязательства, комментарий и ответственного. Ключ из карты сразу
            # удаляется, чтобы при нескольких новых строках с одинаковым (проект, раздел) перенос
            # применился только к первой из них и не задублировался.
            if is_new_quarter and carry_over_map:
                carried = carry_over_map.pop((proj_num, section), None)
                if carried:
                    rollover_summary["matched"] += 1
                    set_clauses = []
                    upd_params = {"id": new_id}
                    if carried["comments"]:
                        set_clauses.append("internal_comment = :internal_comment")
                        upd_params["internal_comment"] = "\n---\n".join(carried["comments"])
                        rollover_summary["comments_copied"] += 1
                    if carried["responsible_user_id"] is not None:
                        set_clauses.append("responsible_user_id = :responsible_user_id")
                        upd_params["responsible_user_id"] = carried["responsible_user_id"]
                    if set_clauses:
                        cur.execute(f"UPDATE fp_rows SET {', '.join(set_clauses)} WHERE id = :id", upd_params)
                    for o in carried["obligations"]:
                        cur.execute("""
                            INSERT INTO obligations (fp_row_id, title, description, responsible_type,
                                responsible_name, due_date, status, created_by, completed_at)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (new_id, o["title"], o["description"], o["responsible_type"],
                              o["responsible_name"], o["due_date"], o["status"], o["created_by"],
                              o["completed_at"]))
                        rollover_summary["obligations_copied"] += 1

            # Повторная загрузка в рамках ТОГО ЖЕ квартала: строка оказалась новой (row_key
            # не совпал, т.к. ключевое поле немного изменилось в файле). Ищем по
            # (project_num, section) в снимке, сделанном до начала импорта, и переносим
            # пользовательские данные — чтобы комментарии/риски/обязательства не терялись.
            elif not is_new_quarter and same_q_user_data and proj_num and section:
                sq_key = (str(proj_num).strip(), str(section).strip())
                sq_carried = same_q_user_data.pop(sq_key, None)
                if sq_carried:
                    sq_clauses = []
                    sq_params = {"id": new_id}
                    if sq_carried["internal_comment"]:
                        sq_clauses.append("internal_comment = :internal_comment")
                        sq_params["internal_comment"] = sq_carried["internal_comment"]
                    if sq_carried["risk_level"] is not None:
                        sq_clauses.append("risk_level = :risk_level")
                        sq_params["risk_level"] = sq_carried["risk_level"]
                    if sq_carried["responsible_user_id"] is not None:
                        sq_clauses.append("responsible_user_id = :responsible_user_id")
                        sq_params["responsible_user_id"] = sq_carried["responsible_user_id"]
                    if sq_clauses:
                        cur.execute(
                            f"UPDATE fp_rows SET {', '.join(sq_clauses)} WHERE id = :id",
                            sq_params,
                        )
                    for o in sq_carried["obligations"]:
                        cur.execute("""
                            INSERT INTO obligations (fp_row_id, title, description, responsible_type,
                                responsible_name, due_date, status, created_by, completed_at)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (new_id, o["title"], o["description"], o["responsible_type"],
                              o["responsible_name"], o["due_date"], o["status"], o["created_by"],
                              o["completed_at"]))

    # Строки, которые были раньше, но не встретились в новой загрузке -> деактивируем (не удаляем,
    # чтобы не потерять привязанные обязательства).
    # ВАЖНО: деактивируем ТОЛЬКО строки того квартала, который сейчас загружается.
    # Строки других кварталов не трогаем — у каждого квартала свой независимый снимок данных.
    cur.execute("""
        SELECT id, row_key, month, client_name, project_num, project_name, section, portfolio,
               contract_num, amount_0_100 FROM fp_rows
        WHERE is_active = 1
          AND (quarter_label = ? OR quarter_label IS NULL OR quarter_label = '')
    """, (upload_ql,))
    all_active = cur.fetchall()
    rows_deactivated = 0
    for row in all_active:
        if row["row_key"] not in seen_keys:
            cur.execute("UPDATE fp_rows SET is_active = 0 WHERE id = ?", (row["id"],))
            rows_deactivated += 1
            diff_deactivated.append({
                "client": row["client_name"], "project": row["project_name"],
                "section": row["section"], "portfolio": row["portfolio"], "amount": row["amount_0_100"],
            })
            pending_events.append(dict(
                fp_row_id=row["id"], event_type="deactivated",
                field_label=None, old_value=None, new_value=None,
                amount_before=row["amount_0_100"] or 0.0, amount_after=0.0,
                month=row["month"], client_name=row["client_name"], project_num=row["project_num"], project_name=row["project_name"],
                section=row["section"], portfolio=row["portfolio"], contract_num=row["contract_num"],
            ))

    # Детектируем переход портфеля 0-100/Возможности → Факт для одной и той же строки.
    # Без этого пользователь видит два несвязанных события: «Пропала» + «Новая».
    FORECAST_PORTFOLIOS = {"0-100", "Возможности"}
    new_by_key = {}
    for ev in pending_events:
        if ev["event_type"] == "new":
            k = (ev["month"], ev["client_name"], ev["project_name"], ev["section"])
            new_by_key[k] = ev

    skip_ids = set()
    portfolio_changed_events = []
    for ev in pending_events:
        if ev["event_type"] == "deactivated" and (ev.get("portfolio") or "") in FORECAST_PORTFOLIOS:
            k = (ev["month"], ev["client_name"], ev["project_name"], ev["section"])
            new_ev = new_by_key.get(k)
            if new_ev and (new_ev.get("portfolio") or "") == "Факт":
                skip_ids.add(id(ev))
                skip_ids.add(id(new_ev))
                del new_by_key[k]  # не допускаем двойного матчинга
                portfolio_changed_events.append(dict(
                    fp_row_id=new_ev["fp_row_id"],
                    event_type="portfolio_changed",
                    field_label="portfolio",
                    old_value=ev["portfolio"],
                    new_value=new_ev["portfolio"],
                    amount_before=ev["amount_before"],
                    amount_after=new_ev["amount_after"],
                    month=ev["month"],
                    client_name=ev["client_name"],
                    project_name=ev["project_name"],
                    section=ev["section"],
                    portfolio=new_ev["portfolio"],
                    contract_num=ev["contract_num"],
                ))
    if skip_ids:
        pending_events = [ev for ev in pending_events if id(ev) not in skip_ids]
        pending_events.extend(portfolio_changed_events)

    # Handle a portfolio transition that also changed month in the same import.
    pending_events = _collapse_portfolio_changes(pending_events)

    # Детектируем смену месяца: та же логика, что portfolio_changed.
    # «Пропала» + «Новая» с одинаковыми client/project/section/portfolio но разными месяцами
    # и близкими суммами → одно событие «Месяц изменился» вместо двух несвязанных событий.
    new_by_identity = {}
    for ev in pending_events:
        if ev["event_type"] == "new":
            k = (ev["client_name"], ev["project_name"], ev["section"], ev.get("portfolio") or "")
            new_by_identity[k] = ev

    month_skip_ids = set()
    month_changed_events = []
    for ev in pending_events:
        if ev["event_type"] == "deactivated":
            k = (ev["client_name"], ev["project_name"], ev["section"], ev.get("portfolio") or "")
            new_ev = new_by_identity.get(k)
            if new_ev and ev.get("month") != new_ev.get("month"):
                # Суммы должны совпадать с допуском 1 рубль
                if abs((ev["amount_before"] or 0) - (new_ev["amount_after"] or 0)) <= 1.0:
                    month_skip_ids.add(id(ev))
                    month_skip_ids.add(id(new_ev))
                    del new_by_identity[k]
                    month_changed_events.append(dict(
                        fp_row_id=new_ev["fp_row_id"],
                        event_type="month_changed",
                        field_label="month",
                        old_value=ev.get("month"),
                        new_value=new_ev.get("month"),
                        amount_before=ev["amount_before"],
                        amount_after=new_ev["amount_after"],
                        month=new_ev.get("month"),
                        client_name=ev["client_name"],
                        project_name=ev["project_name"],
                        section=ev["section"],
                        portfolio=ev.get("portfolio") or new_ev.get("portfolio"),
                        contract_num=ev.get("contract_num"),
                    ))
    if month_skip_ids:
        pending_events = [ev for ev in pending_events if id(ev) not in month_skip_ids]
        pending_events.extend(month_changed_events)

    # Show a same-bank amount transfer as one project change, not new/deactivated.
    pending_events = _collapse_project_moves(pending_events)

    diff_payload = {
        "new": diff_new[:200],
        "changed": diff_changed[:200],
        "deactivated": diff_deactivated[:200],
        "new_truncated": len(diff_new) > 200,
        "changed_truncated": len(diff_changed) > 200,
        "deactivated_truncated": len(diff_deactivated) > 200,
    }

    cur.execute("""
        INSERT INTO import_log (filename, rows_total, rows_new, rows_updated, rows_deactivated, imported_by, diff_json)
        VALUES (?,?,?,?,?,?,?)
    """, (filename, len(seen_keys), rows_new, rows_updated, rows_deactivated, user_id, json.dumps(diff_payload, ensure_ascii=False)))
    log_id = cur.lastrowid

    for ev in pending_events:
        cur.execute("""
            INSERT INTO row_events (fp_row_id, import_log_id, event_type, field_label, old_value, new_value,
                amount_before, amount_after, month, client_name, project_name, section, portfolio, contract_num)
            VALUES (:fp_row_id, :import_log_id, :event_type, :field_label, :old_value, :new_value,
                :amount_before, :amount_after, :month, :client_name, :project_name, :section, :portfolio, :contract_num)
        """, {**ev, "import_log_id": log_id})

    conn.commit()
    conn.close()

    return {
        "rows_total": len(seen_keys),
        "rows_new": rows_new,
        "rows_updated": rows_updated,
        "rows_unchanged": rows_unchanged,
        "rows_deactivated": rows_deactivated,
        "header_mismatches": mismatches,
        "diff": diff_payload,
        "rollover": rollover_summary,
        "quarter_label": upload_ql,
    }
