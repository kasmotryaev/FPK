"""Тесты раздела «Программа максимум»: разбор файла ПМ и сверка с финпланом СЦ."""
import tempfile
import unittest
from pathlib import Path

import openpyxl

import app.db as db
from app import pm_service
from app.importer import EXPECTED_HEADERS, import_excel
from app.pm_parser import PMParseError, fiscal_to_calendar_quarter, parse_workbook

PM_HEADERS = [
    "Клиент", "Договор", "Финдок", "Пресейл", "ПЦ", "Раздел ФП", "Решение", "Страт. решение",
    "Направление", "Доля", "Прогноз + факт по ПЦ (итого)", "Прогноз по ПЦ (итого)",
    "прогноз", "замечание", "автор", "Прогноз 2FQ2026", "Прогноз 3FQ2026",
    "Менеджер УП финдока", "КТ", "Дирекция",
]


def _pm_row(client, section="Проекты", solution="Решение А",
            strategic="Решение А", total=0.0, forecast=0.0, q2=0.0, q3=0.0,
            presale="", manager="Ануфриева А.А."):
    values = {
        "Клиент": client, "Договор": "23 756 Проект", "Пресейл": presale,
        "ПЦ": "ПЦ Финансовые рынки", "Раздел ФП": section,
        "Решение": solution, "Страт. решение": strategic, "Доля": 1,
        "Прогноз + факт по ПЦ (итого)": total, "Прогноз по ПЦ (итого)": forecast,
        "Прогноз 2FQ2026": q2, "Прогноз 3FQ2026": q3,
        "Менеджер УП финдока": manager, "КТ": "", "Дирекция": "ЦРБ-4",
    }
    return [values.get(h) for h in PM_HEADERS]


def _write_pm_file(path, rows, header_row=3):
    """Пишет файл в том же виде, в каком приходит выгрузка: шапка не в первой строке."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TDSheet"
    for _ in range(header_row - 1):
        ws.append(["Распределение по Решениям и Продуктам"])
    ws.append(PM_HEADERS)
    for row in rows:
        ws.append(row)
    wb.save(path)


def _write_fp_file(path, rows):
    """Выгрузка СЦ: (клиент, портфель, сумма)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Данные"
    ws.append(list(EXPECTED_HEADERS))
    for client, portfolio, amount in rows:
        values = {
            "Месяц": "Сентябрь", "Стратегическое решение": "Решение А",
            "ПЦ": "ПЦ Финансовые рынки", "Раздел ФП": "Проекты",
            "Наименование клиента": client, "Номер проекта": "3827137",
            "Наименование проекта": "Проект 1", "Номер договора": "1 к 0404/25",
            "Номер ЭЗ": "3", "Сумма по 0-100, руб": amount, "Портфель": portfolio,
        }
        ws.append([values.get(h) for h in EXPECTED_HEADERS])
    wb.save(path)


class FiscalQuarterTest(unittest.TestCase):
    def test_fiscal_quarters_map_to_calendar(self):
        # финансовый год стартует 1 апреля: ФКВ2 2026 — это календарный июль–сентябрь
        self.assertEqual(fiscal_to_calendar_quarter(1, 2026), "2026-Q2")
        self.assertEqual(fiscal_to_calendar_quarter(2, 2026), "2026-Q3")
        self.assertEqual(fiscal_to_calendar_quarter(3, 2026), "2026-Q4")
        self.assertEqual(fiscal_to_calendar_quarter(4, 2026), "2027-Q1")


class PMParserTest(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.path = Path(self.temp_dir.name) / "pm.xlsx"

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_row_is_split_into_fact_and_quarters(self):
        _write_pm_file(self.path, [
            _pm_row("БАНК А", total=100.0, forecast=90.0, q2=60.0, q3=30.0),
        ])
        parsed = parse_workbook(str(self.path))
        kinds = {(r["kind"], r["quarter_label"]): r["amount"] for r in parsed["rows"]}
        self.assertEqual(kinds[("Факт", "")], 10.0)          # факт = итого − прогноз
        self.assertEqual(kinds[("Прогноз", "2026-Q3")], 60.0)
        self.assertEqual(kinds[("Прогноз", "2026-Q4")], 30.0)

    def test_zero_quarters_do_not_create_rows(self):
        _write_pm_file(self.path, [_pm_row("БАНК", total=50.0, forecast=50.0, q2=50.0, q3=0.0)])
        parsed = parse_workbook(str(self.path))
        self.assertEqual(len(parsed["rows"]), 1)
        self.assertEqual(parsed["rows"][0]["quarter_label"], "2026-Q3")

    def test_strategic_filter_matches_solution_column(self):
        # у части строк стратрешение пустое, а продукт указан в «Решение» — такие тоже наши
        _write_pm_file(self.path, [
            _pm_row("БАНК Б", solution="Продукт Депо", strategic="",
                    total=10.0, forecast=10.0, q2=10.0),
            _pm_row("КОМПАНИЯ В", solution="Продукт 2",
                    strategic="Решение В", total=20.0, forecast=20.0, q2=20.0),
        ])
        parsed = parse_workbook(str(self.path), strategic_filter="Депо")
        self.assertEqual([r["client_name"] for r in parsed["rows"]], ["БАНК Б"])

    def test_broken_file_reports_readable_error(self):
        wb = openpyxl.Workbook()
        wb.active.append(["что-то другое"])
        wb.save(self.path)
        with self.assertRaises(PMParseError):
            parse_workbook(str(self.path))


class PMComparisonTest(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.original_db_path = db.DB_PATH
        db.DB_PATH = Path(self.temp_dir.name) / "pm-test.db"
        db.init_db()

        pm_path = Path(self.temp_dir.name) / "pm.xlsx"
        _write_pm_file(pm_path, [
            # сходится с финпланом
            _pm_row("БАНК Г", total=100.0, forecast=100.0, q2=100.0),
            # расхождение выше порога
            _pm_row("БАНК Б", total=900_000.0, forecast=900_000.0, q2=900_000.0),
            # есть только в ПМ
            _pm_row("БАНК Д", total=500_000.0, forecast=500_000.0, q2=500_000.0),
            # деньги того же клиента под чужим решением
            _pm_row("БАНК Б", solution="Продукт 2",
                    strategic="Решение В",
                    total=300_000.0, forecast=300_000.0, q2=300_000.0),
        ])
        parsed = parse_workbook(str(pm_path))
        self.import_id = pm_service.save_import(parsed, "pm.xlsx", user_id=1)

        fp_path = Path(self.temp_dir.name) / "fp.xlsx"
        _write_fp_file(fp_path, [
            ("БАНК Г", "0-100", 100.0),
            ("БАНК Б", "0-100", 100_000.0),
            ("БАНК Е", "0-100", 700_000.0),      # есть в финплане, в ПМ нет
            ("БАНК Е", "Возможности", 300_000.0),
        ])
        import_excel(str(fp_path), "fp.xlsx", user_id=1, quarter_label="2026-Q3")

    def tearDown(self):
        db.DB_PATH = self.original_db_path
        self.temp_dir.cleanup()

    def _compare(self):
        conn = db.get_conn()
        rows = pm_service.client_comparison(conn, self.import_id, "2026-Q3", strategic="Решение А")
        conn.close()
        return {r["client"]: r for r in rows}

    def test_matching_client_is_ok(self):
        self.assertEqual(self._compare()["БАНК Г"]["status"], "ok")

    def test_difference_above_threshold_is_flagged(self):
        row = self._compare()["БАНК Б"]
        self.assertEqual(row["status"], "diff")
        self.assertEqual(row["delta"], 800_000.0)

    def test_client_only_in_pm(self):
        self.assertEqual(self._compare()["БАНК Д"]["status"], "only_pm")

    def test_client_only_in_fp(self):
        row = self._compare()["БАНК Е"]
        self.assertEqual(row["status"], "only_fp")
        self.assertEqual(row["fp_capability"], 300_000.0)

    def test_other_solutions_shows_money_outside_filter(self):
        conn = db.get_conn()
        # keywords задаём явно: отбор «тематически наших» строк по умолчанию идёт по словам
        # предметной области, а в фикстурах названия решений нейтральные.
        others = pm_service.other_solutions(conn, self.import_id, "2026-Q3", "Решение А",
                                            keywords=("решение в",))
        conn.close()
        self.assertEqual(len(others), 1)
        self.assertEqual(others[0]["client"], "БАНК Б")
        self.assertEqual(others[0]["total"], 300_000.0)

    def test_summary_totals(self):
        conn = db.get_conn()
        summary = pm_service.summary(conn, self.import_id, strategic="Решение А", quarter="2026-Q3")
        conn.close()
        self.assertEqual(summary["forecast_total"], 1_400_100.0)
        self.assertEqual(summary["fact_total"], 0.0)


if __name__ == "__main__":
    unittest.main()
