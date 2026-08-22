# -*- coding: utf-8 -*-
"""Сухая проверка: подойдёт ли .xlsx импортёру ФП-Контроля (без записи в базу)."""
import pathlib
import sys

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent.parent))

import openpyxl
from app.importer import detect_header_mismatch, _build_col_map, REQUIRED_HEADERS

path = sys.argv[1]
ws = openpyxl.load_workbook(path, data_only=True).worksheets[0]
print("лист:", ws.title, "| строк:", ws.max_row, "| колонок:", ws.max_column)
print("расхождения по заголовкам:", detect_header_mismatch(ws) or "нет")
print("распознано колонок:", len(_build_col_map(ws)), "из", len(REQUIRED_HEADERS), "обязательных")
